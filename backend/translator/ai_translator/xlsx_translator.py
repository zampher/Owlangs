# SPDX-FileCopyrightText: 2026 Zamphers
# SPDX-License-Identifier: MPL-2.0
import asyncio
from dataclasses import dataclass
from io import BytesIO
from typing import Self, Literal, List, Optional

import openpyxl
from openpyxl.cell import Cell

from agents.segments_agent import SegmentsTranslateAgentConfig, SegmentsTranslateAgent
from ir.document import Document
from translator.ai_translator.base import AiTranslatorConfig, AiTranslator
from logger.logger import LogModule


@dataclass
class XlsxTranslatorConfig(AiTranslatorConfig):
    insert_mode: Literal["replace", "append", "prepend"] = "replace"
    separator: str = "\n"
    # Specify translation region list.
    # Example: ["Sheet1!A1:B10", "C:D", "E5"]
    # If sheet name is not specified (like "C:D"), it applies to all sheets.
    # If None or empty list, translate all text in the entire file.
    translate_regions: Optional[List[str]] = None


class XlsxTranslator(AiTranslator):
    def __init__(self, config: XlsxTranslatorConfig):
        super().__init__(config=config)
        self.chunk_size = config.chunk_size
        self.translate_agent = None
        if not self.skip_translate:
            agent_config = SegmentsTranslateAgentConfig(
                custom_prompt=config.custom_prompt,
                to_lang=config.to_lang,
                base_url=config.base_url,
                api_key=config.api_key,
                model_id=config.model_id,
                api_type=getattr(config, 'api_type', None) or getattr(config, 'api_protocol', None) or 'openai',
                temperature=config.temperature,
                thinking=config.thinking,
                concurrent=config.concurrent,
                connect_timeout=getattr(config, 'connect_timeout', 15),
                timeout=config.timeout,
                write_timeout=getattr(config, 'write_timeout', None),
                logger=self.logger,
                glossary_dict=config.glossary_dict,
                retry=config.retry,
                max_tokens=getattr(config, 'max_tokens', None)  # Get max_tokens from platform config
            )
            self.translate_agent = SegmentsTranslateAgent(agent_config)
        self.insert_mode = config.insert_mode
        self.separator = config.separator
        # --- New features ---
        self.translate_regions = config.translate_regions

    def _pre_translate(self, document: Document):
        workbook = openpyxl.load_workbook(BytesIO(document.content))
        cells_to_translate = []

        # --- Step 1: Collect text cells that need translation based on whether regions are specified ---

        # If no translation regions are specified, use old logic to translate all cells
        if not self.translate_regions:  # Also handle None or empty list cases
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.data_type == "s":
                            cells_to_translate.append({
                                "sheet_name": sheet.title,
                                "coordinate": cell.coordinate,
                                "original_text": cell.value,
                            })
        # If translation regions are specified, only search within these regions
        else:
            processed_coordinates = set()

            regions_by_sheet = {}
            all_sheet_regions = []
            for region in self.translate_regions:
                if '!' in region:
                    sheet_name, cell_range = region.split('!', 1)
                    if sheet_name not in regions_by_sheet:
                        regions_by_sheet[sheet_name] = []
                    regions_by_sheet[sheet_name].append(cell_range)
                else:
                    all_sheet_regions.append(region)

            for sheet in workbook.worksheets:
                sheet_specific_ranges = regions_by_sheet.get(sheet.title, [])
                total_ranges_for_this_sheet = sheet_specific_ranges + all_sheet_regions

                if not total_ranges_for_this_sheet:
                    continue

                for cell_range in total_ranges_for_this_sheet:
                    try:
                        cells_in_range = sheet[cell_range]

                        # --- START: This is the key part of the modification ---
                        # Flatten to 1D list regardless of whether it returns single cell, 1D tuple (row/column) or 2D tuple (rectangle)
                        flat_cells = []
                        if isinstance(cells_in_range, Cell):
                            flat_cells.append(cells_in_range)
                        elif isinstance(cells_in_range, tuple):
                            for item in cells_in_range:
                                if isinstance(item, Cell):
                                    flat_cells.append(item)  # Handle 1D tuple
                                elif isinstance(item, tuple):
                                    for cell in item:  # Handle 2D tuple
                                        flat_cells.append(cell)
                        # --- END: Modification complete ---

                        # Use simplified single-layer loop
                        for cell in flat_cells:
                            full_coordinate = (sheet.title, cell.coordinate)
                            if full_coordinate in processed_coordinates:
                                continue

                            if isinstance(cell.value, str) and cell.data_type == "s":
                                cell_info = {
                                    "sheet_name": sheet.title,
                                    "coordinate": cell.coordinate,
                                    "original_text": cell.value,
                                }
                                cells_to_translate.append(cell_info)
                                processed_coordinates.add(full_coordinate)

                    except Exception as e:
                        self.logger.warning(LogModule.TRANS, f"Skipping invalid range '{cell_range}' in worksheet '{sheet.title}'. Error: {e}")

        original_texts = [cell["original_text"] for cell in cells_to_translate]
        return workbook, cells_to_translate, original_texts

    def _after_translate(self, workbook, cells_to_translate, translated_texts, original_texts):
        for i, cell_info in enumerate(cells_to_translate):
            sheet_name = cell_info["sheet_name"]
            coordinate = cell_info["coordinate"]
            translated_text = translated_texts[i]
            original_text = original_texts[i]

            # Locate worksheet and cell
            sheet = workbook[sheet_name]
            if self.insert_mode == "replace":
                sheet[coordinate] = translated_text
            elif self.insert_mode == "append":
                sheet[coordinate] = original_text + self.separator + translated_text
            elif self.insert_mode == "prepend":
                sheet[coordinate] = translated_text + self.separator + original_text
            else:
                self.logger.error(LogModule.TRANS, "Invalid XlsxTranslatorConfig parameter")

        workbook_output_stream = BytesIO()
        # Save modified workbook to new file
        try:
            workbook.save(workbook_output_stream)
        finally:
            workbook.close()
        return workbook_output_stream.getvalue()

    def translate(self, document: Document) -> Self:

        # --- Step 1: Try to use cached segments from Extract phase for consistency ---
        task_id = getattr(self, '_task_id', None)
        cached_segments = None
        if task_id:
            try:
                from backend.app.services.task import task_manager
                task_state = task_manager.get_task(task_id) if task_id else None
                if task_state:
                    cache_info = task_state.get("source_chunks_cache", {})
                    cached_segments = cache_info.get("segments", [])
                    if cached_segments:
                        self.logger.info(
                            LogModule.TRANS,
                            f"[XLSX_TRANSLATOR] Found {len(cached_segments)} cached segments from Extract phase. "
                            f"Will use cached segments to ensure consistency with chunk mapping."
                        )
            except Exception as e:
                self.logger.debug(LogModule.TRANS, f"[XLSX_TRANSLATOR] Failed to get cached segments: {e}")

        # --- Step 2: Use cached segments if available, otherwise extract from Excel ---
        if cached_segments and len(cached_segments) > 0:
            # CRITICAL: Always use cached segments from Extract phase to ensure consistency
            # The cached segments are the source of truth, even if _pre_translate extracts differently
            original_texts = [str(s) for s in cached_segments]
            self.logger.info(
                LogModule.TRANS,
                f"[XLSX_TRANSLATOR] Using {len(original_texts)} cached segments from Extract phase "
                f"(skipping re-extraction from Excel file) to ensure consistency with chunk mapping"
            )
            # Rebuild cells_to_translate from segment_info to ensure correct mapping
            workbook = openpyxl.load_workbook(BytesIO(document.content))
            cells_to_translate = []
            try:
                from backend.app.services.task import task_manager
                task_state = task_manager.get_task(task_id) if task_id else None
                if task_state:
                    segments_metadata = task_state.get("segments_metadata", {})
                    segment_info = segments_metadata.get("segment_info", [])
                    if segment_info and len(segment_info) == len(cached_segments):
                        # Rebuild cells_to_translate from segment_info
                        for idx, seg_info in enumerate(segment_info):
                            if isinstance(seg_info, dict) and 'cells' in seg_info:
                                cells = seg_info['cells']
                                if cells and len(cells) > 0:
                                    cell_info = cells[0]  # Use first cell (each segment is one cell)
                                    sheet_name = cell_info.get('sheet', '')
                                    row = cell_info.get('row', 0)
                                    col = cell_info.get('col', 0)
                                    if sheet_name and row > 0 and col > 0:
                                        # Convert row/col to coordinate (e.g., A1)
                                        from openpyxl.utils import get_column_letter
                                        coordinate = f"{get_column_letter(col)}{row}"
                                        cells_to_translate.append({
                                            "sheet_name": sheet_name,
                                            "coordinate": coordinate,
                                            "original_text": original_texts[idx] if idx < len(original_texts) else "",
                                        })
                        self.logger.debug(
                            LogModule.TRANS,
                            f"[XLSX_TRANSLATOR] Rebuilt {len(cells_to_translate)} cells_to_translate from segment_info "
                            f"(matched {len(cached_segments)} cached segments)"
                        )
                    else:
                        self.logger.warning(
                            LogModule.TRANS,
                            f"[XLSX_TRANSLATOR] segment_info not available or count mismatch "
                            f"(segments={len(cached_segments)}, segment_info={len(segment_info) if segment_info else 0}). "
                            f"Will extract cells_to_translate from Excel (may cause mapping issues)."
                        )
                        # Fallback: extract from Excel
                        _, cells_to_translate, _ = self._pre_translate(document)
            except Exception as e:
                self.logger.warning(
                    LogModule.TRANS,
                    f"[XLSX_TRANSLATOR] Failed to rebuild cells_to_translate from segment_info: {e}. "
                    f"Will extract from Excel (may cause mapping issues)."
                )
                # Fallback: extract from Excel
                _, cells_to_translate, _ = self._pre_translate(document)
            
            if not cells_to_translate:
                self.logger.warning(
                    LogModule.TRANS,
                    f"[XLSX_TRANSLATOR] cells_to_translate is empty after rebuild. "
                    f"This may cause issues when writing translated text back to Excel."
                )
        else:
            # Fallback: Extract from Excel if cache not available
            workbook, cells_to_translate, original_texts = self._pre_translate(document)
            if not cells_to_translate:
                print("\nNo plain text content found in specified regions that needs translation.")
                workbook.close()
                return self
        
        if self.glossary_agent:
            self.glossary_dict_gen = self.glossary_agent.send_segments(original_texts, self.chunk_size)
            if self.translate_agent:
                self.translate_agent.update_glossary_dict(self.glossary_dict_gen)
        
        # --- Step 3: Filter excluded segments before translation ---
        excluded_set = self._get_excluded_segments(task_id)
        if excluded_set:
            self.logger.info(LogModule.TRANS, f"[XLSX_TRANSLATOR] Skipping translation for excluded segments: {sorted(excluded_set)}")
        
        translate_indices = [i for i in range(len(original_texts)) if i not in excluded_set]
        texts_for_translation = [original_texts[i] for i in translate_indices]
        
        # --- Step 3: Call translation function ---
        # Use generic chunk translation helper to save segments to cache and translate with chunk merging
        chunk_to_segment_map_for_recording = None
        if self.translate_agent and texts_for_translation:
            try:
                from utils.chunk_translation_helper import translate_segments_with_agent
                from backend.app.services.task import task_manager
                task_state = task_manager.get_task(task_id) if task_id else None
                
                translated_segments, metadata = translate_segments_with_agent(
                    segments=texts_for_translation,
                    chunk_size=self.chunk_size,
                    translate_agent=self.translate_agent,
                    task_id=task_id,
                    task_state=task_state,
                    original_filename=getattr(self, '_original_filename', None),
                    file_contents=document.content,
                )
                chunk_to_segment_map_for_recording = metadata.get("chunk_to_segment_map")
            except Exception as e:
                # Fallback to direct translation if helper fails
                self.logger.warning(LogModule.TRANS, f"[XLSX_TRANSLATOR] Failed to use chunk translation helper: {e}, falling back to direct translation")
                translated_segments = self.translate_agent.send_segments(texts_for_translation, self.chunk_size)
        else:
            translated_segments = texts_for_translation
        
        # Map translated segments back to original indices (including excluded segments)
        final_translated_texts = list(original_texts)
        for idx, segment_idx in enumerate(translate_indices):
            final_translated_texts[segment_idx] = translated_segments[idx] if idx < len(translated_segments) else original_texts[segment_idx]

        # Ensure cells_to_translate matches original_texts length
        if len(cells_to_translate) != len(original_texts):
            self.logger.warning(
                LogModule.TRANS,
                f"[XLSX_TRANSLATOR] cells_to_translate length ({len(cells_to_translate)}) "
                f"does not match original_texts length ({len(original_texts)}). "
                f"Adjusting cells_to_translate to match."
            )
            # Pad or truncate cells_to_translate to match original_texts
            if len(cells_to_translate) < len(original_texts):
                # Pad with empty cell info
                for i in range(len(cells_to_translate), len(original_texts)):
                    cells_to_translate.append({
                        "sheet_name": "Sheet1",  # Default sheet
                        "coordinate": f"A{i+1}",  # Default coordinate
                        "original_text": original_texts[i] if i < len(original_texts) else "",
                    })
            else:
                # Truncate to match
                cells_to_translate = cells_to_translate[:len(original_texts)]

        document.content = self._after_translate(workbook, cells_to_translate, final_translated_texts, original_texts)
        
        # Record translation segments if task_id is provided
        if task_id and len(original_texts) == len(final_translated_texts):
            try:
                from utils.translation_segments import record_translation_segments
                from backend.app.services.task import task_manager
                task_state = task_manager.get_task(task_id)

                if task_state:
                    # Get platform key from task_state (set during task initialization)
                    platform_key = task_state.get("platform_key")
                    self.logger.info(LogModule.TRANS, f"Recording {len(original_texts)} translation segments for task {task_id}")
                    # Get excluded segments using _get_excluded_segments (includes segments_metadata.excluded_segment_indices)
                    excluded_set = self._get_excluded_segments(task_id)
                    excluded_segments = sorted(excluded_set) if excluded_set else None
                    if excluded_segments:
                        self.logger.info(LogModule.TRANS, f"[XLSX_TRANSLATOR] Using {len(excluded_segments)} excluded_segments for recording: {excluded_segments[:10]}...")

                    # Get chunk_to_segment_map from task_state (saved by translate_segments_with_agent)
                    chunk_to_segment_map_from_state = task_state.get("chunk_to_segment_map") if task_state else None
                    # Use metadata's map if available, otherwise use state's map
                    raw_chunk_to_segment_map = chunk_to_segment_map_for_recording or chunk_to_segment_map_from_state
                    
                    # CRITICAL: Map chunk_to_segment_map indices from texts_for_translation to original_texts
                    # chunk_to_segment_map indices are relative to texts_for_translation (filtered segments)
                    # but record_translation_segments expects indices relative to original_texts (all segments)
                    final_chunk_to_segment_map = None
                    if raw_chunk_to_segment_map and translate_indices:
                        final_chunk_to_segment_map = []
                        for chunk_segment_indices in raw_chunk_to_segment_map:
                            # Map each segment index from texts_for_translation to original_texts
                            mapped_indices = [
                                translate_indices[seg_idx] 
                                for seg_idx in chunk_segment_indices 
                                if seg_idx < len(translate_indices)
                            ]
                            if mapped_indices:
                                final_chunk_to_segment_map.append(mapped_indices)
                        self.logger.debug(
                            LogModule.TRANS,
                            f"[XLSX_TRANSLATOR] Mapped chunk_to_segment_map from {len(texts_for_translation)} "
                            f"filtered segments to {len(original_texts)} original segments: "
                            f"{len(raw_chunk_to_segment_map)} chunks -> {len(final_chunk_to_segment_map)} chunks"
                        )
                    else:
                        final_chunk_to_segment_map = raw_chunk_to_segment_map

                    record_translation_segments(
                        task_id=task_id,
                        source_chunks=original_texts,
                        target_chunks=final_translated_texts,
                        original_filename=getattr(self, '_original_filename', None),
                        workflow_type=getattr(self, '_workflow_type', None),
                        source_lang=None,
                        target_lang=self.config.to_lang if hasattr(self.config, 'to_lang') else None,
                        platform_key=platform_key,
                        task_state=task_state,
                        excluded_segments=excluded_segments,
                        chunk_to_segment_map=final_chunk_to_segment_map,
                    )
                    self.logger.info(LogModule.TRANS, f"Successfully recorded translation segments for task {task_id}")
                else:
                    self.logger.warning(LogModule.TRANS, f"Task state not found for task {task_id}, cannot record segments")
            except Exception as e:
                # Log error but don't fail translation
                self.logger.warning(LogModule.TRANS, f"Failed to record translation segments for task {task_id}: {e}", exc_info=True)
        else:
            if not task_id:
                self.logger.debug(LogModule.TRANS, "No task_id provided, skipping segment recording")
            elif len(original_texts) != len(final_translated_texts):
                self.logger.warning(    
                    LogModule.TRANS,
                    f"Source chunks ({len(original_texts)}) and target chunks ({len(final_translated_texts)}) "
                    f"count mismatch, skipping segment recording"
                )
        
        return self
    
    def _get_excluded_segments(self, task_id: str | None) -> set[int]:
        """
        Get excluded segment indices using ExclusionManager (single source of truth).
        This ensures that manually excluded segments from Extract phase are correctly identified.
        """
        if not task_id:
            return set()
        try:
            from backend.app.services.task import task_manager
        except ImportError:
            return set()
        task_state = task_manager.get_task(task_id)
        if not task_state:
            return set()
        
        # CRITICAL: Use ExclusionManager.get_excluded_segments as the single source of truth
        # This ensures that manually excluded segments from Extract phase are correctly identified
        from exclusion.core import ExclusionManager
        excluded_segments_with_reasons = ExclusionManager.get_excluded_segments(task_state)
        excluded_set = set(excluded_segments_with_reasons.keys())
        
        if excluded_set:
                self.logger.info(
                LogModule.TRANS,
                f"[XLSX_TRANSLATOR] Task {task_id}: Retrieved {len(excluded_set)} excluded_segments from ExclusionManager "
                f"(single source of truth). Excluded indices: {sorted(excluded_set)[:20]}{'...' if len(excluded_set) > 20 else ''}"
            )
        else:
            self.logger.debug(
                LogModule.TRANS,
                f"[XLSX_TRANSLATOR] Task {task_id}: No excluded_segments found from ExclusionManager"
            )
        
        return excluded_set

    async def translate_async(self, document: Document, progress_callback=None) -> Self:

        # --- Step 1: Try to use cached segments from Extract phase for consistency ---
        task_id = getattr(self, '_task_id', None)
        cached_segments = None
        if task_id:
            try:
                from backend.app.services.task import task_manager
                task_state = task_manager.get_task(task_id) if task_id else None
                if task_state:
                    cache_info = task_state.get("source_chunks_cache", {})
                    cached_segments = cache_info.get("segments", [])
                    if cached_segments:
                        self.logger.info(
                            LogModule.TRANS,
                            f"[XLSX_TRANSLATOR] Found {len(cached_segments)} cached segments from Extract phase. "
                            f"Will use cached segments to ensure consistency with chunk mapping."
                        )
            except Exception as e:
                self.logger.debug(LogModule.TRANS, f"[XLSX_TRANSLATOR] Failed to get cached segments: {e}")

        # --- Step 2: Use cached segments if available, otherwise extract from Excel ---
        if cached_segments and len(cached_segments) > 0:
            # CRITICAL: Always use cached segments from Extract phase to ensure consistency
            # The cached segments are the source of truth, even if _pre_translate extracts differently
            original_texts = [str(s) for s in cached_segments]
            self.logger.info(
                LogModule.TRANS,
                f"[XLSX_TRANSLATOR] Using {len(original_texts)} cached segments from Extract phase "
                f"(skipping re-extraction from Excel file) to ensure consistency with chunk mapping"
            )
            # Rebuild cells_to_translate from segment_info to ensure correct mapping
            workbook = await asyncio.to_thread(openpyxl.load_workbook, BytesIO(document.content))
            cells_to_translate = []
            try:
                from backend.app.services.task import task_manager
                task_state = task_manager.get_task(task_id) if task_id else None
                if task_state:
                    segments_metadata = task_state.get("segments_metadata", {})
                    segment_info = segments_metadata.get("segment_info", [])
                    if segment_info and len(segment_info) == len(cached_segments):
                        # Rebuild cells_to_translate from segment_info
                        for idx, seg_info in enumerate(segment_info):
                            if isinstance(seg_info, dict) and 'cells' in seg_info:
                                cells = seg_info['cells']
                                if cells and len(cells) > 0:
                                    cell_info = cells[0]  # Use first cell (each segment is one cell)
                                    sheet_name = cell_info.get('sheet', '')
                                    row = cell_info.get('row', 0)
                                    col = cell_info.get('col', 0)
                                    if sheet_name and row > 0 and col > 0:
                                        # Convert row/col to coordinate (e.g., A1)
                                        from openpyxl.utils import get_column_letter
                                        coordinate = f"{get_column_letter(col)}{row}"
                                        cells_to_translate.append({
                                            "sheet_name": sheet_name,
                                            "coordinate": coordinate,
                                            "original_text": original_texts[idx] if idx < len(original_texts) else "",
                                        })
                        self.logger.debug(
                            LogModule.TRANS,
                            f"[XLSX_TRANSLATOR] Rebuilt {len(cells_to_translate)} cells_to_translate from segment_info "
                            f"(matched {len(cached_segments)} cached segments)"
                        )
                    else:
                        self.logger.warning(
                            LogModule.TRANS,
                            f"[XLSX_TRANSLATOR] segment_info not available or count mismatch "
                            f"(segments={len(cached_segments)}, segment_info={len(segment_info) if segment_info else 0}). "
                            f"Will extract cells_to_translate from Excel (may cause mapping issues)."
                        )
                        # Fallback: extract from Excel
                        _, cells_to_translate, _ = await asyncio.to_thread(self._pre_translate, document)
            except Exception as e:
                self.logger.warning(
                    LogModule.TRANS,
                    f"[XLSX_TRANSLATOR] Failed to rebuild cells_to_translate from segment_info: {e}. "
                    f"Will extract from Excel (may cause mapping issues)."
                )
                # Fallback: extract from Excel
                _, cells_to_translate, _ = await asyncio.to_thread(self._pre_translate, document)
            
            if not cells_to_translate:
                self.logger.warning(
                    LogModule.TRANS,
                    f"[XLSX_TRANSLATOR] cells_to_translate is empty after rebuild. "
                    f"This may cause issues when writing translated text back to Excel."
                )
        else:
            # Fallback: Extract from Excel if cache not available
            workbook, cells_to_translate, original_texts = await asyncio.to_thread(self._pre_translate, document)
            if not cells_to_translate:
                print("\nNo plain text content found in specified regions that needs translation.")
                workbook.close()
                return self

        if self.glossary_agent:
            self.glossary_dict_gen = await self.glossary_agent.send_segments_async(original_texts, self.chunk_size)
            if self.translate_agent:
                self.translate_agent.update_glossary_dict(self.glossary_dict_gen)

        # --- Step 3: Filter excluded segments before translation ---
        excluded_set = self._get_excluded_segments(task_id)
        if excluded_set:
            self.logger.info(LogModule.TRANS, f"[XLSX_TRANSLATOR] Skipping translation for excluded segments: {sorted(excluded_set)}")
        
        translate_indices = [i for i in range(len(original_texts)) if i not in excluded_set]
        texts_for_translation = [original_texts[i] for i in translate_indices]
        
        # --- Step 3: Call translation function ---
        # Use generic chunk translation helper to save segments to cache and translate with chunk merging
        chunk_to_segment_map_for_recording = None
        if self.translate_agent and texts_for_translation:
            try:
                from utils.chunk_translation_helper import translate_segments_with_agent_async
                from backend.app.services.task import task_manager
                task_state = task_manager.get_task(task_id) if task_id else None
                
                translated_segments, metadata = await translate_segments_with_agent_async(
                    segments=texts_for_translation,
                    chunk_size=self.chunk_size,
                    translate_agent=self.translate_agent,
                    task_id=task_id,
                    task_state=task_state,
                    original_filename=getattr(self, '_original_filename', None),
                    file_contents=document.content,
                    progress_callback=progress_callback,
                )
                chunk_to_segment_map_for_recording = metadata.get("chunk_to_segment_map")
            except Exception as e:
                # Fallback to direct translation if helper fails
                self.logger.warning(LogModule.TRANS, f"[XLSX_TRANSLATOR] Failed to use chunk translation helper: {e}, falling back to direct translation")
                translated_segments = await self.translate_agent.send_segments_async(texts_for_translation, self.chunk_size, progress_callback)
        else:
            translated_segments = texts_for_translation
        
        # Map translated segments back to original indices (including excluded segments)
        final_translated_texts = list(original_texts)
        for idx, segment_idx in enumerate(translate_indices):
            final_translated_texts[segment_idx] = translated_segments[idx] if idx < len(translated_segments) else original_texts[segment_idx]

        # Ensure cells_to_translate matches original_texts length
        if len(cells_to_translate) != len(original_texts):
            self.logger.warning(
                LogModule.TRANS,
                f"[XLSX_TRANSLATOR] cells_to_translate length ({len(cells_to_translate)}) "
                f"does not match original_texts length ({len(original_texts)}). "
                f"Adjusting cells_to_translate to match."
            )
            # Pad or truncate cells_to_translate to match original_texts
            if len(cells_to_translate) < len(original_texts):
                # Pad with empty cell info
                for i in range(len(cells_to_translate), len(original_texts)):
                    cells_to_translate.append({
                        "sheet_name": "Sheet1",  # Default sheet
                        "coordinate": f"A{i+1}",  # Default coordinate
                        "original_text": original_texts[i] if i < len(original_texts) else "",
                    })
            else:
                # Truncate to match
                cells_to_translate = cells_to_translate[:len(original_texts)]

        document.content = await asyncio.to_thread(self._after_translate, workbook, cells_to_translate,
                                                   final_translated_texts, original_texts)
        
        # Record translation segments if task_id is provided
        if task_id and len(original_texts) == len(final_translated_texts):
            try:
                from utils.translation_segments import record_translation_segments
                from backend.app.services.task import task_manager
                task_state = task_manager.get_task(task_id)

                if task_state:
                    # Get platform key from task_state (set during task initialization)
                    platform_key = task_state.get("platform_key")
                    self.logger.info(LogModule.TRANS, f"Recording {len(original_texts)} translation segments for task {task_id}")
                    # Get excluded segments using _get_excluded_segments (includes segments_metadata.excluded_segment_indices)
                    excluded_set = self._get_excluded_segments(task_id)
                    excluded_segments = sorted(excluded_set) if excluded_set else None
                    if excluded_segments:
                        self.logger.info(LogModule.TRANS, f"[XLSX_TRANSLATOR] Using {len(excluded_segments)} excluded_segments for recording: {excluded_segments[:10]}...")

                    # Get chunk_to_segment_map from task_state (saved by translate_segments_with_agent_async)
                    chunk_to_segment_map_from_state = task_state.get("chunk_to_segment_map") if task_state else None
                    # Use metadata's map if available, otherwise use state's map
                    raw_chunk_to_segment_map = chunk_to_segment_map_for_recording or chunk_to_segment_map_from_state
                    
                    # CRITICAL: Map chunk_to_segment_map indices from texts_for_translation to original_texts
                    # chunk_to_segment_map indices are relative to texts_for_translation (filtered segments)
                    # but record_translation_segments expects indices relative to original_texts (all segments)
                    # Example:
                    # - original_texts = ["text0", "text1", "text2", "text3"] (4 segments)
                    # - excluded_set = {1, 3}
                    # - translate_indices = [0, 2] (indices in original_texts)
                    # - texts_for_translation = ["text0", "text2"] (2 segments)
                    # - chunk_to_segment_map = [[0, 1]] (1 chunk with 2 segments, indices relative to texts_for_translation)
                    # - Need to map to: [[0, 2]] (indices relative to original_texts)
                    final_chunk_to_segment_map = None
                    if raw_chunk_to_segment_map and translate_indices:
                        final_chunk_to_segment_map = []
                        for chunk_segment_indices in raw_chunk_to_segment_map:
                            # Map each segment index from texts_for_translation to original_texts
                            mapped_indices = [
                                translate_indices[seg_idx] 
                                for seg_idx in chunk_segment_indices 
                                if seg_idx < len(translate_indices)
                            ]
                            if mapped_indices:
                                final_chunk_to_segment_map.append(mapped_indices)
                        self.logger.debug(
                            LogModule.TRANS,
                            f"[XLSX_TRANSLATOR] Mapped chunk_to_segment_map from {len(texts_for_translation)} "
                            f"filtered segments to {len(original_texts)} original segments: "
                            f"{len(raw_chunk_to_segment_map)} chunks -> {len(final_chunk_to_segment_map)} chunks"
                        )
                    else:
                        final_chunk_to_segment_map = raw_chunk_to_segment_map

                    record_translation_segments(
                        task_id=task_id,
                        source_chunks=original_texts,
                        target_chunks=final_translated_texts,
                        original_filename=getattr(self, '_original_filename', None),
                        workflow_type=getattr(self, '_workflow_type', None),
                        source_lang=None,
                        target_lang=self.config.to_lang if hasattr(self.config, 'to_lang') else None,
                        platform_key=platform_key,
                        task_state=task_state,
                        excluded_segments=excluded_segments,
                        chunk_to_segment_map=final_chunk_to_segment_map,
                    )
                    self.logger.info(LogModule.TRANS, f"Successfully recorded translation segments for task {task_id}")
                else:
                    self.logger.warning(LogModule.TRANS, f"Task state not found for task {task_id}, cannot record segments")
            except Exception as e:
                # Log error but don't fail translation
                self.logger.warning(LogModule.TRANS, f"Failed to record translation segments for task {task_id}: {e}", exc_info=True)
        else:
            if not task_id:
                self.logger.debug(LogModule.TRANS, "No task_id provided, skipping segment recording")
            elif len(original_texts) != len(final_translated_texts):
                self.logger.warning(
                    LogModule.TRANS,
                    f"Source chunks ({len(original_texts)}) and target chunks ({len(final_translated_texts)}) "
                    f"count mismatch, skipping segment recording"
                )
        
        return self
