# SPDX-FileCopyrightText: 2026 Zampher
# SPDX-License-Identifier: MPL-2.0
"""Tests for XLSX cell type classification and date/value round-tripping."""

from datetime import datetime
from io import BytesIO

import openpyxl
import pytest
from openpyxl.utils.datetime import from_excel, to_excel

from extractor.xlsx_cell_utils import (
    _is_time_only_format,
    apply_xlsx_translated_cell_value,
    classify_excel_numeric_cell,
    excel_serial_to_iso_text,
    format_xlsx_cell_for_extraction,
    nf_has_date_token,
    resolve_date_number_format,
    try_iso_to_excel_serial,
)
from extractor.xlsx_extractor import XlsxExtractor


# ---------------------------------------------------------------------------
# classify_excel_numeric_cell
# ---------------------------------------------------------------------------

def test_general_format_date_serial_detected():
    """Values in the plausible date-serial range (30000-55000) with a
    General / empty number_format are classified as 'date'."""
    assert classify_excel_numeric_cell(46186, "General") == "date"
    assert classify_excel_numeric_cell(100, "General") is None
    assert classify_excel_numeric_cell(0, "General") is None
    assert classify_excel_numeric_cell(29999, "General") is None
    assert classify_excel_numeric_cell(55001, "General") is None


def test_builtin_date_format_detected():
    """Built-in date formats are detected via is_date_flag and/or nf token."""
    assert classify_excel_numeric_cell(45292, "m/d/yy", is_date_flag=True) == "date"
    assert classify_excel_numeric_cell(45292, "yyyy-mm-dd") == "date"


def test_percentage_detected():
    assert classify_excel_numeric_cell(0.75, "0%") == "percentage"
    assert classify_excel_numeric_cell(0.75, "0.00%") == "percentage"


def test_time_only_detected():
    """Pure time fractions (< 1.0) with date-like format are 'time'."""
    assert classify_excel_numeric_cell(0.5, "h:mm", is_date_flag=True) == "time"
    assert classify_excel_numeric_cell(0.5, "h:mm:ss") == "time"


def test_datetime_detected():
    """Serial numbers with non-zero fractional part are 'datetime'."""
    assert classify_excel_numeric_cell(45292.5, "yyyy-mm-dd hh:mm", is_date_flag=True) == "datetime"


def test_bool_not_classified():
    assert classify_excel_numeric_cell(True, "General") is None
    assert classify_excel_numeric_cell(False, "0") is None


def test_non_numeric_not_classified():
    assert classify_excel_numeric_cell("hello", "General") is None
    assert classify_excel_numeric_cell(None, "General") is None


# ---------------------------------------------------------------------------
# nf_has_date_token (fallback for cell.is_date in read_only mode)
# ---------------------------------------------------------------------------

def test_nf_has_date_token_english():
    assert nf_has_date_token("yyyy-mm-dd") is True
    assert nf_has_date_token("m/d/yyyy") is True
    assert nf_has_date_token("dd-mmm-yy") is True
    assert nf_has_date_token("ddd, mmm d") is True
    assert nf_has_date_token("h:mm:ss") is False
    assert nf_has_date_token("0.00") is False


def test_nf_has_date_token_cjk():
    assert nf_has_date_token("yyyy年mm月dd日") is True
    assert nf_has_date_token("yyyy년mm월dd일") is True


def test_nf_has_date_token_russian():
    assert nf_has_date_token("дд.мм.гггг") is True


def test_nf_has_date_token_german():
    assert nf_has_date_token("TT.MM.JJJJ") is True
    assert nf_has_date_token("tt.mm.jjjj") is True


# ---------------------------------------------------------------------------
# _is_time_only_format
# ---------------------------------------------------------------------------

def test_is_time_only():
    assert _is_time_only_format("h:mm") is True
    assert _is_time_only_format("h:mm:ss") is True
    assert _is_time_only_format("h:mm AM/PM") is True
    assert _is_time_only_format("yyyy-mm-dd") is False
    assert _is_time_only_format("m/d/yy h:mm") is False
    assert _is_time_only_format("General") is False


# ---------------------------------------------------------------------------
# excel_serial_to_iso_text
# ---------------------------------------------------------------------------

def test_serial_to_iso_date():
    assert excel_serial_to_iso_text(46186.0) == "2026-06-13"


def test_serial_to_iso_time():
    assert excel_serial_to_iso_text(0.5) == "12:00:00"


def test_serial_to_iso_datetime():
    text = excel_serial_to_iso_text(46186.75, "yyyy-mm-dd hh:mm")
    assert text == "2026-06-13 18:00:00"


# ---------------------------------------------------------------------------
# try_iso_to_excel_serial
# ---------------------------------------------------------------------------

def test_iso_to_serial_date():
    assert try_iso_to_excel_serial("2026-06-13") == pytest.approx(46186.0)


def test_iso_to_serial_datetime():
    serial = try_iso_to_excel_serial("2026-06-13 18:00:00")
    assert serial == pytest.approx(46186.75)


def test_iso_to_serial_time():
    serial = try_iso_to_excel_serial("12:00:00")
    assert serial == pytest.approx(0.5)


def test_iso_to_serial_invalid():
    assert try_iso_to_excel_serial("not a date") is None
    assert try_iso_to_excel_serial("") is None
    assert try_iso_to_excel_serial(None) is None


# ---------------------------------------------------------------------------
# resolve_date_number_format
# ---------------------------------------------------------------------------

def test_resolve_date_number_format_keeps_existing():
    assert resolve_date_number_format("m/d/yy", "date") == "m/d/yy"
    assert resolve_date_number_format("dd/mm/yyyy", "date") == "dd/mm/yyyy"


def test_resolve_date_number_format_fills_general():
    assert resolve_date_number_format("General", "date") == "yyyy-mm-dd"
    assert resolve_date_number_format("", "datetime") == "yyyy-mm-dd hh:mm:ss"
    assert resolve_date_number_format("@", "time") == "hh:mm:ss"


# ---------------------------------------------------------------------------
# format_xlsx_cell_for_extraction  (cell → text + meta)
# ---------------------------------------------------------------------------

def test_format_none_cell():
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws["A1"]
    text, meta = format_xlsx_cell_for_extraction(cell)
    assert text is None
    # None cells produce empty metadata — nothing to preserve
    assert meta == {}
    wb.close()


def test_format_string_cell():
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws["A1"]
    cell.value = "hello"
    cell.data_type = "s"
    text, meta = format_xlsx_cell_for_extraction(cell)
    assert text == "hello"
    assert meta["data_type"] == "s"
    wb.close()


def test_format_date_cell_custom_format():
    """Custom date format (no is_date=True) detected via nf_has_date_token."""
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws["A1"]
    cell.value = 46186
    cell.number_format = "yyyy-mm-dd"
    # Simulate read_only mode: is_date is False for custom formats
    text, meta = format_xlsx_cell_for_extraction(cell)
    assert text == "2026-06-13"
    assert meta["cell_value_kind"] == "date"
    assert meta["excel_serial"] == 46186
    wb.close()


def test_format_datetime_cell():
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws["A1"]
    cell.value = datetime(2026, 6, 13, 18, 30, 0)
    text, meta = format_xlsx_cell_for_extraction(cell)
    assert text == "2026-06-13 18:30:00"
    assert meta["cell_value_kind"] == "datetime"
    wb.close()


def test_format_date_only_datetime_cell():
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws["A1"]
    cell.value = datetime(2026, 6, 13, 0, 0, 0)
    text, meta = format_xlsx_cell_for_extraction(cell)
    assert text == "2026-06-13"
    assert meta["cell_value_kind"] == "date"
    wb.close()


def test_format_percentage_cell():
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws["A1"]
    cell.value = 0.75
    cell.number_format = "0%"
    text, meta = format_xlsx_cell_for_extraction(cell)
    assert text == "75.0%"
    assert meta["cell_value_kind"] == "percentage"
    wb.close()


def test_format_general_format_date_serial():
    """General-format cells in the date-serial range are treated as dates."""
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws["A1"]
    cell.value = 46186
    cell.number_format = "General"
    text, meta = format_xlsx_cell_for_extraction(cell)
    assert text == "2026-06-13"
    assert meta["cell_value_kind"] == "date"
    wb.close()


# ---------------------------------------------------------------------------
# XlsxExtractor integration
# ---------------------------------------------------------------------------

def _make_xlsx_with_cells(cells_spec):
    """Build an in-memory xlsx workbook from a list of cell specs.

    Each spec: (row, col, value, number_format).
    Returns *bytes* of the saved xlsx.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    for row, col, value, nf in cells_spec:
        cell = ws.cell(row=row, column=col)
        cell.value = value
        if nf:
            cell.number_format = nf
    buf = BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def test_extractor_roundtrip_date_cell():
    data = _make_xlsx_with_cells([
        (1, 1, 46186, "yyyy-mm-dd"),
        (1, 2, "plain text", ""),
    ])
    result = XlsxExtractor(data).extract()
    # First segment should be the date
    date_text = result.segments[0]
    assert date_text == "2026-06-13"
    assert result.segment_info[0]["cell_value_kind"] == "date"
    assert result.segment_info[0]["excel_serial"] == 46186
    # Second segment is plain text
    assert result.segments[1] == "plain text"


def test_extractor_roundtrip_datetime_cell():
    data = _make_xlsx_with_cells([
        (1, 1, 46186.75, "yyyy-mm-dd hh:mm"),
    ])
    result = XlsxExtractor(data).extract()
    assert result.segments[0] == "2026-06-13 18:00:00"
    assert result.segment_info[0]["cell_value_kind"] == "datetime"


def test_extractor_roundtrip_time_cell():
    data = _make_xlsx_with_cells([
        (1, 1, 0.5, "h:mm"),
    ])
    result = XlsxExtractor(data).extract()
    assert result.segments[0] == "12:00:00"
    assert result.segment_info[0]["cell_value_kind"] == "time"


def test_extractor_roundtrip_percentage_cell():
    data = _make_xlsx_with_cells([
        (1, 1, 0.75, "0%"),
    ])
    result = XlsxExtractor(data).extract()
    assert result.segments[0] == "75.0%"
    assert result.segment_info[0]["cell_value_kind"] == "percentage"


# ---------------------------------------------------------------------------
# apply_xlsx_translated_cell_value  (write-back preserving types)
# ---------------------------------------------------------------------------

def test_apply_translated_cell_value_date_restores_serial():
    """When the translated text is the same ISO date, the cell keeps its
    numeric serial value and gets a proper date number_format."""
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws["A1"]
    cell.value = 46186
    cell.number_format = "General"
    apply_xlsx_translated_cell_value(
        cell,
        "2026-06-13",
        {
            "number_format": "General",
            "cell_value_kind": "date",
            "excel_serial": 46186,
        },
    )
    assert isinstance(cell.value, (int, float))
    assert int(cell.value) == 46186
    assert cell.number_format == "yyyy-mm-dd"
    assert from_excel(cell.value).strftime("%Y-%m-%d") == "2026-06-13"
    wb.close()


def test_apply_translated_cell_value_date_keeps_unchanged_serial():
    """When AI returns the raw serial number unchanged, it is restored."""
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws["A1"]
    cell.value = 46186
    cell.number_format = "General"
    apply_xlsx_translated_cell_value(
        cell,
        "46186",
        {
            "number_format": "General",
            "cell_value_kind": "date",
            "excel_serial": 46186,
        },
    )
    assert cell.value == 46186
    assert cell.number_format == "yyyy-mm-dd"
    wb.close()


def test_apply_translated_cell_value_preserves_original_date_format():
    """When the original number_format is a meaningful date format, keep it."""
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws["A1"]
    apply_xlsx_translated_cell_value(
        cell,
        "2026-06-13",
        {
            "number_format": "m/d/yy",
            "cell_value_kind": "date",
            "excel_serial": 46186,
        },
    )
    assert cell.number_format == "m/d/yy"
    wb.close()


def test_apply_translated_cell_value_datetime_roundtrip():
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws["A1"]
    cell.value = 46186.75
    cell.number_format = "yyyy-mm-dd hh:mm"
    apply_xlsx_translated_cell_value(
        cell,
        "2026-06-13 18:00:00",
        {
            "number_format": "yyyy-mm-dd hh:mm",
            "cell_value_kind": "datetime",
            "excel_serial": 46186.75,
        },
    )
    assert isinstance(cell.value, float)
    assert cell.value == pytest.approx(46186.75)
    assert cell.number_format == "yyyy-mm-dd hh:mm"
    wb.close()


def test_apply_translated_cell_value_time_roundtrip():
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws["A1"]
    cell.value = 0.5
    cell.number_format = "h:mm:ss"
    apply_xlsx_translated_cell_value(
        cell,
        "12:00:00",
        {
            "number_format": "h:mm:ss",
            "cell_value_kind": "time",
            "excel_serial": 0.5,
        },
    )
    assert isinstance(cell.value, float)
    assert cell.value == pytest.approx(0.5)
    assert cell.number_format == "h:mm:ss"
    wb.close()


def test_apply_translated_cell_value_percentage_roundtrip():
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws["A1"]
    cell.value = "old"
    cell.number_format = "0%"
    apply_xlsx_translated_cell_value(
        cell,
        "75.0%",
        {
            "number_format": "0%",
            "cell_value_kind": "percentage",
        },
    )
    assert cell.value == pytest.approx(0.75)
    assert cell.number_format == "0%"
    wb.close()


def test_apply_translated_cell_value_percentage_no_sign():
    """If the translated text lost the % sign, fall back to string."""
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws["A1"]
    apply_xlsx_translated_cell_value(
        cell,
        "75",
        {
            "number_format": "0%",
            "cell_value_kind": "percentage",
        },
    )
    assert cell.value == "75"
    assert cell.number_format == "0%"
    wb.close()


def test_apply_translated_cell_value_string_fallback():
    """Non-typed cells become plain strings with number_format preserved."""
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws["A1"]
    apply_xlsx_translated_cell_value(
        cell,
        "translated text",
        {
            "number_format": "@",
            "cell_value_kind": None,
        },
    )
    assert cell.value == "translated text"
    # '@' is general-like, so it passes through
    wb.close()


# ---------------------------------------------------------------------------
# Full roundtrip: extract → translate (identity) → write-back
# ---------------------------------------------------------------------------

def test_full_roundtrip_date_identity():
    """Extract a date cell, write the same ISO text back, verify it stays numeric."""
    data = _make_xlsx_with_cells([
        (1, 1, 46186, "General"),
    ])
    extract = XlsxExtractor(data).extract()
    iso_text = extract.segments[0]
    seg_info = extract.segment_info[0]
    assert iso_text == "2026-06-13"

    wb = openpyxl.load_workbook(BytesIO(data))
    cell_ref = seg_info["cells"][0]
    coord = f"{openpyxl.utils.get_column_letter(cell_ref['col'])}{cell_ref['row']}"
    cell = wb[cell_ref["sheet"]][coord]
    apply_xlsx_translated_cell_value(cell, iso_text, seg_info)
    assert isinstance(cell.value, (int, float))
    assert int(cell.value) == 46186
    assert from_excel(cell.value).strftime("%Y-%m-%d") == iso_text
    wb.close()


def test_full_roundtrip_percentage_identity():
    data = _make_xlsx_with_cells([
        (1, 1, 0.75, "0%"),
    ])
    extract = XlsxExtractor(data).extract()
    seg_info = extract.segment_info[0]
    assert extract.segments[0] == "75.0%"

    wb = openpyxl.load_workbook(BytesIO(data))
    cell_ref = seg_info["cells"][0]
    coord = f"{openpyxl.utils.get_column_letter(cell_ref['col'])}{cell_ref['row']}"
    cell = wb[cell_ref["sheet"]][coord]
    apply_xlsx_translated_cell_value(cell, "75.0%", seg_info)
    assert cell.value == pytest.approx(0.75)
    assert cell.number_format == "0%"
    wb.close()
