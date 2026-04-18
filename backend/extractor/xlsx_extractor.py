# SPDX-FileCopyrightText: 2026 Zampher
# SPDX-License-Identifier: MPL-2.0

from __future__ import annotations

from typing import List, Optional, Tuple
from .base import Extractor, ExtractResult


class XlsxExtractor(Extractor):
    """
    Extract text from spreadsheets. Requires openpyxl for .xlsx; for .csv, caller
    should convert to text first or provide a separate path.
    translate_regions: Optional list of regions in A1 notation per sheet like
      ["Sheet1!A1:C10", "Sheet2!B:B"]. If not provided, iterate non-empty cells.
    """

    def __init__(self, file_bytes: bytes, chunk_size: int = 3000, translate_regions: Optional[List[str]] = None):
        """
        Extract text from spreadsheets.
        
        Note: chunk_size is kept for API compatibility but not used in extraction.
        Chunking is handled by chunk_translation_helper during translation.
        """
        self.file_bytes = file_bytes
        self.chunk_size = chunk_size  # Kept for compatibility, not used in extraction
        self.translate_regions = translate_regions or []

    def extract(self) -> ExtractResult:
        try:
            from openpyxl import load_workbook
        except ImportError:
            # Fallback: no extractor available
            return ExtractResult(segments=[])

        from io import BytesIO
        wb = load_workbook(BytesIO(self.file_bytes), data_only=True, read_only=True)

        cells: List[Tuple[str, int, int, str]] = []  # (sheet, row, col, text)

        def add_cell(sheet_name: str, cell):
            try:
                val = cell.value
                if val is None:
                    return
                text = str(val)
                if text.strip():
                    cells.append((sheet_name, cell.row, cell.column, text))
            except Exception:
                return

        if self.translate_regions:
            # very small parsing: Sheet!A1:C10 or Sheet!B:B
            for region in self.translate_regions:
                try:
                    sheet_name, rng = region.split('!')
                    ws = wb[sheet_name]
                    if ':' in rng:
                        # A1:C10
                        for row in ws[rng]:
                            for cell in row:
                                add_cell(sheet_name, cell)
                    else:
                        # Column like B:B or single cell
                        try:
                            for row in ws[rng]:
                                for cell in row:
                                    add_cell(sheet_name, cell)
                        except Exception:
                            add_cell(sheet_name, ws[rng])
                except Exception:
                    continue
        else:
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        add_cell(ws.title, cell)

        # Order by sheet, row, col
        cells.sort(key=lambda t: (t[0], t[1], t[2]))
        # Build segments: each cell is a separate segment (fine-grained extraction)
        # Chunking will be handled by chunk_translation_helper during translation
        segments: List[str] = []
        segment_info: List[dict] = []
        for sheet, row, col, text in cells:
            segments.append(text)
            segment_info.append({'cells': [{'sheet': sheet, 'row': row, 'col': col}]})

        return ExtractResult(segments=segments, segment_info=segment_info)


