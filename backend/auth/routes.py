# SPDX-FileCopyrightText: 2026 Zampher
# SPDX-License-Identifier: MPL-2.0

from fastapi import APIRouter, Request, Response, HTTPException, Form, Depends, UploadFile, File, Body
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from typing import Optional
from pathlib import Path
import time
import os
import json
import ssl
import httpx

from .config import AuthConfig
from .ldap_client import LDAPClient, InvalidCredentials
from .session_manager import AuthSessionManager
from .models import LoginRequest, LoginResponse, LogoutResponse, UserInfo, User, UserRole
from backend.config.config_loader import get_unified_config
import hashlib
from backend.config import get_app_config, save_app_config
from backend.config.secrets_manager import get_secrets_manager
from .local_users import get_local_user_store, LocalUserRole
from logger import unified_logger as logger
from logger.logger import LogModule

def _is_desktop_localhost(request: Request) -> bool:
    """True if request is from localhost with X-Client: desktop (desktop app, no login → treat as admin)."""
    client = getattr(request, "client", None)
    if not client or not isinstance(client, tuple):
        return False
    host = client[0] if len(client) > 0 else None
    if host not in ("127.0.0.1", "::1"):
        return False
    return (request.headers.get("x-client") or "").strip().lower() == "desktop"


# Username masking: keep first and last characters, replace middle with ×
def _mask_username(name: str) -> str:
    try:
        if not name:
            return ""
        if len(name) <= 2:
            return name[0] + ("×" if len(name) == 2 else "")
        return name[0] + ("×" * (len(name) - 2)) + name[-1]
    except Exception:
        return "***"


def _convert_unified_role_to_user_role(unified_role) -> UserRole:
    """Convert unified user role to UserRole enum"""
    from .unified_user_store import UnifiedUserRole
    
    role_mapping = {
        UnifiedUserRole.SUPER_ADMIN: UserRole.ADMIN,
        UnifiedUserRole.ADMIN: UserRole.ADMIN,
        UnifiedUserRole.APP_ADMIN: UserRole.LDAP_APP,
        UnifiedUserRole.USER: UserRole.LDAP_USER,
        UnifiedUserRole.LDAP_ADMIN: UserRole.LDAP_ADMIN,
        UnifiedUserRole.LDAP_APP: UserRole.LDAP_APP,
        UnifiedUserRole.LDAP_USER: UserRole.LDAP_USER,
    }
    
    return role_mapping.get(unified_role, UserRole.LDAP_USER)

# Create router
auth_router = APIRouter(prefix="/auth", tags=["Authentication"])

# Create compatibility router without prefix
auth_compat_router = APIRouter(tags=["Authentication"])

# Template directory: use resource path resolution, compatible with development and PyInstaller
from utils.resource_utils import resource_path
templates = Jinja2Templates(directory=str(resource_path("template")))

# Global variables (should be injected via dependency injection in actual applications)
_auth_config: Optional[AuthConfig] = None
_session_manager: Optional[AuthSessionManager] = None
_ldap_client: Optional[LDAPClient] = None


def init_auth(config: AuthConfig):
    """Initialize authentication module"""
    global _auth_config, _session_manager, _ldap_client
    _auth_config = config
    _session_manager = AuthSessionManager(config)
    if config.ldap_enabled:
        _ldap_client = LDAPClient(config)


def get_auth_config() -> AuthConfig:
    """Get authentication configuration"""
    if _auth_config is None:
        raise HTTPException(status_code=500, detail="Authentication not initialized")
    return _auth_config


def get_session_manager() -> AuthSessionManager:
    """Get session manager"""
    if _session_manager is None:
        raise HTTPException(status_code=500, detail="Session manager not initialized")
    return _session_manager


def get_ldap_client() -> Optional[LDAPClient]:
    """Get LDAP client"""
    return _ldap_client


def _refresh_ldap_client_if_endpoint_changed(old_cfg: "AuthConfig", new_cfg: "AuthConfig") -> None:
    """Safely rebuild LDAP client when LDAP endpoint-related configuration changes."""
    try:
        endpoint_fields = [
            'ldap_enabled', 'ldap_protocol', 'ldap_host', 'ldap_port',
            'ldap_tls_cacertfile', 'ldap_tls_verify'
        ]
        changed = any(getattr(old_cfg, f, None) != getattr(new_cfg, f, None) for f in endpoint_fields)
        if changed:
            global _ldap_client
            if _ldap_client is not None:
                try:
                    _ldap_client.close()
                except Exception:
                    pass
            # Only rebuild when LDAP is enabled
            if new_cfg.ldap_enabled:
                _ldap_client = LDAPClient(new_cfg)
                logger.info(LogModule.AUTH, "[LDAP] Endpoint changed, LDAP client rebuilt")
            else:
                _ldap_client = None
                logger.info(LogModule.AUTH, "[LDAP] LDAP disabled, client released")
    except Exception as e:
        logger.warning(LogModule.AUTH, f"[LDAP] Exception while checking/rebuilding client: {e}")


async def get_current_user(request: Request) -> Optional[User]:
    """Get current user"""
    logger.info(LogModule.AUTH, f"[AUTH] get_current_user called for {request.url.path}")
    # If auth is globally disabled, return a default local user
    try:
        unified_config = get_unified_config()
        auth_required = unified_config.auth_required
        logger.info(LogModule.AUTH, f"[AUTH] auth_required from config: {auth_required}")
        if auth_required is False:
            logger.info(LogModule.AUTH, "[AUTH] Auth disabled, returning default local user")
            return User(
                username="local",
                display_name="Local User",
                email=None,
                is_authenticated=True,
                role=UserRole.ADMIN
            )
    except Exception as e:
        logger.warning(LogModule.AUTH, f"[AUTH] Failed to check auth_required: {e}")
        # Continue with normal auth flow
    logger.info(LogModule.AUTH, "[AUTH] Checking session for user...")
    session_manager = get_session_manager()
    user = await session_manager.get_user(request)
    if user is None:
        # Desktop app from localhost with X-Client: desktop → treat as local admin (no login required)
        if _is_desktop_localhost(request):
            logger.info(LogModule.AUTH, "[AUTH] No session but desktop localhost, returning local admin")
            return User(
                username="local",
                display_name="Local User",
                email=None,
                is_authenticated=True,
                role=UserRole.ADMIN,
            )
        # Return guest user so permission endpoints can return can_access_admin_settings=False
        logger.info(LogModule.AUTH, "[AUTH] No session, returning guest user")
        return User(
            username="guest",
            display_name="Guest",
            email=None,
            is_authenticated=False,
            role=UserRole.LDAP_USER,
        )
    logger.info(LogModule.AUTH, f"[AUTH] Session user: {user.username}")
    return user

@auth_router.get("/config")
async def get_auth_config_endpoint():
    """Expose minimal auth config for frontend routing decisions."""
    # Get latest config (no cache to clear in new config system)
    unified_config = get_unified_config()
    return {"auth_required": unified_config.auth_required}


@auth_router.get("/login", response_class=HTMLResponse)
async def login_page(
    request: Request,
    next_url: Optional[str] = None,
    error: Optional[str] = None
):
    """Login page"""
    from .config import AuthConfig
    config = AuthConfig.get_config()
    return templates.TemplateResponse("login.html", {
        "request": request,
        "next_url": next_url,
        "error": error,
        "ldap_enabled": config.ldap_enabled,
        "login_banner": config.login_banner
    })


@auth_router.post("/login", response_class=JSONResponse)
async def login(
    request: Request,
    response: Response,
    username: str = Form(...),
    password: str = Form(...),
    next_url: Optional[str] = Form(None)
):
    """Handle login request"""
    config = get_auth_config()
    session_manager = get_session_manager()
    
    # Get client IP
    client_ip = request.client.host if request.client else "unknown"
    
    logger.info(LogModule.AUTH, f"Login request received - user: {_mask_username(username)}, IP: {client_ip}")
    logger.info(LogModule.AUTH, f"Auth config - LDAP enabled: {config.ldap_enabled}")
    
    # Check login attempt count
    attempts = session_manager.get_login_attempts(client_ip)
    logger.info(LogModule.AUTH, f"Current login attempts: {attempts}/{config.max_login_attempts}")
    
    if attempts >= config.max_login_attempts:
        logger.warning(LogModule.AUTH, f"IP {client_ip} too many login attempts, locked")
        raise HTTPException(
            status_code=429,
            detail=f"Too many login attempts. Please try again in {config.login_attempt_window // 60} minutes."
        )
    
    try:
        user: User
        
        # Hybrid authentication policy:
        # 1) If username is admin -> always use local auth with ADMIN role
        # 2) If LDAP enabled and username is not admin -> use LDAP auth
        # 3) If LDAP disabled -> only local admin auth is allowed
        
        # Try unified user storage first (for all local users including admin)
        from .unified_user_store import get_unified_user_store
        unified_store = get_unified_user_store()
        unified_user = unified_store.get_user(username)
        
        if unified_user and unified_user.is_active:
            # User exists in unified storage
            logger.info(LogModule.AUTH, f"Using unified user authentication for: {_mask_username(username)}")
            
            if unified_store.verify_credentials(username, password):
                # Update last login time
                unified_store.update_last_login(username)
                
                # Convert unified user to User model
                user = User(
                    username=unified_user.username,
                    display_name=unified_user.display_name or unified_user.username,
                    email=unified_user.email,
                    is_authenticated=True,
                    role=_convert_unified_role_to_user_role(unified_user.role)
                )
                logger.info(LogModule.AUTH, f"Unified user authenticated: {_mask_username(username)}")
            else:
                logger.warning(LogModule.AUTH, f"Unified user authentication failed: {_mask_username(username)}")
                raise InvalidCredentials("Invalid username or password")
        
        # If unified storage is empty and user is default admin, check if we should create default user
        elif not unified_user and username == config.default_username:
            # Check if unified storage is truly empty (no users at all)
            data = unified_store._load()
            if not data.get("users"):
                # First-time setup: create default admin user with the provided password
                logger.info(LogModule.AUTH, f"Unified storage is empty, creating default admin user: {_mask_username(username)}")
                from .unified_user_store import UnifiedUserRole
                if unified_store.create_user(
                    username=username,
                    password=password,
                    role=UnifiedUserRole.SUPER_ADMIN,
                    display_name="Administrator",
                    is_system_user=True,
                    skip_password_validation=True  # Allow any password for first-time setup
                ):
                    # Re-authenticate with the newly created user
                    unified_user = unified_store.get_user(username)
                    if unified_user and unified_store.verify_credentials(username, password):
                        unified_store.update_last_login(username)
                        user = User(
                            username=unified_user.username,
                            display_name=unified_user.display_name or unified_user.username,
                            email=unified_user.email,
                            is_authenticated=True,
                            role=_convert_unified_role_to_user_role(unified_user.role)
                        )
                        logger.info(LogModule.AUTH, f"Default admin user created and authenticated: {_mask_username(username)}")
                    else:
                        logger.warning(LogModule.AUTH, f"Failed to authenticate after creating default admin: {_mask_username(username)}")
                        raise InvalidCredentials("Invalid username or password")
                else:
                    logger.error(LogModule.AUTH, f"Failed to create default admin user: {_mask_username(username)}")
                    raise InvalidCredentials("Failed to initialize admin user")
            else:
                # Unified storage has users but this user doesn't exist
                logger.warning(LogModule.AUTH, f"User not found in unified storage: {_mask_username(username)}")
                raise InvalidCredentials("Invalid username or password")
        
        # No fallback needed - all users should be in unified storage
        elif config.ldap_enabled:
            # Non-admin users use LDAP authentication (ldap3 client)
            logger.info(LogModule.AUTH, f"Using LDAP authentication for user: {_mask_username(username)}")
            try:
                from .ldap_client import LDAPClient
                ldap3_client = LDAPClient(config)
                user = ldap3_client.authenticate(username, password)
            finally:
                try:
                    ldap3_client.close()
                except Exception:
                    pass
            logger.info(LogModule.AUTH, f"LDAP authentication successful, user: {_mask_username(username)}")
        else:
            # LDAP disabled: support local users (except super admin handled above)
            from .local_users import get_local_user_store, LocalUserRole
            logger.info(LogModule.AUTH, "Using local user authentication (LDAP disabled)")
            store = get_local_user_store()
            ok, lu = store.verify_credentials(username, password)
            if not ok or lu is None:
                logger.warning(LogModule.AUTH, f"Local user authentication failed: {_mask_username(username)}")
                raise InvalidCredentials("Invalid username or password")
            # Map local role to system UserRole
            mapped_role = (
                UserRole.ADMIN if lu.role == LocalUserRole.ADMIN else
                UserRole.LDAP_GLOSSARY if lu.role == LocalUserRole.APP_ADMIN else
                UserRole.LDAP_USER
            )
            user = User(
                username=lu.username,
                display_name=lu.display_name or lu.username,
                email=lu.email,
                is_authenticated=True,
                role=mapped_role
            )
            logger.info(LogModule.AUTH, f"Local user authenticated, role mapped: {user.role}")
        
        # Log permission/role info
        try:
            logger.info(
                LogModule.AUTH,
                "User permissions: role=%s, is_admin=%s, is_super_admin=%s, can_access_admin_settings=%s, can_access_glossary_management=%s",
                getattr(user, 'role', None).value if getattr(user, 'role', None) is not None else 'unknown',
                str(user.is_admin() if hasattr(user, 'is_admin') else False),
                str(user.is_super_admin() if hasattr(user, 'is_super_admin') else False),
                str(user.can_access_admin_settings() if hasattr(user, 'can_access_admin_settings') else False),
                str(user.can_access_glossary_management() if hasattr(user, 'can_access_glossary_management') else False)
            )
        except Exception:
            pass

        # Create session
        logger.info(LogModule.AUTH, f"Creating session for user {_mask_username(username)}")
        await session_manager.create_session(request, response, user)
        
        # Get session ID to use as token
        # For API requests, we need to return the session ID directly
        # since cookies might not work properly in Flutter Web
        session_id = session_manager.get_session_id(request)
        if not session_id:
            # If we can't get session ID from cookies, get it from the session we just created
            # This is a workaround for Flutter Web cookie issues
            session_id = session_manager._last_created_session_id
        logger.info(LogModule.AUTH, f"Session created with ID: {session_id}")
        
        # Prepare user information for response
        user_info = {
            "id": user.username,  # Use username as ID
            "username": user.username,
            "email": getattr(user, 'email', ''),
            "full_name": getattr(user, 'display_name', user.username),
            "avatar": None,
            "roles": [user.role.value] if hasattr(user, 'role') else ['user'],
            "permissions": [],  # Will be populated by frontend if needed
            "last_login": None,
            "is_active": True
        }
        
        # Ensure user has a personal profile
        from .user_profile import get_user_profile_manager
        profile_manager = get_user_profile_manager()
        
        # Create default profile if not exists
        if not os.path.exists(f"user_profiles/{username}_profile.json"):
            logger.info(LogModule.AUTH, f"Creating default profile for user {_mask_username(username)}")
            profile_manager.create_default_profile(username)
        else:
            logger.info(LogModule.AUTH, f"User {_mask_username(username)} already has a profile, skipping creation")
        
        # Reset attempts for this IP
        session_manager.reset_login_attempts(client_ip)
        logger.info(LogModule.AUTH, f"Reset login attempts for IP {client_ip}")
        
        # Determine redirect URL
        redirect_url = next_url if next_url and next_url.startswith('/') else "/"
        logger.info(LogModule.AUTH, f"Login successful, redirect URL: {redirect_url}")
        
        return LoginResponse(
            success=True,
            message="Login successful",
            next_url=redirect_url,
            token=session_id,  # Use session ID as token
            user=user_info
        )
        
    except InvalidCredentials as e:
        logger.warning(LogModule.AUTH, f"Authentication failed - invalid credentials: {_mask_username(username)}, error: {e}")
        # Increment login attempts
        session_manager.increment_login_attempts(client_ip)
        logger.info(LogModule.AUTH, f"Incremented login attempts for IP {client_ip}")
        raise HTTPException(status_code=401, detail="Invalid username or password")
    except Exception as e:
        logger.error(LogModule.AUTH, f"Exception during authentication: {_mask_username(username)}, error: {e}")
        logger.error(LogModule.AUTH, f"Exception type: {type(e)}")
        # Increment login attempts
        session_manager.increment_login_attempts(client_ip)
        logger.info(LogModule.AUTH, f"Increased login attempts for IP {client_ip}")
        raise HTTPException(status_code=500, detail=f"Authentication error: {str(e)}")


@auth_router.post("/logout", response_class=JSONResponse)
async def logout(request: Request, response: Response):
    """Handle logout request"""
    session_manager = get_session_manager()
    
    await session_manager.destroy_session(request, response)
    
    return LogoutResponse(
        success=True,
        message="Logout successful"
    )


@auth_router.get("/logout", response_class=RedirectResponse)
async def logout_get(request: Request, response: Response):
    """GET logout, redirect to login page"""
    session_manager = get_session_manager()
    
    await session_manager.destroy_session(request, response)
    
    return RedirectResponse(url="/login", status_code=302)


@auth_router.get("/user", response_model=UserInfo)
async def get_user_info(request: Request):
    """Get current user information"""
    user = await get_current_user(request)
    if not user or not getattr(user, "is_authenticated", True):
        raise HTTPException(status_code=401, detail="Not authenticated")

    return UserInfo(
        username=user.username,
        display_name=user.display_name,
        email=user.email
    )


@auth_router.get("/config")
async def get_auth_config_api(request: Request):
    """Get authentication configuration"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    config = get_auth_config()
    
    # Return configuration but exclude sensitive information like passwords
    return {
        "ldap_enabled": config.ldap_enabled,
        "ldap_protocol": config.ldap_protocol,
        "ldap_host": config.ldap_host,
        "ldap_port": config.ldap_port,
        "ldap_bind_dn_template": config.ldap_bind_dn_template,
        "ldap_base_dn": config.ldap_base_dn,
        "ldap_user_filter": config.ldap_user_filter,
        "ldap_tls_cacertfile": config.ldap_tls_cacertfile,
        "ldap_tls_verify": config.ldap_tls_verify,
        "default_username": config.default_username,
        "session_max_age": config.session_max_age,
        "max_login_attempts": config.max_login_attempts,
        "login_attempt_window": config.login_attempt_window,
    }


@auth_router.post("/config")
async def update_auth_config_api(request: Request, config_data: dict):
    """Update authentication configuration"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
        logger.info(LogModule.AUTH, f"Received configuration update request: {config_data}")
    
    try:
        # Get current configuration
        config = get_auth_config()
        
        # Update configuration
        config.update_from_dict(config_data)
        
        # Save to grouped local.json
        config_file = "local.json"
        if config.save_to_file(config_file):
            logger.info(LogModule.AUTH, "Configuration saved successfully")
            return {"message": "Configuration updated successfully. Please restart the application to take effect."}
        else:
            logger.error(LogModule.AUTH,"Configuration save failed")
            raise HTTPException(status_code=500, detail="Failed to save configuration")
            
    except Exception as e:
        logger.error(LogModule.AUTH, f"Error occurred while updating configuration: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to update configuration: {str(e)}")


@auth_router.post("/test-ldap")
async def test_ldap_connection(request: Request, payload: dict):
    """Test LDAP/LDAPS connection (admin only)
    Input: {"username": "testuser", "password": "***"}
    Execute a simple bind and search using current authentication configuration.
    """
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    if not user.is_admin():
        return JSONResponse(status_code=403, content={"ok": False, "message": "forbidden"})

    username = (payload or {}).get("username", "").strip()
    password = (payload or {}).get("password", "")
    if not username or not password:
        return JSONResponse(status_code=400, content={"ok": False, "message": "username/password required"})

    base_config = get_auth_config()
    # Remove LDAP enabled check - allow testing regardless of current enabled state

    # Allow temporary override with current UI values (not persisted)
    try:
        from dataclasses import asdict
        override = payload or {}
        cfg_dict = asdict(base_config)
        for key in [
            'ldap_protocol', 'ldap_host', 'ldap_port', 'ldap_bind_dn_template', 'ldap_base_dn',
            'ldap_user_filter', 'ldap_admin_group_enabled', 'ldap_glossary_group_enabled',
            'ldap_admin_group', 'ldap_glossary_group', 'ldap_group_base_dn',
            'ldap_tls_cacertfile', 'ldap_tls_verify'
        ]:
            if key in override and override[key] not in (None, ""):
                # Type handling
                if key == 'ldap_port':
                    try:
                        cfg_dict[key] = int(override[key])
                    except Exception:
                        pass
                elif key in ['ldap_tls_verify', 'ldap_admin_group_enabled', 'ldap_glossary_group_enabled']:
                    val = override[key]
                    if isinstance(val, str):
                        cfg_dict[key] = val.lower() in ("true", "1", "yes", "on")
                    else:
                        cfg_dict[key] = bool(val)
                else:
                    cfg_dict[key] = override[key]

        
        # Construct temporary configuration, force enable LDAP for testing
        cfg_dict['ldap_enabled'] = True
        temp_config = AuthConfig(**cfg_dict)

        client = LDAPClient(temp_config)
        user = client.authenticate(username, password)
        
        # Build structured debug information (rendered by frontend i18n)
        groups_enabled = bool(temp_config.ldap_admin_group_enabled or temp_config.ldap_glossary_group_enabled)
        groups_codes = []  # ['admin', 'glossary']
        
        # Check group query status
        if groups_enabled:
            
            # Get user's group membership information (unified use of ldap3, avoid mixing with python-ldap API)
            try:
                from ldap3 import SUBTREE as _LDAP3_SUBTREE
                conn = client._get_connection()
                user_filter = temp_config.ldap_user_filter.format(username=username)
                conn.search(
                    search_base=temp_config.ldap_base_dn,
                    search_filter=user_filter,
                    search_scope=_LDAP3_SUBTREE,
                    attributes=['sAMAccountName', 'displayName', 'mail', 'cn', 'memberOf']
                )

                if conn.entries:
                    user_entry = conn.entries[0]
                    is_admin_member = False
                    is_glossary_member = False

                    # Check admin group
                    if temp_config.ldap_admin_group_enabled:
                        is_admin_member = client._check_admin_group_membership(conn, user_entry)

                    # Check glossary group
                    if temp_config.ldap_glossary_group_enabled:
                        is_glossary_member = client._check_user_group_membership(conn, user_entry)

                    if is_admin_member:
                        groups_codes.append('admin')
                    if is_glossary_member:
                        groups_codes.append('glossary')

            except Exception as e:
                logger.warning(LogModule.AUTH, f"Error occurred while getting group membership information: {e}")
        
        return JSONResponse(content={
            "ok": True,
            "groups_enabled": groups_enabled,
            "groups": groups_codes,
            "user_role": user.role.value,
            "is_admin": user.is_admin(),
            "test_validated": True  # Mark that LDAP test has passed
        })
    except InvalidCredentials:
        return JSONResponse(status_code=401, content={"ok": False, "message": "invalid credentials"})
    except Exception as e:
        return JSONResponse(status_code=502, content={"ok": False, "message": f"{str(e)}"})


@auth_router.get("/user/permissions")
async def get_user_permissions(
    user: User = Depends(get_current_user)
):
    """Get user permission information"""
    return {
        "is_admin": user.is_admin(),
        "is_super_admin": user.is_super_admin(),
        "can_access_admin_settings": user.can_access_admin_settings(),
        "can_access_glossary_management": user.can_access_glossary_management(),
        "allowed_settings": user.get_allowed_settings(),
        "role": user.role.value
    }


@auth_router.get("/app-config")
async def get_app_config_api(
    request: Request,
    user: User = Depends(get_current_user)
):
    """Get application configuration (login required)"""
    logger.info(LogModule.AUTH, f"[CONFIG] ===== App Config Request Started =====")
    logger.info(LogModule.AUTH, f"[CONFIG] Request URL: {request.url}")
    logger.info(LogModule.AUTH, f"[CONFIG] Request method: {request.method}")
    logger.info(LogModule.AUTH, f"[CONFIG] Request headers: {dict(request.headers)}")
    logger.info(LogModule.AUTH, f"[CONFIG] Getting app config for user: {user.username if user else 'None'}")
    
    if not user:
        logger.error(LogModule.AUTH,"[CONFIG] No user found - authentication required")
        from fastapi import HTTPException
        raise HTTPException(status_code=401, detail="Authentication required")
    
    config_dict = {}
    
    # Step 1: Get global configuration (most critical, must succeed)
    try:
        logger.info(LogModule.AUTH, "[CONFIG] Step 1: Loading global configuration...")
        from backend.config.config_loader import get_unified_config, clear_unified_config_cache
        clear_unified_config_cache()
        global_config = get_unified_config()
        logger.info(LogModule.AUTH, f"[CONFIG] Global config object loaded: {type(global_config)}")
        global_config_dict = global_config.get_config_dict(include_api_keys=False, flatten=True)
        logger.info(LogModule.AUTH, f"[CONFIG] Global config dict keys: {list(global_config_dict.keys())[:10]}...")
        if 'ui_language' in global_config_dict:
            logger.info(
                LogModule.AUTH,
                f"[CONFIG] Global config ui_language={global_config_dict.get('ui_language')!r}",
            )
        config_dict.update(global_config_dict)
        ai_platforms_count = len(config_dict.get('ai_platforms', {}))
        logger.info(LogModule.AUTH, f"[CONFIG] Global config loaded successfully, ai_platforms count: {ai_platforms_count}")
        if ai_platforms_count == 0:
            logger.warning(LogModule.AUTH, "[CONFIG] WARNING: ai_platforms is empty!")
    except Exception as e:
        logger.error(LogModule.AUTH, f"[CONFIG] Failed to load global config: {e}", exc_info=True)
        # Return minimal config with error indication
        from fastapi.responses import JSONResponse
        return JSONResponse(
            status_code=500,
            content={
                "error": "Failed to load global configuration",
                "detail": str(e),
                "ai_platforms": {}  # Ensure ai_platforms key exists even on error
            }
        )
    
    # Step 2: Get user personal configuration (optional, can fail gracefully)
    try:
        logger.info(LogModule.AUTH, f"[CONFIG] Step 2: Loading user config for {user.username}...")
        from .user_profile import get_user_profile_manager
        profile_manager = get_user_profile_manager()
        user_profile = profile_manager.get_user_profile(user.username)
        user_config = user_profile.get_config_dict()
        logger.info(LogModule.AUTH, f"[CONFIG] User config dict keys: {list(user_config.keys())[:10]}...")
        if 'ui_language' in user_config:
            logger.info(
                LogModule.AUTH,
                f"[CONFIG] User config ui_language={user_config.get('ui_language')!r} for {user.username}",
            )
        config_dict.update(user_config)
        logger.info(LogModule.AUTH, f"[CONFIG] User config loaded for {user.username}")
    except Exception as e:
        logger.warning(LogModule.AUTH, f"[CONFIG] Failed to load user config for {user.username}: {e}")
        # Continue without user config
    
    # Step 3: Get AppConfig (translator settings including chunk_size) - optional, can fail gracefully
    try:
        logger.info(LogModule.AUTH, "[CONFIG] Step 3: Loading AppConfig (translator settings)...")
        # Always clear AppConfig cache before reading to ensure latest values from disk
        from backend.config import clear_app_config_cache
        clear_app_config_cache()
        app_config = get_app_config()
        app_config_dict = app_config.get_config_dict()
        logger.info(LogModule.AUTH, f"[CONFIG] AppConfig dict keys: {list(app_config_dict.keys())[:10]}...")
        
        # Map translator_chunk_token_size to chunk_size for frontend compatibility
        if 'translator_chunk_token_size' in app_config_dict:
            chunk_size_value = app_config_dict.get('translator_chunk_token_size')
            if chunk_size_value and chunk_size_value != 0:
                config_dict['chunk_size'] = chunk_size_value
                logger.info(LogModule.AUTH, f"[CONFIG] Mapped translator_chunk_token_size={chunk_size_value} to chunk_size")
        
        # Map other translator settings for frontend compatibility
        translator_mappings = {
            'translator_concurrent': 'concurrent',
            'translator_connect_timeout': 'connect_timeout',
            'translator_timeout': 'timeout',
            'translator_retry': 'retry',
            'translator_temperature': 'temperature',
            'translator_thinking_mode': 'thinking',
        }
        for backend_key, frontend_key in translator_mappings.items():
            if backend_key in app_config_dict:
                config_dict[frontend_key] = app_config_dict[backend_key]
        
        # Also include all AppConfig fields (for backward compatibility),
        # but never override user-level settings such as ui_language that were
        # already merged from UserProfile in Step 2.
        app_config_user_level_keys = {
            "ui_language",
            "translator_last_workflow",
            "translator_auto_workflow_enabled",
        }
        for key in app_config_user_level_keys:
            if key in app_config_dict and key in config_dict:
                logger.info(
                    LogModule.AUTH,
                    f"[CONFIG] Skipping AppConfig override for user-level key "
                    f"{key!r}: keeping value from user config {config_dict.get(key)!r}, "
                    f"ignoring global value {app_config_dict.get(key)!r}",
                )
                app_config_dict.pop(key, None)

        config_dict.update(app_config_dict)
        logger.info(LogModule.AUTH, "[CONFIG] AppConfig loaded successfully")
    except Exception as e:
        logger.warning(LogModule.AUTH, f"[CONFIG] Failed to load AppConfig: {e}")
        # Continue without AppConfig
    
    # Map backend snake_case keys to frontend camelCase for parsing engine settings
    # so the frontend can correctly read the user's chosen convert engine.
    if 'translator_convert_engine' in config_dict and config_dict['translator_convert_engine']:
        config_dict['parsingEngine'] = config_dict['translator_convert_engine']
    elif 'parsing_engine' in config_dict and isinstance(config_dict['parsing_engine'], dict):
        config_dict['parsingEngine'] = config_dict['parsing_engine'].get('convert_engine', 'mineru')
    if 'translator_formula_ocr' in config_dict:
        config_dict['formulaOcr'] = config_dict['translator_formula_ocr']
    if 'translator_table_ocr' in config_dict:
        config_dict['tableOcr'] = config_dict['translator_table_ocr']
    
    # Step 4: Get local configuration (optional, can fail gracefully)
    try:
        from backend.config.local_config import LocalConfig
        local_config = LocalConfig.load_from_file()
        local_config_dict = local_config.get_config_dict()
        config_dict.update(local_config_dict)
        logger.info(LogModule.AUTH, "Local config loaded successfully")
    except Exception as e:
        logger.warning(LogModule.AUTH, f"Failed to load local config: {e}")
        # Continue without local config
    
    # Step 5: Get LDAP configuration (optional, can fail gracefully)
    try:
        auth_config = get_auth_config()
        auth_config_dict = auth_config.__dict__
        config_dict.update(auth_config_dict)
        logger.info(LogModule.AUTH, "Auth config loaded successfully")
    except Exception as e:
        logger.warning(LogModule.AUTH, f"Failed to load auth config: {e}")
        # Continue without auth config
    
    # Ensure ai_platforms exists (critical for frontend)
    if 'ai_platforms' not in config_dict:
        logger.warning(LogModule.AUTH, "[CONFIG] ai_platforms not found in config_dict, setting empty dict")
        config_dict['ai_platforms'] = {}
    else:
        logger.info(LogModule.AUTH, f"[CONFIG] Final ai_platforms count: {len(config_dict.get('ai_platforms', {}))}")
    
    # Get default_platform from platforms.json (new config structure only)
    # Frontend expects default_platform at top level
    try:
        logger.info(LogModule.AUTH, "[CONFIG] Loading default_platform from platforms.json...")
        from backend.config.platforms_config import get_platforms_config
        platforms_config = get_platforms_config()
        config_dict['default_platform'] = platforms_config.default_platform
        logger.info(LogModule.AUTH, f"[CONFIG] Default platform: {config_dict['default_platform']}")
    except Exception as e:
        logger.warning(LogModule.AUTH, f"[CONFIG] Failed to load default_platform from platforms.json: {e}")
        config_dict['default_platform'] = 'deepseek'  # Default fallback
        logger.info(LogModule.AUTH, f"[CONFIG] Using fallback default_platform: {config_dict['default_platform']}")
    
    # Filter sensitive configuration based on user permissions
    if not user.is_admin():
        # Non-admin users, only return basic configuration
        filtered_config = {}
        # Allowed basic settings
        allowed_keys = [
            'ui_language', 'translator_last_workflow', 'translator_auto_workflow_enabled',
            'translator_txt_insert_mode', 'translator_txt_separator',
            'translator_xlsx_insert_mode', 'translator_xlsx_separator', 'translator_xlsx_translate_regions',
            'translator_docx_insert_mode', 'translator_docx_separator',
            'translator_srt_insert_mode', 'translator_srt_separator',
            'translator_epub_insert_mode', 'translator_epub_separator',
            'translator_html_insert_mode', 'translator_html_separator',
            'translator_json_paths', 'translator_target_language', 'translator_custom_language',
            'translator_custom_prompt', 'translator_thinking_mode', 'theme',
            'translator_platform_type', 'translator_temperature', 'translator_max_tokens', 'translator_top_p',
            'translator_frequency_penalty', 'translator_presence_penalty',
            'translator_chunk_token_size', 'chunk_size', 'concurrent', 'timeout', 'retry', 'temperature', 'thinking',
            'glossary_generate_enable', 'glossary_agent_config_choice', 'glossary_agent_thinking_mode',
            'glossary_agent_platform_type', 'glossary_agent_temperature', 'glossary_agent_max_tokens', 'glossary_agent_top_p',
            'glossary_agent_frequency_penalty', 'glossary_agent_presence_penalty', 'glossary_agent_to_lang',
            'glossary_agent_chunk_size', 'glossary_agent_concurrent',
            # Non-sensitive settings in global configuration
            'ai_platforms', 'parsing_engine', 'parsingEngine', 'formulaOcr', 'tableOcr', 'translator_settings', 'default_language', 'default_platform',
            # User dimension model override
            'translator_platform_models', 'glossary_agent_platform_models',
            # LDAP configuration (non-sensitive part)
            'ldap_enabled', 'ldap_protocol', 'ldap_host', 'ldap_port'
        ]
        for key in allowed_keys:
            if key in config_dict:
                filtered_config[key] = config_dict[key]
        logger.info(LogModule.AUTH, f"[CONFIG] Filtered config keys: {list(filtered_config.keys())[:20]}...")
        logger.info(LogModule.AUTH, f"[CONFIG] Filtered ai_platforms count: {len(filtered_config.get('ai_platforms', {}))}")
        logger.info(LogModule.AUTH, f"[CONFIG] ===== App Config Response (Non-Admin) =====")
        return filtered_config
    else:
        # Admin users, return all configuration but hide sensitive information
        # Mask API keys (from ai_platforms)
        if 'ai_platforms' in config_dict:
            for platform_key, platform_data in config_dict['ai_platforms'].items():
                if isinstance(platform_data, dict) and 'api_key' in platform_data:
                    api_key = platform_data['api_key']
                    if api_key:
                        platform_data['api_key'] = api_key[:8] + "***" if len(api_key) > 8 else "***"
                    else:
                        platform_data['api_key'] = ""
        
        
        # Mask Mineru Token (loaded from sensitive configuration)
        from backend.config.secrets_manager import get_secrets_manager
        secrets_manager = get_secrets_manager()
        mineru_token = secrets_manager.get_mineru_token()
        if mineru_token:
            config_dict['translator_mineru_token'] = mineru_token[:8] + "***" if len(mineru_token) > 8 else "***"
        else:
            config_dict['translator_mineru_token'] = ""
        
        logger.info(LogModule.AUTH, f"[CONFIG] Final config keys: {list(config_dict.keys())[:20]}...")
        logger.info(LogModule.AUTH, f"[CONFIG] Final ai_platforms count: {len(config_dict.get('ai_platforms', {}))}")
        logger.info(LogModule.AUTH, f"[CONFIG] ===== App Config Response (Admin) =====")
        return config_dict


@auth_router.get("/app-config/raw-secrets")
async def get_raw_secrets_api(
    user: User = Depends(get_current_user)
):
    """Get complete sensitive configuration (admin only)"""
    if not user.is_admin():
        raise HTTPException(status_code=403, detail="Access denied")
    
    from backend.config.secrets_manager import get_secrets_manager
    secrets_manager = get_secrets_manager()
    
    # Get complete API keys and metadata (not masked)
    api_keys_meta = secrets_manager.get_api_keys_meta()
    mineru_meta = secrets_manager.get_mineru_token_meta()
    mineru_local_meta = secrets_manager.get_mineru_local_token_meta()
    # Maintain backward compatibility: provide old fields as well
    api_keys_plain = {k: v.get("key", "") for k, v in api_keys_meta.items()}
    return {
        "platform_api_keys": api_keys_plain,
        "platform_api_keys_meta": api_keys_meta,
        "translator_mineru_token": mineru_meta.get("key", ""),
        "translator_mineru_token_meta": mineru_meta,
        "mineru_local_token": mineru_local_meta.get("key", ""),
        "mineru_local_token_meta": mineru_local_meta
    }

def _get_donor_license_public_key_path():
    """Path to Ed25519 public key for verifying registration codes."""
    from backend.utils.path_utils import get_config_file_path
    return get_config_file_path("donor_license_public.pem")


def _get_donor_edition() -> str:
    """
    Return current donor edition for license validation: PRO (desktop) or PRO-WEB (web).
    Set env DONOR_EDITION=PRO or DONOR_EDITION=PRO-WEB; default PRO for desktop builds.
    """
    edition = (os.environ.get("DONOR_EDITION") or "PRO").strip().upper()
    if edition == "PRO-WEB":
        return "PRO-WEB"
    return "PRO"


def _get_donor_license_private_key_path():
    """Path to Ed25519 private key (for signing year-only code 1037+year). Only used when user submits 1037."""
    from backend.utils.path_utils import get_config_file_path
    return get_config_file_path("donor_license_private.pem")


def _verify_donor_license_token(token: str) -> tuple[bool, Optional[str]]:
    """Re-verify stored license token; return (True, None) if valid, (False, reason) if invalid (e.g. 'Code has expired')."""
    # Legacy simple activation codes (other than 1037) remain always valid
    if token and token.startswith("SIMPLE_"):
        return True, None

    from backend.utils.machine_id import get_machine_id
    from backend.utils.donor_license import load_public_key_from_file, verify_registration_code
    key_path = _get_donor_license_public_key_path()
    if not key_path.exists():
        return False, None
    try:
        pub_pem = load_public_key_from_file(key_path)
        current_mid = get_machine_id()
        current_edition = _get_donor_edition()
        ok, err_msg = verify_registration_code(
            pub_pem, token, current_mid, current_edition=current_edition
        )
        return ok, err_msg
    except Exception:
        return False, None


def _check_can_create_translation_task() -> tuple[bool, Optional[str]]:
    """
    Whether this deployment may create new translation tasks (legacy helper).

    Current design does not hard-block translation tasks based on trial or license;
    this helper always returns (True, None) and is kept for potential future use.
    """
    return True, None


@auth_router.get("/donor/status")
async def get_donor_status(
    user: User = Depends(get_current_user)
):
    """Get donor activation status, machine_id, trial info, and current license info (edition, expiry) for display."""
    from backend.config.secrets_manager import get_secrets_manager
    from backend.utils.donor_license import decode_license_payload
    from backend.utils.donor_trial import is_effective_activated
    from backend.utils.machine_id import get_machine_id
    secrets_manager = get_secrets_manager()
    donor_status = secrets_manager.get_donor_activation()
    activated = donor_status["activated"]
    license_token = donor_status.get("license_token")
    trial_start_date = donor_status.get("trial_start_date")
    expired = False
    license_edition = None
    license_expiry = None
    # Re-verify stored token so that cloning config to another machine or expiry invalidates activation
    if activated and license_token:
        ok, reason = _verify_donor_license_token(license_token)
        if not ok:
            if reason == "Code has expired":
                expired = True
            secrets_manager.update_donor_activation(activated=False, license_token=None)
            activated = False
        # Decode payload for UI: show edition and expiry (no raw token exposed)
        if license_token and not license_token.startswith("SIMPLE_"):
            payload, _ = decode_license_payload(license_token)
            if payload:
                license_edition = (payload.get("license_key") or "").strip() or None
                license_expiry = (payload.get("expiry") or "").strip() or None
    effective_activated, trial_ends_at, trial_expired = is_effective_activated(
        activated, trial_start_date
    )
    edition = _get_donor_edition()
    # Do not hard-block translation tasks even when trial expired.
    # This flag is kept only for UI hints and currently always true.
    can_create_translation_task = True
    machine_id = get_machine_id()
    return {
        "activated": activated,
        "expired": expired,
        "machine_id": machine_id,
        "license_token_masked": "****" if license_token else None,
        "license_edition": license_edition,
        "license_expiry": license_expiry,
        "trial_start_date": trial_start_date,
        "trial_ends_at": trial_ends_at,
        "trial_expired": trial_expired,
        "effective_activated": effective_activated,
        "can_create_translation_task": can_create_translation_task,
        "deployment_edition": edition,
    }

@auth_router.post("/donor/activate")
async def activate_donor(
    request: Request,
    user: User = Depends(get_current_user)
):
    """Activate donor status with machine-bound registration code (no plaintext codes).
    
    Supports two types of codes:
    1. Machine-bound registration codes (signed with Ed25519, machine-specific)
    2. Simple activation codes (legacy whitelist, for friends/testing)
    """
    try:
        data = await request.json()
        registration_code = (data.get("registration_code") or data.get("activation_code", "")).strip()
        if not registration_code:
            raise HTTPException(status_code=400, detail="Registration code is required")
        
        from backend.config.secrets_manager import get_secrets_manager
        from backend.utils.machine_id import get_machine_id
        secrets_manager = get_secrets_manager()
        current_machine_id = get_machine_id()

        # Log activation attempt (for debugging license issues)
        logger.info(
            LogModule.AUTH,
            "[LICENSE] Activation attempt: machine_id={machine_id}, code_prefix={code_prefix}",
            machine_id=current_machine_id,
            code_prefix=registration_code[:16] if registration_code else "",
        )
        
        # Legacy simple exception codes (no machine binding, always valid). Stored as SIMPLE_<code>.
        SIMPLE_EXCEPTION_CODES = {"SKY2000"}
        if registration_code in SIMPLE_EXCEPTION_CODES:
            simple_token = f"SIMPLE_{registration_code}"
            if secrets_manager.update_donor_activation(activated=True, license_token=simple_token):
                logger.info(
                    LogModule.AUTH,
                    f"Donor activated with simple exception code {registration_code} by user {_mask_username(user.username)}",
                )
                return {"success": True, "message": "Donor benefits activated successfully"}
            raise HTTPException(status_code=500, detail="Failed to save donor activation")
        
        # Special code 1037: append current year (e.g. 10372026), sign with private key, expire at end of year.
        # Machine-bound codes (from keygen) are not modified and remain permanent.
        from datetime import datetime
        YEAR_ONLY_SPECIAL_CODE = "1037"
        if registration_code == YEAR_ONLY_SPECIAL_CODE:
            year = datetime.now().year
            code_with_year = f"{YEAR_ONLY_SPECIAL_CODE}{year}"
            expiry_iso = f"{year}-12-31"
            priv_path = _get_donor_license_private_key_path()
            if not priv_path.exists():
                logger.warning(LogModule.AUTH, "Donor license private key not found, cannot sign year-only code 1037")
                raise HTTPException(
                    status_code=503,
                    detail="Activation for this code is not configured. Please contact the administrator.",
                )
            from backend.utils.donor_license import load_private_key_from_file, sign_year_only_code
            priv_pem = load_private_key_from_file(priv_path)
            signed_token = sign_year_only_code(priv_pem, code_with_year, expiry_iso)
            if secrets_manager.update_donor_activation(activated=True, license_token=signed_token):
                logger.info(
                    LogModule.AUTH,
                    f"Donor activated with year-only code {code_with_year} (expires {expiry_iso}) by user {_mask_username(user.username)}",
                )
                return {"success": True, "message": "Donor benefits activated successfully"}
            raise HTTPException(status_code=500, detail="Failed to save donor activation")

        # Machine-bound registration code verification
        from backend.utils.donor_license import load_public_key_from_file, verify_registration_code
        key_path = _get_donor_license_public_key_path()
        if not key_path.exists():
            logger.warning(LogModule.AUTH, "Donor license public key not found, cannot verify registration codes")
            raise HTTPException(status_code=503, detail="Activation is not configured. Please contact the administrator.")
        pub_pem = load_public_key_from_file(key_path)
        current_edition = _get_donor_edition()
        ok, err_msg = verify_registration_code(
            pub_pem, registration_code, current_machine_id, current_edition=current_edition
        )
        if not ok:
            # Log detailed failure reason and decoded payload (best-effort) to help diagnose issues
            try:
                from backend.utils.donor_license import decode_license_payload

                payload, _ = decode_license_payload(registration_code)
            except Exception:
                payload = None
            logger.error(
                LogModule.AUTH,
                "[LICENSE] Activation failed: reason={reason}, machine_id={machine_id}, "
                "code_prefix={code_prefix}, payload={payload}",
                reason=err_msg or "unknown",
                machine_id=current_machine_id,
                code_prefix=registration_code[:24] if registration_code else "",
                payload=payload,
            )
            raise HTTPException(status_code=400, detail=err_msg or "Invalid registration code. Please check and try again.")
        if secrets_manager.update_donor_activation(activated=True, license_token=registration_code):
            logger.info(LogModule.AUTH, f"Donor activated by user {_mask_username(user.username)} (machine_id={current_machine_id[:6]}...)")
            return {"success": True, "message": "Donor benefits activated successfully"}
        raise HTTPException(status_code=500, detail="Failed to save donor activation")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(LogModule.AUTH, f"Failed to activate donor: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to activate donor: {str(e)}")

@auth_router.post("/web/upload-cert")
async def upload_web_cert(
    cert: UploadFile | None = File(None),
    key: UploadFile | None = File(None),
    user: User = Depends(get_current_user)
):
    if not user.is_admin():
        raise HTTPException(status_code=403, detail="Access denied")

    try:
        from pathlib import Path
        from backend.config.config_loader import get_unified_config, save_unified_config
        base_dir = Path(__file__).resolve().parents[2]
        certs_dir = base_dir / "certs"
        certs_dir.mkdir(parents=True, exist_ok=True)

        saved_cert_path = None
        saved_key_path = None

        if cert is not None and cert.filename:
            target = certs_dir / cert.filename
            content = await cert.read()
            target.write_bytes(content)
            saved_cert_path = str(target)

        if key is not None and key.filename:
            target = certs_dir / key.filename
            content = await key.read()
            target.write_bytes(content)
            saved_key_path = str(target)

        if not saved_cert_path and not saved_key_path:
            raise HTTPException(status_code=400, detail="No files uploaded")

        from backend.config.local_config import LocalConfig
        local_config = LocalConfig.load_from_file()
        if saved_cert_path:
            local_config.https.cert_file = saved_cert_path
        if saved_key_path:
            local_config.https.key_file = saved_key_path
        local_config.save_to_file()

        return {"success": True, "cert": saved_cert_path, "key": saved_key_path}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(LogModule.AUTH, f"Certificate upload failed: {e}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@auth_router.post("/web/test-https")
async def test_https_available(
    request: Request,
    payload: dict,
    user: User = Depends(get_current_user)
):
    """Test current certificate and HTTPS availability (admin only)
    Logic:
    1) Read passed certificate/private key paths (use global config if not passed)
    2) Verify certificate/private key files exist and are readable
    3) Try to load into SSLContext (equivalent to Uvicorn usage)
    4) Make one HTTPS request to self (verify=False), return status code
    """
    if not user.is_admin():
        raise HTTPException(status_code=403, detail="Access denied")

    try:
        from backend.config.config_loader import get_unified_config
        from backend.config.local_config import LocalConfig
        from backend.config.secrets_manager import get_secrets_manager

        gc = get_unified_config()
        local_config = LocalConfig.load_from_file()
        sm = get_secrets_manager()
        cert_file = local_config.https.cert_file
        key_file = local_config.https.key_file
        key_password = (payload or {}).get('https_key_password') or sm.get_web_tls_password()

        details = {
            "cert_exists": bool(cert_file and os.path.exists(cert_file)),
            "key_exists": bool(key_file and os.path.exists(key_file)),
        }

        # Check if openssl is available (for auto-generation or user troubleshooting)
        try:
            import shutil
            details["openssl_available"] = bool(shutil.which("openssl"))
        except Exception:
            details["openssl_available"] = False

        if not details["cert_exists"] or not details["key_exists"]:
            return JSONResponse(status_code=400, content={
                "ok": False,
                "message": "Certificate or key file not found" + ("; please install openssl to auto-generate dev cert" if not details.get("openssl_available") else ""),
                **details
            })

        # 3) Load into SSLContext
        try:
            ctx = ssl.create_default_context(purpose=ssl.Purpose.CLIENT_AUTH)
            # Allow no password
            ctx.load_cert_chain(certfile=cert_file, keyfile=key_file, password=key_password)
            details["load_sslcontext_ok"] = True
        except Exception as e:
            details["load_sslcontext_ok"] = False
            details["load_error"] = str(e)
            return JSONResponse(status_code=400, content={
                "ok": False,
                "message": "Failed to load cert/key into SSL context",
                **details
            })

        # 4) Self-test: make one HTTPS request to self (disable verification to support self-signed)
        port = getattr(request.app.state, 'port_to_use', 8800)
        try:
            async with httpx.AsyncClient(verify=False, timeout=2.5) as client:
                r = await client.get(f"https://127.0.0.1:{port}/login")
                details["probe_status"] = r.status_code
        except Exception as e:
            details["probe_status"] = None
            details["probe_error"] = str(e)

        return {"ok": True, "message": "HTTPS test completed", **details}
    except Exception as e:
        logger.error(LogModule.AUTH, f"HTTPS test failed: {e}")
        return JSONResponse(status_code=500, content={"ok": False, "message": str(e)})


# === Glossary Management API ===

@auth_router.get("/glossaries")
async def get_glossaries_list(
    user: User = Depends(get_current_user)
):
    """Get glossary list"""
    from glossary.manager import get_glossary_manager
    
    manager = get_glossary_manager()
    
    # Get global glossaries
    global_glossaries = manager.get_global_glossaries()
    
    # Get user personal glossary
    personal_glossary = manager.get_user_personal_glossary(user.username)
    
    # Get user selection
    user_selection = manager.get_user_selection(user.username)
    
    # Get version information
    versions = manager.get_all_versions()
    
    return {
        "global_glossaries": [
            {
                "id": g.id,
                "name": g.name,
                "owner": g.owner,
                "is_global": g.is_global,
                "created_at": g.created_at.isoformat(),
                "updated_at": g.updated_at.isoformat(),
                "item_count": g.item_count,
                "description": g.description
            }
            for g in global_glossaries
        ],
        "personal_glossary": {
            "id": personal_glossary.id,
            "name": personal_glossary.name,
            "owner": personal_glossary.owner,
            "is_global": personal_glossary.is_global,
            "created_at": personal_glossary.created_at.isoformat(),
            "updated_at": personal_glossary.updated_at.isoformat(),
            "item_count": personal_glossary.item_count,
            "description": personal_glossary.description
        } if personal_glossary else None,
        "user_selection": {
            "username": user_selection.username,
            "selected_global_glossaries": user_selection.selected_global_glossaries,
            "personal_glossary": user_selection.personal_glossary
        },
        "versions": versions
    }


@auth_router.get("/glossaries/simple-list")
async def get_glossaries_simple_list(user: User = Depends(get_current_user)):
    """Get simplified glossary list for UI selection (ID and name only)"""
    from glossary.manager import get_glossary_manager
    
    manager = get_glossary_manager()
    
    # Get global glossaries
    global_glossaries = manager.get_global_glossaries()
    
    # Get user personal glossary
    personal_glossary = manager.get_user_personal_glossary(user.username)
    
    result = []
    
    # Add global glossaries
    for glossary in global_glossaries:
        result.append({
            "id": glossary.id,
            "name": glossary.name,
            "type": "global",
            "item_count": glossary.item_count
        })
    
    # Add personal glossary if exists
    if personal_glossary:
        result.append({
            "id": personal_glossary.id,
            "name": personal_glossary.name,
            "type": "personal",
            "item_count": personal_glossary.item_count
        })
    
    return {"glossaries": result}


@auth_router.get("/glossaries/check-updates")
async def check_glossaries_updates(
    request: Request,
    user: User = Depends(get_current_user)
):
    """Check glossary updates"""
    from glossary.manager import get_glossary_manager
    
    manager = get_glossary_manager()
    current_versions = manager.get_all_versions()
    
    # Get user's last checked version
    last_check = request.cookies.get('glossaries_last_check', '{}')
    try:
        last_versions = json.loads(last_check)
    except:
        last_versions = {}
    
    # Check if there are updates
    has_updates = False
    for glossary_id, current_version in current_versions.items():
        last_version = last_versions.get(glossary_id, 0)
        if current_version > last_version:
            has_updates = True
            break
    
    return {
        "has_updates": has_updates,
        "current_versions": current_versions
    }


@auth_router.get("/glossaries/export-all")
async def export_all_glossaries(
    target_language: str | None = None,  # Only target_lang is kept, source_lang removed
    category: str | None = None,
    search: str | None = None,
    format: str | None = None,  # xlsx (default) | csvzip
    user: User = Depends(get_current_user)
):
    """Export all glossaries.

    - format=xlsx (default): single XLSX multi-sheet
    - format=csvzip: ZIP archive with per-glossary CSVs (five columns)
    """
    from glossary.manager import get_glossary_manager
    from fastapi.responses import FileResponse
    from openpyxl import Workbook
    import tempfile
    import re
    import io
    import csv
    import zipfile

    manager = get_glossary_manager()

    # Collect glossaries: all global + user's personal (if exists)
    global_glossaries = manager.get_global_glossaries()
    personal_glossary = manager.get_user_personal_glossary(user.username)

    # Helper: sanitize sheet name (max 31 chars, no []:*?/\\)
    def sanitize_sheet_name(name: str) -> str:
        name = re.sub(r"[\[\]:\\/*?]", "_", name)
        return name[:31] if len(name) > 31 else name

    # Helper: apply filters
    def entry_matches_filters(k: str, entry: dict) -> bool:
        if search:
            s = search.lower()
            if s not in k.lower() and s not in (entry.get('dst', '') or '').lower():
                return False
        if category is not None:
            cat = (entry.get('category', '') or '').strip()
            if category == 'unclassified':
                if cat:
                    return False
            else:
                if cat != category:
                    return False
        # language filter (only target_lang is kept, source_lang removed)
        tl = (entry.get('target_lang', '') or '').strip()
        if target_language and tl and tl != target_language:
            return False
        return True

    fmt = (format or "xlsx").lower()
    if fmt == "csvzip":
        # Build zip with per-glossary CSVs
        tmp = tempfile.NamedTemporaryFile(prefix="glossaries_", suffix=".zip", delete=False)
        with zipfile.ZipFile(tmp.name, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            def add_csv_for_glossary(display_name: str, entries: dict[str, dict]):
                csv_buf = io.StringIO()
                writer = csv.writer(csv_buf)
                writer.writerow(['src', 'dst', 'category', 'target_lang'])
                for k, v in entries.items():
                    if not isinstance(v, dict):
                        continue
                    if not entry_matches_filters(k, v):
                        continue
                    writer.writerow([
                        k,
                        v.get('dst', '') or '',
                        v.get('category', '') or '',
                        v.get('target_lang', '') or ''
                    ])
                csv_bytes = csv_buf.getvalue().encode('utf-8-sig')
                safe_name = sanitize_sheet_name(display_name) or "glossary"
                zf.writestr(f"{safe_name}.csv", csv_bytes)

            for g in global_glossaries:
                data = manager.get_glossary_content_with_languages(g.id) or {}
                add_csv_for_glossary(g.name or g.id, data)
            if personal_glossary:
                data = manager.get_glossary_content_with_languages(personal_glossary.id) or {}
                add_csv_for_glossary(personal_glossary.name or personal_glossary.id, data)

        filename = "glossaries_export.zip"
        return FileResponse(tmp.name, filename=filename, media_type="application/zip")

    # Default: XLSX multi-sheet
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    used_names: set[str] = set()
    def add_sheet_for_glossary(display_name: str, entries: dict[str, dict]):
        sheet_name = sanitize_sheet_name(display_name)
        suffix = 1
        while sheet_name in used_names or not sheet_name:
            sheet_name = sanitize_sheet_name(f"{display_name}_{suffix}")
            suffix += 1
        used_names.add(sheet_name)

        ws = wb.create_sheet(title=sheet_name)
        ws.append(["src", "dst", "category", "target_lang"])
        for k, v in entries.items():
            if not isinstance(v, dict):
                continue
            if not entry_matches_filters(k, v):
                continue
            ws.append([
                k,
                v.get('dst', '') or '',
                v.get('category', '') or '',
                v.get('target_lang', '') or ''
            ])

    # Global glossaries
    for g in global_glossaries:
        data = manager.get_glossary_content_with_languages(g.id) or {}
        add_sheet_for_glossary(g.name or g.id, data)

    # Personal glossary
    if personal_glossary:
        data = manager.get_glossary_content_with_languages(personal_glossary.id) or {}
        add_sheet_for_glossary(personal_glossary.name or personal_glossary.id, data)

    # Ensure at least one visible sheet exists
    if not wb.worksheets:
        ws = wb.create_sheet(title="Glossaries")
        ws.append(["src", "dst", "category", "target_lang"])

    # Save file
    tmp = tempfile.NamedTemporaryFile(prefix="glossaries_", suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.flush()

    filename = "glossaries_export.xlsx"
    return FileResponse(tmp.name, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@auth_router.post("/glossaries/upload")
async def upload_glossary(
    request: Request,
    user: User = Depends(get_current_user)
):
    """Upload glossary"""
    from glossary.manager import get_glossary_manager
    
    manager = get_glossary_manager()
    
    try:
        form = await request.form()
        file = form.get("file")
        name = form.get("name", "").strip()
        description = form.get("description", "").strip()
        is_global = form.get("is_global", "false").lower() == "true"
        
        if not file or not name:
            raise HTTPException(status_code=400, detail="File name and glossary name cannot be empty")
        
        # Check permissions
        if is_global and not user.is_admin():
            raise HTTPException(status_code=403, detail="Only administrators can upload global glossaries")
        
        # Read file content with encoding fallback
        # Supported languages: zh, en, ja, ko, fr, de, es, ru, ar, pt
        # Encoding priority: UTF-8 > Chinese > Japanese > Korean > Western European > Russian > Arabic > Fallback
        content = await file.read()
        content_str = None
        for encoding in [
            'utf-8-sig', 'utf-8',  # UTF-8 (universal, highest priority)
            'gbk', 'gb2312', 'gb18030',  # Chinese (Simplified)
            'big5', 'big5hkscs',  # Chinese (Traditional)
            'shift_jis', 'euc-jp', 'iso-2022-jp',  # Japanese
            'euc-kr', 'cp949',  # Korean
            'iso-8859-1', 'latin-1', 'windows-1252',  # Western European (en, fr, de, es, pt)
            'windows-1251', 'koi8-r', 'cp866', 'iso-8859-5',  # Russian
            'windows-1256', 'iso-8859-6', 'cp1256',  # Arabic
        ]:
            try:
                content_str = content.decode(encoding)
                break
            except (UnicodeDecodeError, LookupError):
                continue
        if content_str is None:
            content_str = content.decode('utf-8-sig', errors='replace')
        
        # Parse CSV
        import csv
        from io import StringIO
        
        glossary_dict = {}
        reader = csv.DictReader(StringIO(content_str))
        for row in reader:
            src = row.get('src', '').strip()
            dst = row.get('dst', '').strip()
            if src and dst:
                glossary_dict[src] = dst
        
        # Allow creating an empty glossary (header-only CSV)
        creating_empty = len(glossary_dict) == 0
        
        # Validate glossary
        if not creating_empty:
            is_valid, message = manager.validate_glossary_dict(glossary_dict)
            if not is_valid:
                raise HTTPException(status_code=400, detail=message)
        
        # Save glossary
        if is_global:
            glossary = manager.create_global_glossary(name, glossary_dict, user.username, description)
            logger.info(LogModule.AUTH, f"Administrator {user.username} created global glossary: {name}")
            return {
                "success": True,
                "message": "Glossary uploaded successfully",
                "item_count": len(glossary_dict),
                "glossary_id": glossary.id,
                "glossary_name": glossary.name
            }
        else:
            # Personal glossary
            success = manager.save_user_personal_glossary(user.username, glossary_dict)
            if not success:
                raise HTTPException(status_code=500, detail="Failed to save personal glossary")
            logger.info(LogModule.AUTH, f"User {user.username} updated personal glossary")
            return {
                "success": True,
                "message": "Glossary uploaded successfully",
                "item_count": len(glossary_dict),
                "glossary_id": f"personal_{user.username}",
                "glossary_name": f"{user.username}'s Personal Glossary"
            }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(LogModule.AUTH, f"Glossary upload failed: {e}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@auth_router.get("/glossaries/{glossary_id}/download")
async def download_glossary(
    glossary_id: str,
    user: User = Depends(get_current_user)
):
    """Download glossary with category support"""
    from glossary.manager import get_glossary_manager
    from fastapi.responses import FileResponse
    
    manager = get_glossary_manager()
    
    # Get glossary content with languages (fallback handled later if needed)
    glossary_dict = manager.get_glossary_content_with_languages(glossary_id)
    # Distinguish between not found (None) and empty glossary (empty dict)
    if glossary_dict is None:
        raise HTTPException(status_code=404, detail="Glossary not found")
    
    # Generate temporary CSV file with three columns
    import tempfile
    import csv
    
    temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8-sig')
    writer = csv.writer(temp_file)
    writer.writerow(['src', 'dst', 'category', 'target_lang'])
    for src, entry in glossary_dict.items():
        dst = entry.get('dst', '')
        category = entry.get('category', '')  # 保留空白
        target_lang = entry.get('target_lang', '')
        writer.writerow([src, dst, category, target_lang])
    temp_file.close()
    
    # Determine filename
    if glossary_id.startswith('global_'):
        global_glossaries = manager.get_global_glossaries()
        for g in global_glossaries:
            if g.id == glossary_id:
                filename = f"{g.name}.csv"
                break
        else:
            filename = "glossary.csv"
    else:
        filename = "personal_glossary.csv"
    
    return FileResponse(
        path=temp_file.name,
        filename=filename,
        media_type='text/csv'
    )


@auth_router.get("/glossaries/{glossary_id}/entries")
async def list_glossary_entries(
    glossary_id: str,
    search: str | None = None,
    category: str | None = None,
    target_language: str | None = None,  # Only target_lang is kept, source_lang removed
    language_pair: str | None = None,
    user: User = Depends(get_current_user)
):
    """List entries of a glossary with category and language filters."""
    from glossary.manager import get_glossary_manager

    manager = get_glossary_manager()
    data = manager.get_glossary_content_with_languages(glossary_id)
    if not data:
        # fallback to categories-only format
        data = manager.get_glossary_content_with_categories(glossary_id) or {}

    # parse language_pair if provided (only target_lang is kept)
    if language_pair and not target_language:
        try:
            _, tgt_l = language_pair.split("-", 1)
            target_language = target_language or tgt_l
        except Exception:
            pass

    def match(k: str, entry: dict) -> bool:
        # search filter
        if search:
            s = search.lower()
            if s not in k.lower() and s not in entry.get('dst', '').lower():
                return False

        # category filter
        if category is not None:
            if category == 'unclassified':
                if entry.get('category', '').strip():
                    return False
            else:
                if entry.get('category', '').strip() != category:
                    return False

        # language filter (only target_lang is kept, source_lang removed)
        if target_language and entry.get('target_lang', '').strip() != target_language:
            return False

        return True

    entries = []
    for k, entry in data.items():
        if not match(k, entry):
            continue
        eid = _compute_entry_id(k, entry)
        entries.append({
            "id": eid,
            "src": k,
            "dst": entry.get('dst', ''),
            "category": entry.get('category', ''),
            "target_lang": entry.get('target_lang', ''),
        })

    return {"success": True, "entries": entries, "total": len(entries)}


@auth_router.post("/glossaries/{glossary_id}/import")
async def import_glossary_entries(
    glossary_id: str,
    request: Request,
    merge_mode: str | None = None,  # replace | append | update
    user: User = Depends(get_current_user)
):
    """Import CSV into specified glossary with merge strategy.
    - replace: replace entire glossary with imported
    - append: add new keys only, keep existing
    - update: upsert keys (default)
    """
    from glossary.manager import get_glossary_manager

    form = await request.form()
    file = form.get("file")
    if file is None:
        raise HTTPException(status_code=400, detail="file is required")
    # allow query/body merge_mode override
    mode = (form.get("merge_mode") or merge_mode or "update").lower()
    if mode not in {"replace", "append", "update"}:
        mode = "update"

    content_bytes = await file.read()
    # Try multiple encodings for better compatibility
    # Supported languages: zh, en, ja, ko, fr, de, es, ru, ar, pt
    # Encoding priority: UTF-8 > Chinese > Japanese > Korean > Western European > Russian > Arabic > Fallback
    content_str = None
    for encoding in [
        'utf-8-sig', 'utf-8',  # UTF-8 (universal, highest priority)
        'gbk', 'gb2312', 'gb18030',  # Chinese (Simplified)
        'big5', 'big5hkscs',  # Chinese (Traditional)
        'shift_jis', 'euc-jp', 'iso-2022-jp',  # Japanese
        'euc-kr', 'cp949',  # Korean
        'iso-8859-1', 'latin-1', 'windows-1252',  # Western European (en, fr, de, es, pt)
        'windows-1251', 'koi8-r', 'cp866', 'iso-8859-5',  # Russian
        'windows-1256', 'iso-8859-6', 'cp1256',  # Arabic
    ]:
        try:
            content_str = content_bytes.decode(encoding)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if content_str is None:
        # Fallback: decode with error replacement
        content_str = content_bytes.decode('utf-8-sig', errors='replace')

    # parse csv with language and category support
    # Note: source_lang is removed, only target_lang is kept
    import csv
    from io import StringIO
    reader = csv.DictReader(StringIO(content_str))
    imported: dict[str, dict[str, str]] = {}
    for row in reader:
        src = (row.get('src') or '').strip()
        dst = (row.get('dst') or '').strip()
        category = (row.get('category') or '').strip()  # 保留空白
        target_lang = (row.get('target_lang') or '').strip()
        if src and dst:
            imported[src] = {
                'dst': dst,
                'category': category,
                'target_lang': target_lang,
            }
    if not imported:
        raise HTTPException(status_code=400, detail="Empty CSV")

    manager = get_glossary_manager()
    current = manager.get_glossary_content_with_languages(glossary_id) or {}

    if mode == "replace":
        new_map = imported
    elif mode == "append":
        new_map = {**current}
        for k, v in imported.items():
            if k not in new_map:
                new_map[k] = v
    else:  # update
        new_map = {**current}
        new_map.update(imported)

    # Save with languages
    success = manager.save_glossary_with_languages(glossary_id, new_map, user.username)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to import entries")

    # Calculate statistics for detailed feedback
    existing_keys = set(current.keys())
    imported_keys = set(imported.keys())
    new_terms = len(imported_keys - existing_keys)  # Terms that didn't exist before
    updated_terms = len(imported_keys & existing_keys)  # Terms that were updated
    
    return {
        "success": True,
        "message": "Imported successfully",
        "imported_count": len(imported),
        "total": len(new_map),
        "new_terms": new_terms,
        "updated_terms": updated_terms,
        "glossary_created": False
    }


@auth_router.post("/glossaries/batch-merge")
async def batch_merge_glossaries(
    request: Request,
    user: User = Depends(get_current_user)
):
    """Batch merge glossary entries into multiple target glossaries.
    
    Accepts a CSV file and a list of target glossary IDs, merges the entries
    into all specified glossaries using the specified merge mode.
    """
    from glossary.manager import get_glossary_manager

    form = await request.form()
    file = form.get("file")
    target_ids_str = form.get("target_ids", "").strip()
    merge_mode = (form.get("merge_mode") or "update").lower()
    
    if not file:
        raise HTTPException(status_code=400, detail="file is required")
    if not target_ids_str:
        raise HTTPException(status_code=400, detail="target_ids is required (comma-separated list)")
    
    if merge_mode not in {"replace", "append", "update"}:
        merge_mode = "update"
    
    # Parse target glossary IDs
    target_ids = [tid.strip() for tid in target_ids_str.split(",") if tid.strip()]
    if not target_ids:
        raise HTTPException(status_code=400, detail="At least one target glossary ID is required")
    
    # Check permissions: only admins can merge into global glossaries
    manager = get_glossary_manager()
    for tid in target_ids:
        if tid.startswith('global_') and not user.is_admin():
            raise HTTPException(status_code=403, detail=f"Only administrators can merge into global glossary: {tid}")
    
    # Parse CSV content with encoding fallback
    # Supported languages: zh, en, ja, ko, fr, de, es, ru, ar, pt
    # Encoding priority: UTF-8 > Chinese > Japanese > Korean > Western European > Russian > Arabic > Fallback
    content_bytes = await file.read()
    content_str = None
    for encoding in [
        'utf-8-sig', 'utf-8',  # UTF-8 (universal, highest priority)
        'gbk', 'gb2312', 'gb18030',  # Chinese (Simplified)
        'big5', 'big5hkscs',  # Chinese (Traditional)
        'shift_jis', 'euc-jp', 'iso-2022-jp',  # Japanese
        'euc-kr', 'cp949',  # Korean
        'iso-8859-1', 'latin-1', 'windows-1252',  # Western European (en, fr, de, es, pt)
        'windows-1251', 'koi8-r', 'cp866', 'iso-8859-5',  # Russian
        'windows-1256', 'iso-8859-6', 'cp1256',  # Arabic
    ]:
        try:
            content_str = content_bytes.decode(encoding)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if content_str is None:
        content_str = content_bytes.decode('utf-8-sig', errors='replace')
    
    import csv
    from io import StringIO
    reader = csv.DictReader(StringIO(content_str))
    imported: dict[str, dict[str, str]] = {}
    for row in reader:
        src = (row.get('src') or '').strip()
        dst = (row.get('dst') or '').strip()
        category = (row.get('category') or '').strip()
        target_lang = (row.get('target_lang') or '').strip()
        if src and dst:
            imported[src] = {
                'dst': dst,
                'category': category,
                'target_lang': target_lang,
            }
    
    if not imported:
        raise HTTPException(status_code=400, detail="Empty CSV or no valid entries")
    
    # Merge into each target glossary
    results = []
    for tid in target_ids:
        try:
            current = manager.get_glossary_content_with_languages(tid) or {}
            
            if merge_mode == "replace":
                new_map = imported
            elif merge_mode == "append":
                new_map = {**current}
                for k, v in imported.items():
                    if k not in new_map:
                        new_map[k] = v
            else:  # update
                new_map = {**current}
                new_map.update(imported)
            
            success = manager.save_glossary_with_languages(tid, new_map, user.username)
            if success:
                results.append({
                    "glossary_id": tid,
                    "success": True,
                    "imported_count": len(imported),
                    "total": len(new_map)
                })
            else:
                results.append({
                    "glossary_id": tid,
                    "success": False,
                    "error": "Failed to save glossary"
                })
        except Exception as e:
            results.append({
                "glossary_id": tid,
                "success": False,
                "error": str(e)
            })
    
    success_count = sum(1 for r in results if r.get("success"))
    return {
        "success": success_count > 0,
        "message": f"Merged into {success_count}/{len(target_ids)} glossaries",
        "results": results,
        "total_imported": len(imported)
    }


@auth_router.post("/glossaries/create-and-import")
async def create_and_import_glossary(
    request: Request,
    user: User = Depends(get_current_user)
):
    """Create a new glossary and import entries in one operation.
    
    This is an optimized endpoint that combines createEmptyGlossary and import
    into a single API call, reducing round-trips.
    """
    from glossary.manager import get_glossary_manager
    
    form = await request.form()
    file = form.get("file")
    name = form.get("name", "").strip()
    description = form.get("description", "").strip()
    is_global = form.get("is_global", "false").lower() == "true"
    merge_mode = (form.get("merge_mode") or "update").lower()
    
    if not name:
        raise HTTPException(status_code=400, detail="name is required")
    
    # Check permissions
    if is_global and not user.is_admin():
        raise HTTPException(status_code=403, detail="Only administrators can create global glossaries")
    
    if merge_mode not in {"replace", "append", "update"}:
        merge_mode = "update"
    
    manager = get_glossary_manager()
    
    # Parse CSV if provided with encoding fallback
    # Supported languages: zh, en, ja, ko, fr, de, es, ru, ar, pt
    # Encoding priority: UTF-8 > Chinese > Japanese > Korean > Western European > Russian > Arabic > Fallback
    glossary_dict = {}
    if file:
        content_bytes = await file.read()
        content_str = None
        for encoding in [
            'utf-8-sig', 'utf-8',  # UTF-8 (universal, highest priority)
            'gbk', 'gb2312', 'gb18030',  # Chinese (Simplified)
            'big5', 'big5hkscs',  # Chinese (Traditional)
            'shift_jis', 'euc-jp', 'iso-2022-jp',  # Japanese
            'euc-kr', 'cp949',  # Korean
            'iso-8859-1', 'latin-1', 'windows-1252',  # Western European (en, fr, de, es, pt)
            'windows-1251', 'koi8-r', 'cp866', 'iso-8859-5',  # Russian
            'windows-1256', 'iso-8859-6', 'cp1256',  # Arabic
        ]:
            try:
                content_str = content_bytes.decode(encoding)
                break
            except (UnicodeDecodeError, LookupError):
                continue
        if content_str is None:
            content_str = content_bytes.decode('utf-8-sig', errors='replace')
        
        import csv
        from io import StringIO
        reader = csv.DictReader(StringIO(content_str))
        for row in reader:
            src = (row.get('src') or '').strip()
            dst = (row.get('dst') or '').strip()
            category = (row.get('category') or '').strip()
            target_lang = (row.get('target_lang') or '').strip()
            if src and dst:
                glossary_dict[src] = {
                    'dst': dst,
                    'category': category,
                    'target_lang': target_lang,
                }
    
    # Allow creating empty glossary
    creating_empty = len(glossary_dict) == 0
    
    # Validate if not empty
    if not creating_empty:
        is_valid, message = manager.validate_glossary_dict({k: v.get('dst', '') for k, v in glossary_dict.items()})
        if not is_valid:
            raise HTTPException(status_code=400, detail=message)
    
    # Create glossary
    if is_global:
        glossary = manager.create_global_glossary(name, {k: v.get('dst', '') for k, v in glossary_dict.items()}, user.username, description)
        # If has language/category info, save with languages
        if glossary_dict and any(v.get('category') or v.get('target_lang') for v in glossary_dict.values()):
            manager.save_glossary_with_languages(glossary.id, glossary_dict, user.username)
        logger.info(LogModule.AUTH, f"Administrator {user.username} created global glossary: {name} with {len(glossary_dict)} entries")
        return {
            "success": True,
            "message": "Glossary created and imported successfully",
            "glossary_id": glossary.id,
            "glossary_name": glossary.name,
            "item_count": len(glossary_dict)
        }
    else:
        # Personal glossary
        success = manager.save_user_personal_glossary(user.username, {k: v.get('dst', '') for k, v in glossary_dict.items()})
        if not success:
            raise HTTPException(status_code=500, detail="Failed to create personal glossary")
        # If has language/category info, save with languages
        if glossary_dict and any(v.get('category') or v.get('target_lang') for v in glossary_dict.values()):
            personal_id = f"personal_{user.username}"
            manager.save_glossary_with_languages(personal_id, glossary_dict, user.username)
        logger.info(LogModule.AUTH, f"User {user.username} created/updated personal glossary with {len(glossary_dict)} entries")
        return {
            "success": True,
            "message": "Glossary created and imported successfully",
            "glossary_id": f"personal_{user.username}",
            "glossary_name": f"{user.username}'s Personal Glossary",
            "item_count": len(glossary_dict)
        }


@auth_router.post("/glossaries/{glossary_id}/apply-to-task/{task_id}")
async def apply_glossary_to_task(
    glossary_id: str,
    task_id: str,
    user: User = Depends(get_current_user)
):
    """Apply a glossary to a translation task.
    
    This endpoint allows applying a glossary to an existing translation task,
    which will be used in subsequent translation operations for that task.
    """
    from glossary.manager import get_glossary_manager
    from backend.app.services.task import task_manager
    
    manager = get_glossary_manager()
    
    # Get glossary content
    glossary_dict = manager.get_glossary_content_with_languages(glossary_id)
    if glossary_dict is None:
        # Fallback to category-only format
        glossary_dict = manager.get_glossary_content_with_categories(glossary_id) or {}
    
    if not glossary_dict:
        raise HTTPException(status_code=404, detail="Glossary not found or empty")
    
    # Check if task exists
    task_state = task_manager.get_task(task_id)
    if task_state is None:
        raise HTTPException(status_code=404, detail=f"Task ID '{task_id}' not found")
    
    # Convert to simple dict format (src -> dst) for task payload
    simple_dict = {}
    for src, entry in glossary_dict.items():
        if isinstance(entry, dict):
            simple_dict[src] = entry.get('dst', '')
        else:
            simple_dict[src] = entry
    
    # Store glossary in task state for use in translation
    task_manager.update_task(task_id, {"applied_glossary": {
        "glossary_id": glossary_id,
        "glossary_dict": simple_dict,
        "applied_by": user.username,
        "applied_at": time.time()
    }})
    
    logger.info(LogModule.AUTH, f"[DEBUG] User {user.username} applied glossary {glossary_id} to task {task_id}")
    logger.info(LogModule.AUTH, f"[DEBUG] Applied glossary dict size: {len(simple_dict)} entries")
    logger.info(LogModule.AUTH, f"[DEBUG] Applied glossary sample (first 10 entries): {dict(list(simple_dict.items())[:10])}")
    logger.info(LogModule.AUTH, f"User {user.username} applied glossary {glossary_id} to task {task_id}")
    
    return {
        "success": True,
        "message": f"Glossary applied to task {task_id}",
        "glossary_id": glossary_id,
        "entry_count": len(simple_dict)
    }


def _save_glossary_by_id(manager, glossary_id: str, data: dict, username: str) -> None:
    """Helper to persist glossary changes based on id type."""
    if glossary_id.startswith('global_'):
        ok = manager.update_global_glossary(glossary_id, data, username)
        if not ok:
            raise HTTPException(status_code=500, detail="Failed to update global glossary")
    elif glossary_id.startswith('personal_'):
        # personal id format: personal_<username>
        ok = manager.save_user_personal_glossary(username, data)
        if not ok:
            raise HTTPException(status_code=500, detail="Failed to save personal glossary")
    else:
        raise HTTPException(status_code=400, detail="Invalid glossary id")


def _compute_entry_id(src: str, entry: dict) -> str:
    category = (entry.get('category') or '').strip()
    tl = (entry.get('target_lang') or '').strip()
    raw = f"{src}|{category}|{tl}"
    return hashlib.sha1(raw.encode('utf-8')).hexdigest()[:12]


@auth_router.post("/glossaries/{glossary_id}/entries")
async def create_glossary_entry(
    glossary_id: str,
    payload: dict,
    user: User = Depends(get_current_user)
):
    """Create or upsert a glossary entry with language and category support."""
    from glossary.manager import get_glossary_manager

    src = (payload.get("src") or "").strip()
    dst = (payload.get("dst") or "").strip()
    category = (payload.get("category") or "").strip()
    target_lang = (payload.get("target_lang") or "").strip()
    if not src or not dst:
        raise HTTPException(status_code=400, detail="'src' and 'dst' are required")

    manager = get_glossary_manager()
    data = manager.get_glossary_content_with_languages(glossary_id) or {}
    data[src] = {
        'dst': dst,
        'category': category,
        'target_lang': target_lang,
    }

    # Save with languages
    success = manager.save_glossary_with_languages(glossary_id, data, user.username)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to save entry")
    eid = _compute_entry_id(src, data[src])
    return {"success": True, "message": "Entry saved", "entry": {"id": eid, "src": src, "dst": dst, "category": category, "target_lang": target_lang}}


@auth_router.put("/glossaries/{glossary_id}/entries/{entry_id}")
async def update_glossary_entry(
    glossary_id: str,
    entry_id: str,
    payload: dict,
    user: User = Depends(get_current_user)
):
    """Update existing entry with language and category support. If 'src' changes, rename key."""
    from glossary.manager import get_glossary_manager

    new_src = (payload.get("src") or '').strip()
    new_dst = (payload.get("dst") or "").strip()
    new_category = (payload.get("category") or "").strip()
    new_target_lang = (payload.get("target_lang") or "").strip()
    if not new_dst:
        raise HTTPException(status_code=400, detail="'dst' is required")

    manager = get_glossary_manager()
    data = manager.get_glossary_content_with_languages(glossary_id) or {}
    # locate original by computed id
    original_key = None
    for k, v in data.items():
        if _compute_entry_id(k, v) == entry_id:
            original_key = k
            break

    if original_key is None and (not new_src or new_src not in data):
        data[new_src] = {
            'dst': new_dst,
            'category': new_category,
            'target_lang': new_target_lang,
        }
    else:
        if not new_src:
            new_src = original_key or ''
        if original_key and new_src != original_key:
            data.pop(original_key, None)
        data[new_src] = {
            'dst': new_dst,
            'category': new_category,
            'target_lang': new_target_lang,
        }

    success = manager.save_glossary_with_languages(glossary_id, data, user.username)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to update entry")

    eid = _compute_entry_id(new_src, data[new_src])
    return {"success": True, "message": "Entry updated", "entry": {"id": eid, "src": new_src, "dst": new_dst, "category": new_category, "target_lang": new_target_lang}}


@auth_router.delete("/glossaries/{glossary_id}/entries/{entry_id}")
async def delete_glossary_entry(
    glossary_id: str,
    entry_id: str,
    user: User = Depends(get_current_user)
):
    """Delete an entry by id (src) with language support."""
    from glossary.manager import get_glossary_manager

    manager = get_glossary_manager()
    data = manager.get_glossary_content_with_languages(glossary_id) or {}
    # find by computed id
    target_key = None
    for k, v in data.items():
        if _compute_entry_id(k, v) == entry_id:
            target_key = k
            break
    if target_key is None:
        raise HTTPException(status_code=404, detail="Entry not found")
    data.pop(target_key, None)

    success = manager.save_glossary_with_languages(glossary_id, data, user.username)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to delete entry")

    return {"success": True, "message": "Entry deleted"}


@auth_router.put("/glossaries/selection")
async def update_glossary_selection(
    request: Request,
    user: User = Depends(get_current_user)
):
    """Update user glossary selection"""
    from glossary.manager import get_glossary_manager
    from glossary.models import UserGlossarySelection
    
    manager = get_glossary_manager()
    
    try:
        data = await request.json()
        logger.info(LogModule.AUTH, f"[LDAP-API] Received update request: {data}")
        selected_global_glossaries = data.get("selected_global_glossaries", [])
        personal_glossary = data.get("personal_glossary")
        
        # Verify selected global glossaries exist
        global_glossaries = manager.get_global_glossaries()
        valid_global_ids = [g.id for g in global_glossaries]
        
        for glossary_id in selected_global_glossaries:
            if glossary_id not in valid_global_ids:
                raise HTTPException(status_code=400, detail=f"Glossary {glossary_id} not found")
        
        # Verify personal glossary
        if personal_glossary and personal_glossary != f"personal_{user.username}":
            raise HTTPException(status_code=400, detail="Invalid personal glossary ID")
        
        # Save selection
        selection = UserGlossarySelection(
            username=user.username,
            selected_global_glossaries=selected_global_glossaries,
            personal_glossary=personal_glossary
        )
        manager.save_user_selection(selection)
        
        logger.info(LogModule.AUTH, f"User {user.username} updated glossary selection")
        
        return {"success": True, "message": "Glossary selection updated"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(LogModule.AUTH, f"Failed to update glossary selection: {e}")
        raise HTTPException(status_code=500, detail=f"Update failed: {str(e)}")


@auth_router.delete("/glossaries/{glossary_id}")
async def delete_glossary(
    glossary_id: str,
    user: User = Depends(get_current_user)
):
    """Delete glossary"""
    from glossary.manager import get_glossary_manager
    
    manager = get_glossary_manager()
    
    # Check permissions
    if glossary_id.startswith('global_'):
        if not user.is_admin():
            raise HTTPException(status_code=403, detail="Only administrators can delete global glossaries")
        
        success = manager.delete_global_glossary(glossary_id)
        if success:
            logger.info(LogModule.AUTH, f"Administrator {user.username} deleted global glossary: {glossary_id}")
        else:
            raise HTTPException(status_code=404, detail="Glossary not found")
    else:
        # Personal glossary - users can only delete their own
        if not glossary_id.startswith(f"personal_{user.username}"):
            raise HTTPException(status_code=403, detail="Can only delete own personal glossary")
        
        # Clear personal glossary
        success = manager.save_user_personal_glossary(user.username, {})
        if success:
            logger.info(LogModule.AUTH, f"User {user.username} cleared personal glossary")
        else:
            raise HTTPException(status_code=500, detail="Failed to delete personal glossary")
    
    return {"success": True, "message": "Glossary deleted"}


# === Prompt Management API ===

@auth_router.get("/prompts")
async def get_prompts_list(
    user: User = Depends(get_current_user)
):
    """Get prompt list"""
    from prompts.manager import get_prompt_manager
    
    manager = get_prompt_manager()
    
    # Get global prompts
    global_prompts = manager.get_global_prompts()
    
    # Get user personal prompts
    personal_prompt = manager.get_user_personal_prompt(user.username)
    
    # Get user selection
    user_selection = manager.get_user_selection(user.username)
    
    # Get version information
    versions = manager.get_all_versions()
    
    return {
        "global_prompts": [
            {
                "id": p.id,
                "name": p.name,
                "owner": p.owner,
                "is_global": p.is_global,
                "created_at": p.created_at.isoformat(),
                "updated_at": p.updated_at.isoformat(),
                "item_count": p.item_count,
                "description": p.description
            }
            for p in global_prompts
        ],
        "personal_prompt": {
            "id": personal_prompt.id,
            "name": personal_prompt.name,
            "owner": personal_prompt.owner,
            "is_global": personal_prompt.is_global,
            "created_at": personal_prompt.created_at.isoformat(),
            "updated_at": personal_prompt.updated_at.isoformat(),
            "item_count": personal_prompt.item_count,
            "description": personal_prompt.description
        } if personal_prompt else None,
        "user_selection": {
            "username": user_selection.username,
            "selected_global_prompts": user_selection.selected_global_prompts,
            "personal_prompt": user_selection.personal_prompt
        },
        "versions": versions
    }


@auth_router.get("/prompts/check-updates")
async def check_prompts_updates(
    request: Request,
    user: User = Depends(get_current_user)
):
    """Check prompt updates"""
    from prompts.manager import get_prompt_manager
    
    manager = get_prompt_manager()
    
    try:
        # Get current version information
        current_versions = manager.get_all_versions()
        
        # More complex update checking logic can be added here
        # e.g. checking file modification time, etc.
        
        return {
            "has_updates": False,  # Simplified implementation, always returns no updates
            "current_versions": current_versions
        }
        
    except Exception as e:
        logger.error(LogModule.AUTH, f"Failed to check prompt updates: {e}")
        return {
            "has_updates": False,
            "current_versions": {}
        }


@auth_router.post("/prompts/upload")
async def upload_prompt(
    request: Request,
    user: User = Depends(get_current_user)
):
    """Upload prompt"""
    from prompts.manager import get_prompt_manager
    
    manager = get_prompt_manager()
    
    try:
        form = await request.form()
        file = form.get("file")
        name = form.get("name", "").strip()
        description = form.get("description", "").strip()
        is_global = form.get("is_global", "false").lower() == "true"
        
        if not file or not name:
            raise HTTPException(status_code=400, detail="File name and prompt name cannot be empty")
        
        # Check permissions
        if is_global and not user.is_admin():
            raise HTTPException(status_code=403, detail="Only administrators can upload global prompts")
        
        # Read file content
        content = await file.read()
        content_str = content.decode('utf-8-sig')
        
        # Parse JSON
        import json
        try:
            prompts_dict = json.loads(content_str)
        except json.JSONDecodeError as e:
            raise HTTPException(status_code=400, detail=f"JSON format error: {str(e)}")
        
        if not prompts_dict:
            raise HTTPException(status_code=400, detail="Prompts cannot be empty")
        
        # Validate prompts
        is_valid, message = manager.validate_prompt_dict(prompts_dict)
        if not is_valid:
            raise HTTPException(status_code=400, detail=message)
        
        # Save prompts
        if is_global:
            prompt = manager.create_global_prompt(name, prompts_dict, user.username, description)
            logger.info(LogModule.AUTH, f"Administrator {user.username} created global prompt: {name}")
        else:
            # Personal prompts
            success = manager.save_user_personal_prompt(user.username, prompts_dict)
            if not success:
                raise HTTPException(status_code=500, detail="Failed to save personal prompt")
            logger.info(LogModule.AUTH, f"User {user.username} updated personal prompt")
        
        return {
            "success": True,
            "message": "Prompts uploaded successfully",
            "item_count": len(prompts_dict)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(LogModule.AUTH, f"Failed to upload prompts: {e}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@auth_router.get("/prompts/{prompt_id}/download")
async def download_prompt(
    prompt_id: str,
    user: User = Depends(get_current_user)
):
    """Download prompt"""
    from prompts.manager import get_prompt_manager
    
    manager = get_prompt_manager()
    
    # Get prompt file
    if prompt_id.startswith('global_'):
        global_prompts = manager.get_global_prompts()
        prompt_file = None
        for p in global_prompts:
            if p.id == prompt_id:
                prompt_file = p
                break
        
        if not prompt_file:
            raise HTTPException(status_code=404, detail="Prompt not found")
        
        # Read prompt content
        prompts_dict = manager.storage.load_prompts_from_json(
            manager.storage.global_dir / manager.storage.global_prompts[prompt_id]['file_path']
        )
        
        filename = f"{prompt_file.name}.json"
        
    elif prompt_id.startswith(f"personal_{user.username}"):
        # Personal prompt
        personal_prompt = manager.get_user_personal_prompt(user.username)
        if not personal_prompt:
            raise HTTPException(status_code=404, detail="Personal prompt not found")
        
        prompts_dict = manager.storage.load_prompts_from_json(
            manager.storage.users_dir / f"{user.username}_prompts.json"
        )
        filename = f"{user.username}_personal_prompts.json"
        
    else:
        raise HTTPException(status_code=404, detail="Prompt not found")
    
    # Generate JSON content
    import json
    content = json.dumps(prompts_dict, ensure_ascii=False, indent=2)
    
    return Response(
        content=content,
        headers={"Content-Disposition": f"attachment; filename={filename}"},
        media_type='application/json'
    )


@auth_router.put("/prompts/selection")
async def update_prompt_selection(
    request: Request,
    user: User = Depends(get_current_user)
):
    """Update user prompt selection"""
    from prompts.manager import get_prompt_manager
    from prompts.models import UserPromptSelection
    
    manager = get_prompt_manager()
    
    try:
        data = await request.json()
        logger.info(LogModule.AUTH, f"[PROMPT-API] Received update request: {data}")
        selected_global_prompts = data.get("selected_global_prompts", [])
        personal_prompt = data.get("personal_prompt")
        
        # Verify selected global prompts exist
        global_prompts = manager.get_global_prompts()
        valid_global_ids = [p.id for p in global_prompts]
        
        for prompt_id in selected_global_prompts:
            if prompt_id not in valid_global_ids:
                raise HTTPException(status_code=400, detail=f"Prompt {prompt_id} not found")
        
        # Verify personal prompt
        if personal_prompt and personal_prompt != f"personal_{user.username}":
            raise HTTPException(status_code=400, detail="Invalid personal prompt ID")
        
        # Save selection
        selection = UserPromptSelection(
            username=user.username,
            selected_global_prompts=selected_global_prompts,
            personal_prompt=personal_prompt
        )
        manager.save_user_selection(selection)
        
        logger.info(LogModule.AUTH, f"User {user.username} updated prompt selection")
        
        return {"success": True, "message": "Prompt selection updated"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(LogModule.AUTH, f"Failed to update prompt selection: {e}")
        raise HTTPException(status_code=500, detail=f"Update failed: {str(e)}")


@auth_router.delete("/prompts/personal")
async def delete_personal_prompt(
    user: User = Depends(get_current_user)
):
    """Delete user personal prompt"""
    from prompts.manager import get_prompt_manager
    
    manager = get_prompt_manager()
    
    try:
        # Check if personal prompt exists
        personal_prompt = manager.get_user_personal_prompt(user.username)
        if not personal_prompt:
            raise HTTPException(status_code=404, detail="Personal prompt not found")
        
        # Delete personal prompt file
        success = manager.storage.delete_user_personal_prompt(user.username)
        if not success:
            raise HTTPException(status_code=500, detail="Failed to delete personal prompt")
        
        logger.info(LogModule.AUTH, f"User {user.username} deleted personal prompt")
        
        return {"success": True, "message": "Personal prompt deleted"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(LogModule.AUTH, f"Failed to delete personal prompt: {e}")
        raise HTTPException(status_code=500, detail=f"Delete failed: {str(e)}")


@auth_router.delete("/prompts/{prompt_id}")
async def delete_prompt(
    prompt_id: str,
    user: User = Depends(get_current_user)
):
    """Delete prompt"""
    from prompts.manager import get_prompt_manager
    
    manager = get_prompt_manager()
    
    # Check permissions
    if prompt_id.startswith('global_'):
        if not user.is_admin():
            raise HTTPException(status_code=403, detail="Only administrators can delete global prompts")
        
        success = manager.delete_global_prompt(prompt_id)
        if success:
            logger.info(LogModule.AUTH, f"Administrator {user.username} deleted global prompt: {prompt_id}")
        else:
            raise HTTPException(status_code=404, detail="Prompt not found")
    else:
        # Personal prompt - users can only delete their own
        if not prompt_id.startswith(f"personal_{user.username}"):
            raise HTTPException(status_code=403, detail="Can only delete own personal prompt")
        
        # Clear personal prompt
        success = manager.save_user_personal_prompt(user.username, {})
        if success:
            logger.info(LogModule.AUTH, f"User {user.username} cleared personal prompt")
        else:
            raise HTTPException(status_code=500, detail="Failed to delete personal prompt")
    
    return {"success": True, "message": "Prompt deleted"}


@auth_router.get("/prompts/merged")
async def get_merged_prompts(
    user: User = Depends(get_current_user)
):
    """Get user merged prompts"""
    from prompts.manager import get_prompt_manager
    
    manager = get_prompt_manager()
    merged_prompts = manager.get_merged_prompts(user.username)
    
    return {
        "prompts": merged_prompts,
        "count": len(merged_prompts)
    }


# === Simplified Prompt Management API ===

@auth_router.get("/prompts/simple")
async def get_simple_prompts(
    user: User = Depends(get_current_user)
):
    """Get simplified global prompt list"""
    from prompts.manager import get_prompt_manager
    
    manager = get_prompt_manager()
    
    # Get global prompts collection
    global_prompts = manager.get_global_prompts()
    
    # Find global prompt collection named "Simple Prompts"
    simple_prompts_collection = None
    for prompt_file in global_prompts:
        if prompt_file.name == "Simple Prompts":
            simple_prompts_collection = prompt_file
            break
    
    if simple_prompts_collection:
        # Load prompt content
        prompts_dict = manager.storage.load_prompts_from_json(
            Path(simple_prompts_collection.file_path)
        )
        
        # Convert to simplified format
        simple_prompts = [
            {"id": f"global_{i}", "name": name, "content": content}
            for i, (name, content) in enumerate(prompts_dict.items())
        ]
        
        return simple_prompts
    else:
        return []


@auth_router.post("/prompts/simple")
async def add_simple_prompt(
    request: Request,
    user: User = Depends(get_current_user)
):
    """Add simplified global prompt"""
    from prompts.manager import get_prompt_manager
    
    manager = get_prompt_manager()
    
    try:
        data = await request.json()
        name = data.get("name", "").strip()
        content = data.get("content", "").strip()
        
        if not name or not content:
            raise HTTPException(status_code=400, detail="Prompt description and content cannot be empty")
        
        # Get global prompts collection
        global_prompts = manager.get_global_prompts()
        
        # Find global prompt collection named "Simple Prompts"
        simple_prompts_collection = None
        for prompt_file in global_prompts:
            if prompt_file.name == "Simple Prompts":
                simple_prompts_collection = prompt_file
                break
        
        if simple_prompts_collection:
            # Load existing prompts
            prompts_dict = manager.storage.load_prompts_from_json(
                Path(simple_prompts_collection.file_path)
            )
        else:
            # Create new global prompt collection
            prompts_dict = {}
            simple_prompts_collection = manager.create_global_prompt(
                name="Simple Prompts",
                prompts_dict={},
                owner=user.username,
                description="Simplified global prompt collection"
            )
        
        # Add new prompt
        prompts_dict[name] = content
        
        # Update global prompt
        success = manager.update_global_prompt(
            simple_prompts_collection.id,
            prompts_dict,
            user.username
        )
        if not success:
            raise HTTPException(status_code=500, detail="Failed to save prompt")
        
        logger.info(LogModule.AUTH, f"User {user.username} added global prompt: {name}")
        
        return {"success": True, "message": "Prompt added successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(LogModule.AUTH, f"Failed to add prompt: {e}")
        raise HTTPException(status_code=500, detail=f"Add failed: {str(e)}")


@auth_router.delete("/prompts/simple/{prompt_id}")
async def delete_simple_prompt(
    prompt_id: str,
    user: User = Depends(get_current_user)
):
    """Delete simplified global prompt"""
    from prompts.manager import get_prompt_manager
    
    manager = get_prompt_manager()
    
    try:
        # Parse prompt ID
        if not prompt_id.startswith("global_"):
            raise HTTPException(status_code=400, detail="Invalid prompt ID")
        
        index = int(prompt_id.replace("global_", ""))
        
        # Get global prompts collection
        global_prompts = manager.get_global_prompts()
        
        # Find global prompt collection named "Simple Prompts"
        simple_prompts_collection = None
        for prompt_file in global_prompts:
            if prompt_file.name == "Simple Prompts":
                simple_prompts_collection = prompt_file
                break
        
        if not simple_prompts_collection:
            raise HTTPException(status_code=404, detail="Global prompt collection not found")
        
        # Load prompts
        prompts_dict = manager.storage.load_prompts_from_json(
            Path(simple_prompts_collection.file_path)
        )
        
        # Get prompt name to delete
        prompt_names = list(prompts_dict.keys())
        if index >= len(prompt_names):
            raise HTTPException(status_code=404, detail="Prompt not found")
        
        prompt_name = prompt_names[index]
        
        # Delete prompt
        del prompts_dict[prompt_name]
        
        # Update global prompt
        success = manager.update_global_prompt(
            simple_prompts_collection.id,
            prompts_dict,
            user.username
        )
        if not success:
            raise HTTPException(status_code=500, detail="Failed to save prompt")
        
        logger.info(LogModule.AUTH, f"User {user.username} deleted global prompt: {prompt_name}")
        
        return {"success": True, "message": "Prompt deleted successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(LogModule.AUTH, f"Failed to delete prompt: {e}")
        raise HTTPException(status_code=500, detail=f"Delete failed: {str(e)}")


@auth_router.post("/app-config")
async def update_app_config_api(
    request: Request,
    user: User = Depends(get_current_user)
):
    """Update application configuration (requires administrator or management group permissions; only super administrator can change default password)"""
    if not user.is_admin():
        raise HTTPException(status_code=403, detail="Access denied: Admin privileges required")
    
    try:
        config_data = await request.json()
        
        # Separate LDAP-related keys from App configuration keys to avoid LDAP keys being mistakenly included in app_config
        ldap_keys = {
            'ldap_enabled','ldap_protocol','ldap_host','ldap_port','ldap_bind_dn_template','ldap_base_dn',
            'ldap_user_filter','ldap_tls_cacertfile','ldap_tls_verify','ldap_admin_group_enabled','ldap_admin_group',
            'ldap_glossary_group_enabled','ldap_glossary_group','ldap_group_base_dn'
        }
        ldap_updates = {k: v for k, v in config_data.items() if k in ldap_keys}
        config_data = {k: v for k, v in config_data.items() if k not in ldap_keys}
        
        # First handle LDAP updates (unified to new keys) and write to auth_config
        if ldap_updates:
            try:
                from .config import get_auth_config as _get_auth_cfg, save_auth_config as _save_auth_cfg
                auth_cfg = _get_auth_cfg()
                # Backup old endpoint-related values before saving
                import copy
                old_for_endpoint = copy.deepcopy(auth_cfg)
                auth_cfg.update_from_dict(ldap_updates)
                if _save_auth_cfg():
                    logger.info(LogModule.AUTH, f"[APP-CONFIG] Successfully synchronized LDAP configuration: {list(ldap_updates.keys())}")
                    # Synchronously refresh in-memory instance in this module to ensure subsequent GET reads latest values
                    try:
                        global _auth_config
                        if _auth_config is not None:
                            _auth_config.update_from_dict(ldap_updates)
                            logger.info(LogModule.AUTH, "[APP-CONFIG] Successfully synchronized _auth_config in module")
                        # Hot reload LDAP client (if endpoint changes)
                        _refresh_ldap_client_if_endpoint_changed(old_for_endpoint, auth_cfg)
                    except Exception as _e:
                        logger.warning(LogModule.AUTH, f"[APP-CONFIG] Failed to synchronize in-memory module: {_e}")
                else:
                    logger.warning(LogModule.AUTH, "[APP-CONFIG] Failed to synchronize LDAP configuration")
            except Exception as _e:
                logger.error(LogModule.AUTH, f"[APP-CONFIG] Exception when processing LDAP configuration: {_e}")

        app_config = get_app_config()
        
        # Remove any platform_api_keys from frontend (sensitive information not saved in application configuration)
        if 'platform_api_keys' in config_data:
            del config_data['platform_api_keys']
        
        
        # Handle MinerU Token (save to sensitive configuration) - supports {key, configured}
        if 'translator_mineru_token' in config_data:
            token_val = config_data['translator_mineru_token']
            from backend.config.secrets_manager import get_secrets_manager
            secrets_manager = get_secrets_manager()
            if isinstance(token_val, dict):
                raw = token_val.get('key', '')
                configured = token_val.get('configured')
                if raw and not str(raw).endswith('***'):
                    secrets_manager.update_mineru_token(str(raw), configured)
            else:
                raw = token_val
                if raw and not str(raw).endswith('***'):
                    secrets_manager.update_mineru_token(str(raw))
            del config_data['translator_mineru_token']
        
        
        # Handle Web/HTTPS related fields and write to local configuration
        from backend.config.config_loader import get_unified_config, save_unified_config
        from backend.config.local_config import LocalConfig
        global_cfg = get_unified_config()
        local_cfg = LocalConfig.load_from_file()

        # Handle HTTPS configuration
        if 'https' in config_data:
            https_config = config_data['https']
            if isinstance(https_config, dict):
                if 'enabled' in https_config:
                    local_cfg.https.enabled = bool(https_config['enabled'])
                if 'force_redirect' in https_config:
                    local_cfg.https.force_redirect = bool(https_config['force_redirect'])
                if 'cert_file' in https_config:
                    local_cfg.https.cert_file = https_config['cert_file'] or None
                if 'key_file' in https_config:
                    local_cfg.https.key_file = https_config['key_file'] or None
        

        # If HTTPS is requested to be enabled, perform strong validation before saving (ensure it has passed testing)
        try:
            if bool(local_cfg.https.enabled):
                cert_file = local_cfg.https.cert_file
                key_file = local_cfg.https.key_file
                if not (cert_file and key_file and os.path.exists(cert_file) and os.path.exists(key_file)):
                    raise HTTPException(status_code=400, detail="Enable HTTPS failed: certificate or key not found")
                import ssl as _ssl
                from backend.config.secrets_manager import get_secrets_manager as _get_sm
                _pwd = _get_sm().get_web_tls_password()
                ctx = _ssl.create_default_context(purpose=_ssl.Purpose.CLIENT_AUTH)
                ctx.load_cert_chain(certfile=cert_file, keyfile=key_file, password=_pwd)
        except HTTPException:
            raise
        except Exception as _e:
            raise HTTPException(status_code=400, detail=f"Enable HTTPS failed: {str(_e)}")

        # Handle AI platform configuration updates
        ai_platform_updates = {}
        parsing_engine_updates = {}
        
        if 'ai_platforms' in config_data:
            ai_platforms_data = config_data['ai_platforms']
            # Remove API keys from platform data (they are stored separately)
            # Also remove default_platform if present (it's stored in platforms.json, not global_config.json)
            for platform_key, platform_data in ai_platforms_data.items():
                if platform_key == 'default_platform':
                    # Handle default_platform separately - save to platforms.json only
                    if isinstance(platform_data, str) and platform_data.strip():
                        try:
                            from backend.config.platforms_config import get_platforms_config, save_platforms_config
                            platforms_config = get_platforms_config()
                            platforms_config.default_platform = platform_data.strip()
                            if not save_platforms_config():
                                logger.warning(LogModule.AUTH, f"[APP-CONFIG] Failed to save default_platform to platforms.json")
                            else:
                                logger.info(LogModule.AUTH, f"[APP-CONFIG] Default platform updated to: {platform_data.strip()}")
                        except Exception as _e:
                            logger.warning(LogModule.AUTH, f"[APP-CONFIG] Failed to update default_platform: {_e}")
                    continue
                if isinstance(platform_data, dict):
                    platform_data = platform_data.copy()
                    platform_data.pop('api_key', None)
                    ai_platforms_data[platform_key] = platform_data
            
            # Remove default_platform from ai_platforms_data before saving to global_config
            ai_platforms_data.pop('default_platform', None)
            ai_platform_updates['ai_platforms'] = ai_platforms_data
            del config_data['ai_platforms']
        
        # Handle parsing_engine (new name) or translator_settings (backward compatibility)
        if 'parsing_engine' in config_data:
            parsing_engine_updates['parsing_engine'] = config_data['parsing_engine']
            del config_data['parsing_engine']
        elif 'translator_settings' in config_data:
            # Backward compatibility: map translator_settings to parsing_engine
            parsing_engine_updates['parsing_engine'] = config_data['translator_settings']
            del config_data['translator_settings']
        
        # Update global configuration with new structured data
        if ai_platform_updates or parsing_engine_updates:
            global_cfg.update_from_dict({**ai_platform_updates, **parsing_engine_updates})
        
        # Handle default language, write to global configuration root fields
        if 'default_language' in config_data:
            try:
                dl = str(config_data.get('default_language') or '').lower()
                if dl in ('zh', 'en'):
                    setattr(global_cfg, 'default_language', dl)
                else:
                    # Simple fallback: unexpected values default to en
                    setattr(global_cfg, 'default_language', 'en')
            except Exception as _e:
                logger.warning(LogModule.AUTH, f"[APP-CONFIG] Failed to update default language: {_e}")
            finally:
                # Avoid writing to user-level App configuration simultaneously
                del config_data['default_language']

        # Update other configurations (user-level App configuration)
        app_config.update_from_dict({k: v for k, v in config_data.items() if k not in ['https']})
        
        # Save configuration
        # Save HTTPS private key password to sensitive configuration
        from backend.config.secrets_manager import get_secrets_manager
        secrets_manager = get_secrets_manager()
        if 'https_key_password' in config_data:
            secrets_manager.update_web_tls_password(config_data.get('https_key_password') or None)

        ok1 = save_app_config()
        ok2 = save_unified_config()
        ok3 = local_cfg.save_to_file()
        if ok1 and ok2 and ok3:
            logger.info(LogModule.AUTH, f"Application configuration updated by user {_mask_username(user.username)}")
            return {"success": True, "message": "Configuration updated successfully"}
        else:
            raise HTTPException(status_code=500, detail="Failed to save configuration")
    except Exception as e:
        logger.error(LogModule.AUTH, f"Failed to update application configuration: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to update configuration: {str(e)}")


@auth_router.post("/settings/batch")
async def batch_update_settings(
    request: Request,
    user: User = Depends(get_current_user)
):
    """Batch update settings (grouped by type: global/user/sensitive)"""
    try:
        data = await request.json()
        config_type = data.get('type')  # 'global' | 'user' | 'sensitive'
        changes = data.get('changes', {})
        
        if not config_type or not changes:
            raise HTTPException(status_code=400, detail="type and changes are required")
        
        if config_type not in ['global', 'user', 'sensitive']:
            raise HTTPException(status_code=400, detail=f"Invalid config type: {config_type}. Must be one of: global, user, sensitive")
        
        from .user_profile import get_user_profile_manager
        from backend.config.config_loader import get_unified_config, save_unified_config
        from backend.config.secrets_manager import get_secrets_manager
        from backend.config.platforms_config import get_platforms_config, save_platforms_config
        
        if config_type == 'global':
            # Update global configuration (admin only)
            if not user.is_admin():
                raise HTTPException(status_code=403, detail="Admin required for global settings")
            
            global_config = get_unified_config()
            
            # Process each change
            for key, value in changes.items():
                # Handle special keys
                if key == 'ai_platforms_default_platform':
                    # Save to platforms.json only (new config structure)
                    try:
                        if not isinstance(value, str) or not value.strip():
                            raise ValueError("default_platform must be a non-empty string")
                        
                        # Save to new platforms.json structure
                        platforms_config = get_platforms_config()
                        platforms_config.default_platform = value.strip()
                        if not save_platforms_config():
                            raise HTTPException(status_code=500, detail="Failed to save default_platform to platforms.json")
                        
                        logger.info(LogModule.AUTH, f"Default platform updated to: {value.strip()}")
                    except HTTPException:
                        raise
                    except Exception as e:
                        logger.error(LogModule.AUTH, f"Failed to update default_platform: {e}", exc_info=True)
                        raise HTTPException(status_code=400, detail=f"Invalid value for ai_platforms_default_platform: {str(e)}")
                
                elif key in ['parsingEngine', 'translator_convert_engine', 'translator_mineru_model_version', 
                             'translator_formula_ocr', 'translator_table_ocr', 'translator_skip_translate',
                             'translator_pdf_split_enabled', 'translator_pdf_split_max_pages',
                             'translator_pdf_split_max_workers', 'translator_request_retry_count']:
                    # Handle parsing_engine fields (accept both camelCase from frontend and snake_case)
                    parsing_engine_updates = {}
                    if key in ('parsingEngine', 'translator_convert_engine'):
                        parsing_engine_updates['convert_engine'] = value
                    elif key == 'translator_mineru_model_version':
                        parsing_engine_updates['mineru_model_version'] = value
                    elif key == 'translator_formula_ocr':
                        parsing_engine_updates['formula_ocr'] = value
                    elif key == 'translator_table_ocr':
                        parsing_engine_updates['table_ocr'] = value
                    elif key == 'translator_skip_translate':
                        parsing_engine_updates['skip_translate'] = value
                    elif key == 'translator_pdf_split_enabled':
                        parsing_engine_updates['pdf_split_enabled'] = value
                    elif key == 'translator_pdf_split_max_pages':
                        parsing_engine_updates['pdf_split_max_pages'] = value
                    elif key == 'translator_pdf_split_max_workers':
                        parsing_engine_updates['pdf_split_max_workers'] = value
                    elif key == 'translator_request_retry_count':
                        parsing_engine_updates['request_retry_count'] = value
                    
                    if parsing_engine_updates:
                        # Update parsing engine via update_from_dict
                        global_config.update_from_dict({'parsing_engine': parsing_engine_updates})
                
                elif key == 'ai_platforms':
                    # Update platforms.json (new structure)
                    try:
                        if not isinstance(value, dict):
                            raise ValueError("ai_platforms must be a dictionary")
                        platforms_config = get_platforms_config()
                        logger.info(LogModule.AUTH, f"[AI_PLATFORMS] Incoming platforms count: {len(value.keys())}")
                        logger.debug(LogModule.AUTH, f"[AI_PLATFORMS] Incoming platforms keys: {list(value.keys())}")
                        # Merge incoming platforms into existing platforms_config
                        for p_key, p_val in value.items():
                            if not isinstance(p_val, dict):
                                logger.warning(LogModule.AUTH, f"Invalid platform config for {p_key}, skipping")
                                continue
                            from backend.config.platforms_config import AIPlatformConfig
                            try:
                                cfg = AIPlatformConfig(
                                    name=p_val.get('name', ''),
                                    url=p_val.get('url', ''),
                                    model=p_val.get('model', ''),
                                    max_tokens=p_val.get('max_tokens', 4096),
                                    temperature=float(p_val.get('temperature', 0.3)),
                                    temperature_min=float(p_val.get('temperature_min', 0.0)),
                                    temperature_max=float(p_val.get('temperature_max', 2.0)),
                                    thinking_mode_supported=p_val.get('thinking_mode_supported', False),
                                    thinking_mode=p_val.get('thinking_mode', 'disable'),
                                    recommended_tokens=p_val.get('recommended_tokens'),
                                    performance_note=p_val.get('performance_note'),
                                    platform_type=p_val.get('platform_type', 'llm'),
                                    parser_subtype=p_val.get('parser_subtype'),
                                    api_protocol=p_val.get('api_protocol', 'openai'),
                                    requires_api_key=p_val.get('requires_api_key', True),
                                    description=p_val.get('description'),
                                    token_link=p_val.get('token_link'),
                                    api_endpoints=p_val.get('api_endpoints') if p_val.get('api_endpoints') is not None else {},
                                    chunk_size=int(p_val.get('chunk_size', 3000)) if p_val.get('chunk_size') is not None else None,
                                    concurrent=int(p_val.get('concurrent', 5)) if p_val.get('concurrent') is not None else None,
                                )
                                platforms_config.update_platform_config(p_key, cfg)
                                logger.info(LogModule.AUTH, f"[AI_PLATFORMS] Updated platform '{p_key}': url={cfg.url}, model={cfg.model}")
                            except Exception as e:
                                logger.warning(LogModule.AUTH, f"Failed to parse platform {p_key}: {e}")
                        # Determine target path for extra visibility
                        try:
                            from utils.path_utils import get_config_file_path  # may not exist; fallback below
                        except Exception:
                            get_config_file_path = None
                        save_ok = save_platforms_config()
                        if not save_ok:
                            raise HTTPException(status_code=500, detail="Failed to save platforms.json")
                        try:
                            if get_config_file_path:
                                target_path = get_config_file_path("platforms.json")
                                logger.info(LogModule.AUTH, f"[AI_PLATFORMS] platforms.json saved to: {target_path}")
                        except Exception:
                            pass
                        logger.info(LogModule.AUTH, f"[AI_PLATFORMS] Updated platforms saved successfully. Keys: {list(value.keys())}")
                    except HTTPException:
                        raise
                    except Exception as e:
                        logger.error(LogModule.AUTH, f"Failed to update ai_platforms: {e}", exc_info=True)
                        raise HTTPException(status_code=400, detail=f"Invalid ai_platforms payload: {str(e)}")

                elif key == 'exclusion_defaults':
                    # Update exclusion default settings in system.json
                    if isinstance(value, dict):
                        global_config.update_from_dict({'exclusion_defaults': value})
                        # Clear system config cache so new values take effect immediately
                        from backend.config.system_config import clear_system_config_cache
                        clear_system_config_cache()
                        logger.info(LogModule.AUTH, f"Exclusion defaults updated: {value}")
                    else:
                        logger.warning(LogModule.AUTH, f"Invalid exclusion_defaults value (expected dict): {value}")

                elif key in ['default_language', 'smart_glossary_matching_enabled', 'auth_required']:
                    # Update via update_from_dict for these fields
                    global_config.update_from_dict({key: value})
                elif hasattr(global_config, key):
                    # Direct attribute update (for properties)
                    # Note: UnifiedConfig uses properties, so we update via update_from_dict
                    global_config.update_from_dict({key: value})
                
                else:
                    logger.warning(LogModule.AUTH, f"Unknown global setting key: {key}, skipping")
            
            # Save global configuration
            if not save_unified_config():
                raise HTTPException(status_code=500, detail="Failed to save global config")
            
            logger.info(LogModule.AUTH, f"Global settings updated by {user.username}: {list(changes.keys())}")
            
        elif config_type == 'user':
            # Update user configuration
            profile_manager = get_user_profile_manager()
            
            # Special handling: Some user settings need to be synced to global app_config.json
            # This ensures backend can read the latest values from app_config.json
            app_config_needs_save = False
            
            for key, value in changes.items():
                # Map frontend keys to backend keys
                backend_key = key
                if key in ['chunkSize', 'translationChunkSize']:
                    backend_key = 'chunk_size'
                elif key == 'translationConcurrent':
                    backend_key = 'concurrent'
                elif key == 'translationTimeout':
                    backend_key = 'timeout'
                
                # Update user profile
                if not profile_manager.update_user_setting(user.username, key, value):
                    raise HTTPException(status_code=500, detail=f"Failed to save user setting: {key}")
                
                # Sync certain settings to app_config.json for backend consistency
                # These settings are user preferences but need to be in global config for backend to read
                if backend_key in ['chunk_size', 'concurrent', 'timeout', 'retry']:
                    try:
                        from config import get_app_config, save_app_config
                        app_config = get_app_config()
                        
                        # Map to app_config field names
                        if backend_key == 'chunk_size':
                            app_config.translator_chunk_token_size = int(value) if value else 8000
                            logger.info(LogModule.AUTH, f"[SETTINGS] Synced chunk_size={value} to app_config.translator_chunk_token_size")
                        elif backend_key == 'concurrent':
                            app_config.translator_concurrent = int(value) if value else 10
                            logger.info(LogModule.AUTH, f"[SETTINGS] Synced concurrent={value} to app_config.translator_concurrent")
                        elif backend_key == 'connect_timeout':
                            app_config.translator_connect_timeout = int(value) if value else 15
                            logger.info(LogModule.AUTH, f"[SETTINGS] Synced connect_timeout={value} to app_config.translator_connect_timeout")
                        elif backend_key == 'timeout':
                            app_config.translator_timeout = int(value) if value else 30
                            logger.info(LogModule.AUTH, f"[SETTINGS] Synced timeout={value} to app_config.translator_timeout")
                        elif backend_key == 'retry':
                            app_config.translator_retry = int(value) if value else 5
                            logger.info(LogModule.AUTH, f"[SETTINGS] Synced retry={value} to app_config.translator_retry")
                        
                        app_config_needs_save = True
                    except Exception as e:
                        logger.warning(LogModule.AUTH, f"[SETTINGS] Failed to sync {backend_key} to app_config.json: {e}")
                        # Continue even if sync fails (user profile is still saved)
                    
                    # Also sync chunk_size and concurrent to the default platform config (per-platform settings)
                    if backend_key in ['chunk_size', 'concurrent']:
                        try:
                            platforms_config = get_platforms_config()
                            default_platform = platforms_config.default_platform
                            if default_platform:
                                platform_cfg = platforms_config.get_platform_config(default_platform)
                                if platform_cfg:
                                    if backend_key == 'chunk_size':
                                        platform_cfg.chunk_size = int(value) if value else 3000
                                        logger.info(LogModule.AUTH, f"[SETTINGS] Synced chunk_size={value} to platform '{default_platform}' config")
                                    elif backend_key == 'concurrent':
                                        platform_cfg.concurrent = int(value) if value else 5
                                        logger.info(LogModule.AUTH, f"[SETTINGS] Synced concurrent={value} to platform '{default_platform}' config")
                                    if save_platforms_config():
                                        logger.info(LogModule.AUTH, f"[SETTINGS] Saved platforms.json after syncing {backend_key}")
                        except Exception as e:
                            logger.warning(LogModule.AUTH, f"[SETTINGS] Failed to sync {backend_key} to platforms.json: {e}")
            
            # Save app_config.json if any settings were synced
            if app_config_needs_save:
                try:
                    from config import save_app_config
                    if save_app_config():
                        logger.info(LogModule.AUTH, f"[SETTINGS] Successfully synced user settings to app_config.json")
                    else:
                        logger.warning(LogModule.AUTH, f"[SETTINGS] Failed to save app_config.json after syncing user settings")
                except Exception as e:
                    logger.warning(LogModule.AUTH, f"[SETTINGS] Exception while saving app_config.json: {e}")
            
            logger.info(LogModule.AUTH, f"User settings updated by {user.username}: {list(changes.keys())}")
            
        elif config_type == 'sensitive':
            # Update sensitive configuration: admin, or guest (e.g. desktop/setup wizard without login)
            if not user.is_admin() and user.username != "guest":
                raise HTTPException(status_code=403, detail="Admin required for sensitive settings")
            
            secrets_manager = get_secrets_manager()
            for key, value in changes.items():
                if key == 'platform_api_keys' or key == 'api_keys':
                    # Handle API keys (dictionary)
                    if not isinstance(value, dict):
                        raise HTTPException(status_code=400, detail="platform_api_keys must be a dictionary")
                    
                    # Update each platform key
                    for platform, key_data in value.items():
                        if isinstance(key_data, dict):
                            # New format: {key: "...", configured: true}
                            api_key = key_data.get('key', '')
                            configured = key_data.get('configured')
                            if not secrets_manager.update_platform_api_key(platform, api_key, configured):
                                raise HTTPException(status_code=500, detail=f"Failed to save API key for platform {platform}")
                        elif isinstance(key_data, str):
                            # Old format: direct string
                            if not secrets_manager.update_platform_api_key(platform, key_data):
                                raise HTTPException(status_code=500, detail=f"Failed to save API key for platform {platform}")
                        else:
                            logger.warning(LogModule.AUTH, f"Invalid API key format for platform {platform}, skipping")
                
                elif key == 'mineru_token' or key == 'translator_mineru_token':
                    # Handle MinerU token
                    if isinstance(value, dict):
                        # New format: {key: "...", configured: true}
                        token = value.get('key', '')
                        configured = value.get('configured')
                        if not secrets_manager.update_mineru_token(token, configured):
                            raise HTTPException(status_code=500, detail="Failed to save MinerU token")
                    elif isinstance(value, str):
                        # Old format: direct string
                        if not secrets_manager.update_mineru_token(value):
                            raise HTTPException(status_code=500, detail="Failed to save MinerU token")
                    else:
                        raise HTTPException(status_code=400, detail="mineru_token must be a string or dict with 'key' field")
                
                elif key == 'mineru_local_token':
                    # Handle MinerU Local token
                    if isinstance(value, dict):
                        # New format: {key: "...", configured: true}
                        token = value.get('key', '')
                        configured = value.get('configured')
                        if not secrets_manager.update_mineru_local_token(token, configured):
                            raise HTTPException(status_code=500, detail="Failed to save MinerU Local token")
                    elif isinstance(value, str):
                        # Direct string format
                        if not secrets_manager.update_mineru_local_token(value):
                            raise HTTPException(status_code=500, detail="Failed to save MinerU Local token")
                    else:
                        raise HTTPException(status_code=400, detail="mineru_local_token must be a string or dict with 'key' field")
                
                else:
                    logger.warning(LogModule.AUTH, f"Unknown sensitive key: {key}, skipping")
            
            logger.info(LogModule.AUTH, f"Sensitive settings updated by {user.username}: {list(changes.keys())}")
        
        return {
            "success": True,
            "message": f"{config_type} settings updated successfully",
            "updated_keys": list(changes.keys())
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(LogModule.AUTH, f"Failed to update settings batch: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to update settings: {str(e)}")


@auth_router.post("/app-config/setting")
async def update_single_setting(
    request: Request,
    user: User = Depends(get_current_user)
):
    """Update single setting item"""
    try:
        data = await request.json()
        key = data.get('key')
        value = data.get('value')
        
        if not key:
            raise HTTPException(status_code=400, detail="Setting key is required")
        
        from .user_profile import get_user_profile_manager
        from backend.config.config_loader import get_unified_config, save_unified_config
        from backend.config.secrets_manager import get_secrets_manager
        
        profile_manager = get_user_profile_manager()
        global_config = get_unified_config()

        # Define sensitive configuration keys (only administrators can modify, saved to secrets.json)
        sensitive_config_keys = [
            'translator_mineru_token',
            'mineru_local_token',
            'platform_api_keys'
        ]
        
        # Define global configuration keys (only administrators can modify)
        global_config_keys = [
            'translator_convert_engine', 'translator_mineru_model_version',
            'translator_formula_ocr', 'translator_table_ocr', 'translator_skip_translate',
            'platform_urls', 'platform_models', 'active_task_ids',
            # AI Platforms default selection
            'ai_platforms_default_platform',
            # LDAP configuration keys
            'ldap_enabled', 'ldap_protocol', 'ldap_host', 'ldap_port', 'ldap_bind_dn_template',
            'ldap_base_dn', 'ldap_user_filter', 'ldap_tls_cacertfile', 'ldap_tls_verify',
            # Exclusion defaults
            'exclusion_defaults',
        ]

        # Define user configuration keys (all users can modify)
        user_config_keys = [
            'ui_language', 'translator_last_workflow', 'translator_auto_workflow_enabled',
            'translator_txt_insert_mode', 'translator_txt_separator',
            'translator_xlsx_insert_mode', 'translator_xlsx_separator', 'translator_xlsx_translate_regions',
            'translator_docx_insert_mode', 'translator_docx_separator',
            'translator_srt_insert_mode', 'translator_srt_separator',
            'translator_epub_insert_mode', 'translator_epub_separator',
            'translator_html_insert_mode', 'translator_html_separator',
            'translator_json_paths', 'translator_target_language', 'translator_custom_language',
            'translator_custom_prompt', 'translator_thinking_mode', 'theme',
            'translator_platform_type', 'translator_temperature', 'temperature', 'translator_max_tokens', 'translator_top_p',
            'translator_frequency_penalty', 'translator_presence_penalty',
            'chunk_size', 'chunkSize', 'translationChunkSize', 'concurrent', 'translationConcurrent',
            'timeout', 'translationTimeout', 'retry',
            'glossary_generate_enable', 'glossary_agent_config_choice', 'glossary_agent_thinking_mode',
            'glossary_agent_platform_type', 'glossary_agent_temperature', 'glossary_agent_max_tokens', 'glossary_agent_top_p',
            'glossary_agent_frequency_penalty', 'glossary_agent_presence_penalty', 'glossary_agent_to_lang',
            'glossary_agent_chunk_size', 'glossary_agent_concurrent',
            # User dimension model override dictionary keys
            'translator_platform_models', 'glossary_agent_platform_models'
        ]
        
        # Permission check
        if key in sensitive_config_keys:
            # Sensitive configuration, only administrators can modify
            if not user.is_admin():
                logger.warning(LogModule.AUTH, f"LDAP user {_mask_username(user.username)} attempted to modify sensitive configuration: {key}")
                raise HTTPException(status_code=403, detail="Access denied: Only admin can modify sensitive settings")
            # Default password can only be changed by super administrator
        elif key in global_config_keys:
            # Global configuration, only administrators can modify
            if not user.is_admin():
                logger.warning(LogModule.AUTH, f"LDAP user {_mask_username(user.username)} attempted to modify global configuration: {key}")
                raise HTTPException(status_code=403, detail="Access denied: Only admin can modify global settings")
        elif key in user_config_keys:
            # User configuration, all users can modify
            pass
        else:
            # Unknown configuration key
            logger.warning(LogModule.AUTH, f"User {_mask_username(user.username)} attempted to modify unknown configuration: {key}")
            raise HTTPException(status_code=400, detail=f"Unknown setting key: {key}")
        
        # Update based on configuration type
        if key in sensitive_config_keys:
            # Update sensitive configuration (save to secrets.json)
            secrets_manager = get_secrets_manager()
            
            if key == 'translator_mineru_token':
                print(f"=== DEBUG: API endpoint calling update_mineru_token with value: {value[:20]}... ===")
                logger.info(LogModule.AUTH, f"=== DEBUG: API endpoint calling update_mineru_token ===")
                if secrets_manager.update_mineru_token(value):
                    logger.info(LogModule.AUTH, f"MinerU token updated by user {_mask_username(user.username)}")
                    return {"success": True, "message": "MinerU token updated successfully"}
                else:
                    raise HTTPException(status_code=500, detail="Failed to save MinerU token")
            
            elif key == 'mineru_local_token':
                if secrets_manager.update_mineru_local_token(value):
                    logger.info(LogModule.AUTH, f"MinerU Local token updated by user {_mask_username(user.username)}")
                    return {"success": True, "message": "MinerU Local token updated successfully"}
                else:
                    raise HTTPException(status_code=500, detail="Failed to save MinerU Local token")
            
            elif key == 'platform_api_keys':
                # Handle platform API keys dictionary
                if isinstance(value, dict):
                    updated_any = False
                    for platform, api_key in value.items():
                        # Compatibility: value might be {platform: str} or {platform: {key, configured}}
                        configured_flag = None
                        if isinstance(api_key, dict):
                            configured_flag = api_key.get('configured')
                            api_key = api_key.get('key', '')
                        if api_key and str(api_key).strip():  # Only save non-empty keys
                            if secrets_manager.update_api_key(platform, str(api_key), configured_flag):
                                updated_any = True
                    # Synchronously refresh in-memory global configuration to ensure latest masked keys are visible after page refresh
                    if updated_any:
                        try:
                            from backend.config.config_loader import get_unified_config
                            global_config = get_unified_config()
                            for platform, api_key in value.items():
                                if isinstance(api_key, dict):
                                    raw_key = api_key.get('key', '')
                                else:
                                    raw_key = str(api_key) if api_key is not None else ''
                                if raw_key and raw_key.strip():
                                    # API keys are managed by secrets_manager, not unified_config
                                    # The update was already done via secrets_manager.update_platform_api_key above
                                    pass
                        except Exception as _e:
                            logger.warning(LogModule.AUTH, f"Failed to refresh in-memory global API keys: {_e}")
                    logger.info(LogModule.AUTH, f"Platform API keys updated by user {_mask_username(user.username)}")
                    return {"success": True, "message": "Platform API keys updated successfully"}
                else:
                    raise HTTPException(status_code=400, detail="Platform API keys must be a dictionary")
            
            # session_secret_key and redis_password are now managed by local.json
                if secrets_manager.update_auth_secret(key, value):
                    logger.info(LogModule.AUTH, f"Authentication sensitive configuration {key} updated by user {_mask_username(user.username)}")
                    return {"success": True, "message": f"Auth secret {key} updated successfully"}
                else:
                    raise HTTPException(status_code=500, detail=f"Failed to save auth secret {key}")
            
            elif key == 'docling_auth':
                if isinstance(value, dict):
                    if get_secrets_manager().update_docling_auth(value):
                        logger.info(LogModule.AUTH, f"Docling authentication updated by user {_mask_username(user.username)}")
                        return {"success": True, "message": "Docling auth updated successfully"}
                    else:
                        raise HTTPException(status_code=500, detail="Failed to save Docling auth")
                else:
                    raise HTTPException(status_code=400, detail="Docling auth must be a dictionary")
            else:
                raise HTTPException(status_code=400, detail=f"Unknown sensitive setting key: {key}")
        
        elif key in global_config_keys:
            # Update global configuration
            if key.startswith('platform_') and key.endswith('_model_id'):
                # Handle platform models
                platform = key.replace('translator_platform_', '').replace('_model_id', '')
                global_config.update_platform_model(platform, value)
            elif key.startswith('glossary_agent_platform_') and key.endswith('_model_id'):
                # Handle glossary platform models
                platform = key.replace('glossary_agent_platform_', '').replace('_model_id', '')
                global_config.update_glossary_platform_model(platform, value)
            elif key.startswith('ldap_'):
                # Handle LDAP configuration
                from .config import get_auth_config, save_auth_config
                auth_config = get_auth_config()
                if hasattr(auth_config, key):
                    setattr(auth_config, key, value)
                    if save_auth_config():
                        logger.info(LogModule.AUTH, f"LDAP setting {key} updated by user {_mask_username(user.username)}")
                        return {"success": True, "message": f"LDAP setting {key} updated successfully"}
                    else:
                        raise HTTPException(status_code=500, detail="Failed to save LDAP configuration")
                else:
                    raise HTTPException(status_code=400, detail=f"Unknown LDAP setting key: {key}")
            elif key == 'exclusion_defaults':
                # Update exclusion default settings
                if isinstance(value, dict):
                    global_config.update_from_dict({'exclusion_defaults': value})
                    from backend.config.system_config import clear_system_config_cache
                    clear_system_config_cache()
                    logger.info(LogModule.AUTH, f"Exclusion defaults updated via single setting: {value}")
                else:
                    raise HTTPException(status_code=400, detail="exclusion_defaults must be a dictionary")
            else:
                # Handle regular global configuration items
                if hasattr(global_config, key):
                    setattr(global_config, key, value)
                elif key == 'ai_platforms_default_platform':
                    try:
                        # Validate platform value
                        if not isinstance(value, str) or not value.strip():
                            raise ValueError("default_platform must be a non-empty string")
                        
                        # Save to new platforms.json structure only
                        from backend.config.platforms_config import get_platforms_config, save_platforms_config
                        platforms_config = get_platforms_config()
                        platforms_config.default_platform = value.strip()
                        if not save_platforms_config():
                            raise HTTPException(status_code=500, detail="Failed to save default_platform to platforms.json")
                        
                        logger.info(LogModule.AUTH, f"Default platform updated to: {value.strip()}")
                    except HTTPException:
                        raise
                    except Exception as e:
                        logger.error(LogModule.AUTH, f"Failed to update default_platform: {e}", exc_info=True)
                        raise HTTPException(status_code=400, detail=f"Invalid value for ai_platforms_default_platform: {str(e)}")
                elif key in ['parsingEngine', 'translator_convert_engine', 'translator_mineru_model_version', 'translator_formula_ocr', 'translator_table_ocr', 'translator_skip_translate']:
                    # Handle fields in parsing_engine (accept both camelCase from frontend and snake_case)
                    if key in ('parsingEngine', 'translator_convert_engine'):
                        global_config.parsing_engine.convert_engine = value
                    elif key == 'translator_mineru_model_version':
                        global_config.parsing_engine.mineru_model_version = value
                    elif key == 'translator_formula_ocr':
                        global_config.parsing_engine.formula_ocr = value
                    elif key == 'translator_table_ocr':
                        global_config.parsing_engine.table_ocr = value
                    elif key == 'translator_skip_translate':
                        global_config.parsing_engine.skip_translate = value
                else:
                    raise HTTPException(status_code=400, detail=f"Unknown global setting key: {key}")
            
            # Save global configuration
            if save_unified_config():
                logger.info(LogModule.AUTH, f"Global setting {key} updated by user {_mask_username(user.username)}")
                return {"success": True, "message": f"Global setting {key} updated successfully"}
            else:
                raise HTTPException(status_code=500, detail="Failed to save global configuration")
        
        else:
            # Update user configuration (including user-dimension model keys)
            if profile_manager.update_user_setting(user.username, key, value):
                logger.info(LogModule.AUTH, f"User setting {key} updated by user {_mask_username(user.username)}")
                return {"success": True, "message": f"User setting {key} updated successfully"}
            else:
                raise HTTPException(status_code=500, detail="Failed to save user configuration")
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(LogModule.AUTH, f"Failed to update setting: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to update setting: {str(e)}")


# === LDAP Configuration Dedicated Read/Write Interface (Unified Entry Point) ===
@auth_router.get("/ldap-config")
async def get_ldap_config_api(user: User = Depends(get_current_user)):
    """Read LDAP-related configuration (readable after login; sensitive information not returned)"""
    config = get_auth_config()
    return {
        "ldap_enabled": config.ldap_enabled,
        "ldap_protocol": config.ldap_protocol,
        "ldap_host": config.ldap_host,
        "ldap_port": config.ldap_port,
        "ldap_bind_dn_template": config.ldap_bind_dn_template,
        "ldap_base_dn": config.ldap_base_dn,
        "ldap_user_filter": config.ldap_user_filter,
        "ldap_tls_cacertfile": config.ldap_tls_cacertfile,
        "ldap_tls_verify": config.ldap_tls_verify,
        "ldap_admin_group_enabled": config.ldap_admin_group_enabled,
        "ldap_admin_group": config.ldap_admin_group,
        "ldap_glossary_group_enabled": getattr(config, 'ldap_glossary_group_enabled', False),
        "ldap_glossary_group": getattr(config, 'ldap_glossary_group', ''),
        "ldap_group_base_dn": config.ldap_group_base_dn,
    }


@auth_router.post("/ldap-config")
async def update_ldap_config_api(request: Request, user: User = Depends(get_current_user)):
    """Unified update of LDAP-related configuration (requires administrator or management group permissions)."""
    if not user.is_admin():
        raise HTTPException(status_code=403, detail="Access denied: Admin privileges required")

    try:
        data = await request.json()

        # Only handle new key names

        # Only extract LDAP-related fields
        allowed = {
            'ldap_enabled', 'ldap_protocol', 'ldap_host', 'ldap_port', 'ldap_bind_dn_template', 'ldap_base_dn',
            'ldap_user_filter', 'ldap_tls_cacertfile', 'ldap_tls_verify', 'ldap_admin_group_enabled', 'ldap_admin_group',
            'ldap_glossary_group_enabled', 'ldap_glossary_group', 'ldap_group_base_dn'
        }
        update_payload = {k: v for k, v in data.items() if k in allowed}

        # Type processing
        if 'ldap_port' in update_payload:
            try:
                update_payload['ldap_port'] = int(update_payload['ldap_port'])
            except Exception:
                pass
        for b in ['ldap_enabled', 'ldap_tls_verify', 'ldap_admin_group_enabled', 'ldap_glossary_group_enabled']:
            if b in update_payload and isinstance(update_payload[b], str):
                update_payload[b] = update_payload[b].lower() in ("true", "1", "yes", "on")

        # Check if trying to enable LDAP without test validation
        if update_payload.get('ldap_enabled', False):
            # Check if this is a test validation request
            test_validated = data.get('ldap_test_validated', False)
            if not test_validated:
                return JSONResponse(
                    status_code=400, 
                    content={
                        "ok": False, 
                        "message": "LDAP test must be performed and passed before enabling LDAP. Please test the connection first."
                    }
                )

        # Update and save
        from .config import get_auth_config as _get_auth_cfg, save_auth_config as _save_auth_cfg
        auth_cfg = _get_auth_cfg()
        logger.info(LogModule.AUTH, f"[LDAP-API] Normalized update fields: {update_payload}")
        auth_cfg.update_from_dict(update_payload)
        saved = _save_auth_cfg()
        # Synchronously update in-memory global configuration in this module to avoid requiring restart
        try:
            local_cfg = get_auth_config()
            local_cfg.update_from_dict(update_payload)
            logger.info(LogModule.AUTH, "[LDAP-API] Successfully synchronized in-memory configuration")
        except Exception:
            pass
        if saved:
            logger.info(LogModule.AUTH, f"LDAP configuration updated by user {_mask_username(user.username)}")
            # Synchronously refresh in-memory instance in this module to avoid reading old values after page refresh
            try:
                global _auth_config
                if _auth_config is not None:
                    _auth_config.update_from_dict(update_payload)
                    logger.info(LogModule.AUTH, "[LDAP-API] Successfully synchronized _auth_config in module")
            except Exception as _e:
                logger.warning(LogModule.AUTH, f"[LDAP-API] Failed to synchronize in-memory configuration: {_e}")
            return {"success": True, "message": "LDAP configuration updated"}
        else:
            raise HTTPException(status_code=500, detail="Failed to save LDAP configuration")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(LogModule.AUTH, f"Failed to update LDAP configuration: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to update LDAP configuration: {str(e)}")


# === Message Configuration Dedicated Read/Write Interface ===
@auth_router.get("/message-config")
async def get_message_config_api():
    """Read message-related configuration (public interface, no authentication required)"""
    from .config import AuthConfig
    config = AuthConfig.get_config()
    return {
        "login_banner": config.login_banner,
        "usage_message": config.usage_message,
    }


@auth_router.get("/app-config/public")
async def get_public_app_config_api():
    """Read public application configuration (public interface, no authentication required)"""
    from .config import AuthConfig
    config = AuthConfig.get_config()
    return {
        "app_name": "Owlangs",
        "login_banner": config.login_banner,
        "usage_message": config.usage_message,
        "auth_enabled": True,
        "features": {
            "translation": True,
            "anonymization": True,
            "collaboration": True
        }
    }


@auth_router.post("/message-config")
async def update_message_config_api(request: Request, user: User = Depends(get_current_user)):
    """Update message-related configuration (requires administrator privileges)"""
    if not user.is_admin():
        raise HTTPException(status_code=403, detail="Access denied: Admin privileges required")

    try:
        data = await request.json()

        # Only extract message-related fields
        allowed = {'login_banner', 'usage_message'}
        update_payload = {k: v for k, v in data.items() if k in allowed}

        # Update and save
        from .config import AuthConfig
        auth_cfg = AuthConfig.get_config()
        logger.info(LogModule.AUTH, f"[Message-API] Update fields: {update_payload}")
        auth_cfg.update_from_dict(update_payload)
        saved = auth_cfg.save_to_file()
        
        # Synchronously update in-memory global configuration in this module
        try:
            local_cfg = get_auth_config()
            local_cfg.update_from_dict(update_payload)
            logger.info(LogModule.AUTH, "[Message-API] Successfully synchronized in-memory configuration")
        except Exception:
            pass
            
        if saved:
            logger.info(LogModule.AUTH, f"Message configuration updated by user {_mask_username(user.username)}")
            return {"success": True, "message": "Message configuration updated"}
        else:
            raise HTTPException(status_code=500, detail="Failed to save message configuration")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(LogModule.AUTH, f"Failed to update message configuration: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to update message configuration: {str(e)}")

# Compatibility routes (without /auth prefix)
@auth_compat_router.get("/login")
async def login_page_compat(request: Request, next_url: Optional[str] = None):
    """Compatibility login page (without /auth prefix)"""
    return await login_page(request, next_url)


@auth_compat_router.post("/login")
async def login_compat(
    request: Request,
    response: Response,
    username: str = Form(...),
    password: str = Form(...),
    next_url: Optional[str] = Form(None)
):
    """Compatibility login handling (without /auth prefix)"""
    return await login(request, response, username, password, next_url)


@auth_compat_router.get("/logout")
async def logout_get_compat(request: Request, response: Response):
    """Compatibility logout (without /auth prefix)"""
    return await logout_get(request, response)


@auth_compat_router.get("/user", response_model=UserInfo)
async def get_user_info_compat(request: Request):
    """Compatibility get user info (without /auth prefix)"""
    return await get_user_info(request)


@auth_compat_router.get("/user/permissions")
async def get_user_permissions_compat(
    user: User = Depends(get_current_user)
):
    """Compatibility get user permissions (without /auth prefix)"""
    return await get_user_permissions(user)


@auth_compat_router.get("/app-config")
async def get_app_config_compat(
    user: User = Depends(get_current_user)
):
    """Compatibility get app config (without /auth prefix)"""
    return await get_app_config_api(user)


@auth_compat_router.post("/app-config/setting")
async def update_single_setting_compat(request: Request, user: User = Depends(get_current_user)):
    """Compatibility update single setting (without /auth prefix)"""
    return await update_single_setting(request, user)


@auth_router.get("/ai-platform-status")
async def get_ai_platform_status(user: User = Depends(get_current_user)):
    """Return persisted AI platform test status (single source of truth for frontend)."""
    from backend.config.ai_platform_status import get_status
    return get_status()


@auth_router.post("/test-all-platforms")
async def test_all_ai_platforms(
    user: User = Depends(get_current_user)
):
    """
    Test all configured AI platforms and return results.
    Uses the same logic as the hourly scheduler.
    """
    try:
        from .ai_platform_scheduler import run_one_round_ai_platform_tests
        
        logger.info(LogModule.AUTH, "[TEST_ALL_PLATFORMS] Starting batch test of all platforms")
        await run_one_round_ai_platform_tests()
        
        # Return current status after testing
        from backend.config.ai_platform_status import get_status
        status = get_status()
        
        return {
            "success": True,
            "message": "All configured platforms tested",
            "status": status
        }
    except Exception as e:
        logger.error(LogModule.AUTH, f"[TEST_ALL_PLATFORMS] Batch test failed: {e}", exc_info=True)
        return {"success": False, "error": f"Batch test failed: {str(e)}"}


@auth_router.post("/test-ai-platform")
async def test_ai_platform(
    request: Request,
    user: User = Depends(get_current_user)
):
    """Test AI platform connection (delegates to service)."""
    try:
        data = await request.json()
        platform_type = data.get('platform_type') or data.get('platform')
        
        if not platform_type:
            raise HTTPException(status_code=400, detail="Missing required parameter: platform_type")

        from backend.config.secrets_manager import get_secrets_manager
        from backend.config.config_loader import get_unified_config, clear_unified_config_cache
        
        # Clear config cache to ensure latest platform configuration is loaded
        # This is important for api_key_optional and other platform settings
        clear_unified_config_cache()
        
        secrets_manager = get_secrets_manager()
        global_config = get_unified_config()

        # Initial log to confirm endpoint trigger
        try:
            tmp = await request.json()
            platform_type_dbg = tmp.get('platform_type') if isinstance(tmp, dict) else None
            logger.info(LogModule.AUTH, f"[TEST_AI_PLATFORM] Request received, platform_type={platform_type_dbg}")
            print(f"[TEST_AI_PLATFORM] Request received, platform_type={platform_type_dbg}")
        except Exception:
            logger.info(LogModule.AUTH, "[TEST_AI_PLATFORM] Request received (no JSON body)")
            print("[TEST_AI_PLATFORM] Request received (no JSON body)")

        # Get API key (from request or secrets)
        api_key = data.get('api_key')
        if not api_key:
            if platform_type == "mineru":
                api_key = secrets_manager.get_mineru_token()
            elif platform_type == "mineru_local":
                api_key = secrets_manager.get_mineru_local_token()
            else:
                api_keys = secrets_manager.get_api_keys()
                api_key = api_keys.get(platform_type)
        # Log masked api key (first 10 chars, rest as *)
        try:
            def _mask_key(k):
                if not isinstance(k, str):
                    k_str = str(k) if k is not None else ""
                else:
                    k_str = k
                n = len(k_str)
                if n <= 0:
                    return ""
                head = k_str[:10]
                stars = "*" * max(0, n - 10)
                return head + stars
            masked = _mask_key(api_key)
            logger.info(LogModule.AUTH, f"[TEST_AI_PLATFORM] Platform={platform_type}, api_key(masked)={masked} (len={len(str(api_key) if api_key is not None else 0)})")
            print(f"[TEST_AI_PLATFORM] Platform={platform_type}, api_key(masked)={masked}")
        except Exception:
            pass
        
        # Get platform configuration to check if API key is optional and extract base_url/model_name
        base_url = data.get('base_url')
        model_name = data.get('model_name')
        requires_api_key = True
        
        # DEBUG: Log raw values from request
        logger.info(LogModule.AUTH, f"[TEST_AI_PLATFORM DEBUG] Raw request data: base_url='{base_url}', model_name='{model_name}'")
        print(f"[TEST_AI_PLATFORM DEBUG] Raw request data: base_url='{base_url}', model_name='{model_name}'")
        
        # Get platform configuration
        # Force reload to ensure we have the latest config
        clear_unified_config_cache()
        global_config = get_unified_config()
        ai_platforms = global_config.ai_platforms
        platform_config = ai_platforms.get(platform_type)
        
        # DEBUG logging
        logger.info(LogModule.AUTH, f"[TEST_AI_PLATFORM DEBUG] platform_config type={type(platform_config)}")
        print(f"[TEST_AI_PLATFORM DEBUG] platform_config type={type(platform_config)}")
        if platform_config:
            if isinstance(platform_config, dict):
                logger.info(LogModule.AUTH, f"[TEST_AI_PLATFORM DEBUG] platform_config keys={list(platform_config.keys())}")
                print(f"[TEST_AI_PLATFORM DEBUG] platform_config keys={list(platform_config.keys())}")
                logger.info(LogModule.AUTH, f"[TEST_AI_PLATFORM DEBUG] requires_api_key from dict={platform_config.get('requires_api_key')}")
                print(f"[TEST_AI_PLATFORM DEBUG] requires_api_key from dict={platform_config.get('requires_api_key')}")
            else:
                logger.info(LogModule.AUTH, f"[TEST_AI_PLATFORM DEBUG] platform_config has requires_api_key={hasattr(platform_config, 'requires_api_key')}")
                print(f"[TEST_AI_PLATFORM DEBUG] platform_config has requires_api_key={hasattr(platform_config, 'requires_api_key')}")
                if hasattr(platform_config, 'requires_api_key'):
                    logger.info(LogModule.AUTH, f"[TEST_AI_PLATFORM DEBUG] requires_api_key value={getattr(platform_config, 'requires_api_key')}")
                    print(f"[TEST_AI_PLATFORM DEBUG] requires_api_key value={getattr(platform_config, 'requires_api_key')}")
        
        if platform_config:
            # platform_config may be an object or a dict; handle both
            if not base_url:
                try:
                    base_url = getattr(platform_config, 'url') if hasattr(platform_config, 'url') else platform_config.get('url')
                except Exception:
                    base_url = None
            if not model_name:
                try:
                    model_name = getattr(platform_config, 'model') if hasattr(platform_config, 'model') else platform_config.get('model')
                except Exception:
                    model_name = None
            # Check if API key is required for this platform
            try:
                requires_api_key = getattr(platform_config, 'requires_api_key') if hasattr(platform_config, 'requires_api_key') else platform_config.get('requires_api_key', True)
            except Exception:
                requires_api_key = True
        
        logger.info(LogModule.AUTH, f"[TEST_AI_PLATFORM DEBUG] final requires_api_key={requires_api_key}, base_url={base_url}, model_name={model_name}")
        print(f"[TEST_AI_PLATFORM DEBUG] final requires_api_key={requires_api_key}, base_url={base_url}, model_name={model_name}")
        
        # API key is required unless the platform config explicitly sets requires_api_key: false
        if not api_key and requires_api_key:
            raise HTTPException(status_code=400, detail=f"No API key found for platform: {platform_type}")
        
        # If still missing, use defaults or raise error
        if not base_url or (not model_name and platform_type not in ("volcengine_ark", "doubao", "ark")):
            # For MinerU (cloud and local), use special handling
            if platform_type in ("mineru", "mineru_local"):
                # MinerU has special test endpoint, delegate to it
                from .mineru_service import test_mineru_connectivity
                from backend.config.ai_platform_status import update_platform_status
                result = await test_mineru_connectivity(api_key, base_url=base_url, platform_key=platform_type)
                update_platform_status(
                    platform_type,
                    result.get("success", False),
                    result.get("error"),
                )
                return result
            else:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Missing required parameters for platform {platform_type}: base_url{' and model_name' if platform_type not in ('volcengine_ark','doubao','ark') else ''} must be provided or available in platform configuration"
                )

        # Delegate to service
        from .ai_platform_service import test_ai_platform_connectivity
        from backend.config.platforms_config import get_platforms_config, save_platforms_config, clear_platforms_config_cache
        
        logger.info(LogModule.AUTH, f"[TEST_AI_PLATFORM] Using base_url={base_url}, model_name={model_name}")
        print(f"[TEST_AI_PLATFORM] Using base_url={base_url}, model_name={model_name}")
        safe_model = model_name or ""
        
        # Detect max_tokens if requested (default: True)
        detect_max_tokens = data.get('detect_max_tokens', True)
        result = await test_ai_platform_connectivity(platform_type, base_url, safe_model, api_key, detect_max_tokens=detect_max_tokens)
        
        # If max_tokens was detected and test was successful, update platform configuration
        if result.get('success') and 'max_tokens' in result:
            detected_max_tokens = result['max_tokens']
            # Validate: reject suspiciously low values (< 1024) to prevent configuration errors
            if detected_max_tokens and detected_max_tokens < 1024:
                logger.warning(
                    LogModule.SYSTEM,
                    f"[TEST_AI_PLATFORM] Detected suspiciously low max_tokens={detected_max_tokens} (< 1024) "
                    f"for platform '{platform_type}'. This may be a detection error. Skipping auto-update."
                )
            else:
                try:
                    platforms_config = get_platforms_config()
                    platform_config = platforms_config.get_platform_config(platform_type)
                    
                    if platform_config:
                        old_max_tokens = platform_config.max_tokens
                        if old_max_tokens != detected_max_tokens:
                            platform_config.max_tokens = detected_max_tokens
                            if save_platforms_config():
                                clear_platforms_config_cache()
                                logger.info(LogModule.AUTH, f"[TEST_AI_PLATFORM] Updated max_tokens for platform '{platform_type}': {old_max_tokens} -> {detected_max_tokens}")
                                result['message'] = result.get('message', '') + f" (updated max_tokens: {old_max_tokens} -> {detected_max_tokens})"
                            else:
                                logger.warning(LogModule.AUTH, f"[TEST_AI_PLATFORM] Failed to save max_tokens update for platform '{platform_type}'")
                        else:
                            logger.debug(LogModule.AUTH, f"[TEST_AI_PLATFORM] max_tokens for platform '{platform_type}' unchanged: {detected_max_tokens}")
                    else:
                        logger.warning(LogModule.AUTH, f"[TEST_AI_PLATFORM] Platform config not found for '{platform_type}', cannot update max_tokens")
                except Exception as e:
                    logger.error(LogModule.AUTH, f"[TEST_AI_PLATFORM] Failed to update max_tokens for platform '{platform_type}': {e}", exc_info=True)
        
        # Persist test result so frontend and Quick Settings use backend as single source of truth
        from backend.config.ai_platform_status import update_platform_status
        update_platform_status(
            platform_type,
            result.get("success", False),
            result.get("error") or result.get("message"),
        )
        # Ensure frontend gets a message when test failed (use error if message missing)
        if not result.get("success") and not result.get("message") and result.get("error"):
            result["message"] = result["error"]
        try:
            logger.info(
                LogModule.AUTH,
                f"[TEST_AI_PLATFORM] Result: success={result.get('success')} error={result.get('error')!r} message={result.get('message')!r}",
            )
        except Exception:
            pass
        return result

    except HTTPException:
        raise
    except Exception as e:
        logger.error(LogModule.AUTH, f"AI platform test failed: {e}", exc_info=True)
        return {"success": False, "error": f"Test failed: {str(e)}"}


@auth_router.post("/ai-platform/list-models")
async def list_ai_platform_models(
    request: Request,
    user: User = Depends(get_current_user)
):
    """List available models for an AI platform."""
    try:
        data = await request.json()
        platform_type = data.get('platform_type') or data.get('platform')
        base_url = data.get('base_url')
        api_key = data.get('api_key')

        if not platform_type:
            raise HTTPException(status_code=400, detail="Missing required parameter: platform_type")
        if not base_url:
            raise HTTPException(status_code=400, detail="Missing required parameter: base_url")
        
        # Check if API key is required based on platform config
        # Force reload to ensure we have the latest config
        from backend.config import clear_unified_config_cache, get_unified_config
        clear_unified_config_cache()
        global_config = get_unified_config()
        ai_platforms = global_config.ai_platforms
        platform_config = ai_platforms.get(platform_type)
        api_protocol = data.get("api_protocol") or data.get("apiProtocol")
        if not api_protocol and platform_config:
            if isinstance(platform_config, dict):
                api_protocol = platform_config.get("api_protocol")
            else:
                api_protocol = getattr(platform_config, "api_protocol", None)

        logger.info(
            LogModule.AUTH,
            f"[LIST_MODELS DEBUG] Received request: platform_type='{platform_type}', "
            f"api_protocol='{api_protocol}', base_url='{base_url}'",
        )
        print(
            f"[LIST_MODELS DEBUG] Received request: platform_type='{platform_type}', "
            f"api_protocol='{api_protocol}', base_url='{base_url}'"
        )

        requires_api_key = True
        if platform_config:
            if isinstance(platform_config, dict):
                requires_api_key = platform_config.get('requires_api_key', True)
            else:
                requires_api_key = getattr(platform_config, 'requires_api_key', True)
        
        if not api_key and requires_api_key:
            # Try to get from secrets for non-local platforms
            from backend.config.secrets_manager import get_secrets_manager
            secrets_manager = get_secrets_manager()
            if platform_type == "mineru":
                api_key = secrets_manager.get_mineru_token()
            else:
                api_keys = secrets_manager.get_api_keys()
                api_key = api_keys.get(platform_type)
        
        if not api_key and requires_api_key:
            raise HTTPException(status_code=400, detail=f"No API key found for platform: {platform_type}")
        
        # Delegate to service
        from .ai_platform_service import list_platform_models
        models = await list_platform_models(platform_type, base_url, api_key, api_protocol=api_protocol)
        return {"success": True, "models": models}
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(LogModule.AUTH, f"List AI platform models failed: {e}", exc_info=True)
        return {"success": False, "error": f"Failed to list models: {str(e)}", "models": []}


@auth_router.post("/mineru/test-connection")
async def test_mineru_connection(
    request: Request,
    user: User = Depends(get_current_user)
):
    """Test MinerU connection"""
    try:
        # Get MinerU token (from request or secrets)
        data = {}
        try:
            content_type = request.headers.get("content-type", "")
            if "application/json" in content_type:
                data = await request.json()
        except Exception:
            # If no body or invalid JSON, use empty dict
            pass
        
        api_key = data.get('api_key')
        
        from backend.config.secrets_manager import get_secrets_manager
        secrets_manager = get_secrets_manager()
        
        # Use API key from request if provided, otherwise get from secrets
        if api_key:
            mineru_token = api_key
        else:
            mineru_token = secrets_manager.get_mineru_token()
        
        if not mineru_token:
            result = {"success": False, "message": "MinerU API Key not configured"}
            # Persist status so frontend can reflect MinerU availability
            try:
                from backend.config.ai_platform_status import update_platform_status
                update_platform_status("mineru", result.get("success", False), result.get("message"))
            except Exception as e:
                logger.error(LogModule.AUTH, f"[MINERU_TEST] Failed to update platform status for missing token: {e}", exc_info=True)
            return result
        # Log masked mineru token
        try:
            def _mask_key(k):
                if not isinstance(k, str):
                    k_str = str(k) if k is not None else ""
                else:
                    k_str = k
                n = len(k_str)
                if n <= 0:
                    return ""
                head = k_str[:10]
                stars = "*" * max(0, n - 10)
                return head + stars
            logger.info(LogModule.AUTH, f"[TEST_MINERU] api_key(masked)={_mask_key(mineru_token)} (len={len(str(mineru_token) if mineru_token is not None else 0)})")
        except Exception:
            pass
        
        # Delegate to service module
        from .mineru_service import test_mineru_connectivity
        from backend.config.ai_platform_status import update_platform_status
        result = await test_mineru_connectivity(mineru_token, platform_key="mineru")
        update_platform_status(
            "mineru",
            result.get("success", False),
            result.get("error") or result.get("message"),
        )
        return result

    except HTTPException:
        raise
    except Exception as e:
        logger.error(LogModule.AUTH, f"MinerU connection test failed: {e}", exc_info=True)
        return {"success": False, "message": f"Connection test failed: {str(e)}"}

# Note: MinerU test connection is now integrated into /test-ai-platform endpoint


@auth_router.get("/certificate-list")
async def get_certificate_list(user: User = Depends(get_current_user)):
    """Get list of certificates in certs directory"""
    try:
        import os
        import subprocess
        from pathlib import Path
        from datetime import datetime
        
        certs_dir = Path("certs")
        certificates = []
        
        if certs_dir.exists():
            for file_path in certs_dir.iterdir():
                if file_path.is_file() and file_path.suffix in ['.crt', '.key', '.pem']:
                    stat = file_path.stat()
                    file_type = 'cert' if file_path.suffix in ['.crt', '.pem'] else 'key'
                    
                    cert_info = {
                        'name': file_path.name,
                        'type': file_type,
                        'size': f"{stat.st_size} bytes",
                        'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                    }
                    
                    # For certificate files, try to get validity period
                    if file_type == 'cert':
                        try:
                            # Use openssl to get certificate validity
                            result = subprocess.run([
                                'openssl', 'x509', '-in', str(file_path), '-noout', '-dates'
                            ], capture_output=True, text=True, check=True)
                            
                            # Parse the output to extract dates
                            output = result.stdout
                            not_before = None
                            not_after = None
                            
                            for line in output.split('\n'):
                                if line.startswith('notBefore='):
                                    not_before = line.split('=', 1)[1].strip()
                                elif line.startswith('notAfter='):
                                    not_after = line.split('=', 1)[1].strip()
                            
                            if not_after:
                                # Parse the date and check if it's expired
                                try:
                                    # OpenSSL date format: Oct  2 19:42:59 2025 GMT
                                    from datetime import datetime
                                    parsed_date = datetime.strptime(not_after, '%b %d %H:%M:%S %Y %Z')
                                    now = datetime.now()
                                    
                                    cert_info['valid_until'] = parsed_date.strftime('%Y-%m-%d %H:%M:%S')
                                    cert_info['is_expired'] = parsed_date < now
                                    
                                    # Calculate days until expiration
                                    days_left = (parsed_date - now).days
                                    if days_left < 0:
                                        cert_info['days_left'] = f"Expired {abs(days_left)} days ago"
                                    elif days_left == 0:
                                        cert_info['days_left'] = "Expires today"
                                    else:
                                        cert_info['days_left'] = f"{days_left} days left"
                                        
                                except ValueError:
                                    cert_info['valid_until'] = not_after
                                    cert_info['days_left'] = "Unknown"
                                    cert_info['is_expired'] = False
                                    
                        except (subprocess.CalledProcessError, FileNotFoundError):
                            # If openssl is not available or fails, skip validity info
                            pass
                    
                    certificates.append(cert_info)
        
        return {"certificates": certificates}
        
    except Exception as e:
        logger.error(LogModule.AUTH, f"Failed to get certificate list: {e}")
        raise HTTPException(status_code=500, detail="Failed to get certificate list")


@auth_router.post("/generate-certificate")
async def generate_certificate(request: Request, user: User = Depends(get_current_user)):
    """Generate temporary SSL certificate"""
    if not user.is_admin():
        raise HTTPException(status_code=403, detail="Access denied: Admin privileges required")
    
    try:
        data = await request.json()
        platform = data.get('platform', 'linux')
        
        import os
        import subprocess
        from pathlib import Path
        
        # Create certs directory if it doesn't exist
        certs_dir = Path("certs")
        certs_dir.mkdir(exist_ok=True)
        
        # Change to certs directory
        os.chdir(certs_dir)
        
        try:
            # Generate private key
            subprocess.run([
                'openssl', 'genrsa', '-out', 'server.key', '2048'
            ], check=True, capture_output=True)
            
            # Generate CSR
            subprocess.run([
                'openssl', 'req', '-new', '-key', 'server.key', '-out', 'server.csr',
                '-subj', '/C=US/ST=State/L=City/O=Organization/CN=localhost'
            ], check=True, capture_output=True)
            
            # Generate self-signed certificate
            subprocess.run([
                'openssl', 'x509', '-req', '-days', '365', '-in', 'server.csr',
                '-signkey', 'server.key', '-out', 'server.crt'
            ], check=True, capture_output=True)
            
            # Set proper permissions (Linux/Unix)
            if platform == 'linux':
                os.chmod('server.key', 0o600)
                os.chmod('server.crt', 0o644)
            
            # Clean up CSR file
            if Path('server.csr').exists():
                Path('server.csr').unlink()
            
            logger.info(LogModule.AUTH, f"Certificate generated successfully by user {_mask_username(user.username)}")
            return {"success": True, "message": "Certificate generated successfully"}
            
        except subprocess.CalledProcessError as e:
            logger.error(LogModule.AUTH, f"OpenSSL command failed: {e}")
            return {"success": False, "message": f"Certificate generation failed: {e.stderr.decode() if e.stderr else str(e)}"}
        
        finally:
            # Change back to original directory
            os.chdir('..')
            
    except Exception as e:
        logger.error(LogModule.AUTH, f"Certificate generation error: {e}")
        raise HTTPException(status_code=500, detail=f"Certificate generation failed: {str(e)}")


def _require_admin(user: Optional[User]):
    if not user or not (user.is_admin() or user.is_super_admin()):
        raise HTTPException(status_code=403, detail="Admin permission required")


@auth_router.get("/local-users", response_class=JSONResponse)
async def list_local_users(current_user: Optional[User] = Depends(get_current_user)):
    """List local users (admin only)."""
    _require_admin(current_user)
    store = get_local_user_store()
    users = store.list_users()
    # Hide password hashes and convert to list format
    safe_users = []
    for user_data in users:
        if isinstance(user_data, dict):
            # If it's already a dict with user info
            safe_user = {k: v for k, v in user_data.items() if k != "password_hash"}
            safe_users.append(safe_user)
        else:
            # If it's a username string, get the full user data
            user = store.get_user(user_data)
            if user:
                safe_user = {
                    "username": user.username,
                    "role": user.role.value,
                    "display_name": user.display_name,
                    "email": user.email,
                    "created_at": getattr(user, 'created_at', None),
                    "last_login": getattr(user, 'last_login', None),
                    "is_active": getattr(user, 'is_active', True)
                }
                safe_users.append(safe_user)
    return {"users": safe_users}


@auth_router.post("/local-users", response_class=JSONResponse)
async def create_local_user(
    request: Request,
    current_user: Optional[User] = Depends(get_current_user)
):
    """Create local user (admin only)."""
    _require_admin(current_user)
    payload = await request.json()
    username = payload.get("username")
    password = payload.get("password")
    role = payload.get("role", "user")
    display_name = payload.get("display_name")
    email = payload.get("email")
    if not username or not password:
        raise HTTPException(status_code=400, detail="username and password are required")
    try:
        store = get_local_user_store()
        store.create_user(username, password, LocalUserRole(role), display_name, email)
        return {"ok": True}
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))


@auth_router.put("/local-users/{username}", response_class=JSONResponse)
async def update_local_user(username: str, request: Request, current_user: Optional[User] = Depends(get_current_user)):
    """Update local user basic info (admin only)."""
    _require_admin(current_user)
    payload = await request.json()
    role = payload.get("role")
    display_name = payload.get("display_name")
    email = payload.get("email")
    # Super admin guard: do not allow modifying super admin basic info
    auth_cfg = get_auth_config()
    if username == auth_cfg.default_username:
        raise HTTPException(status_code=403, detail="Cannot modify super admin")
    store = get_local_user_store()
    try:
        store.update_user(
            username,
            role=LocalUserRole(role) if role is not None else None,
            display_name=display_name,
            email=email
        )
        return {"ok": True}
    except ValueError as ve:
        raise HTTPException(status_code=404, detail=str(ve))


@auth_router.post("/local-users/{username}/reset-password", response_class=JSONResponse)
async def reset_local_user_password(username: str, request: Request, current_user: Optional[User] = Depends(get_current_user)):
    """Reset local user password (admin only, cannot reset super admin)."""
    _require_admin(current_user)
    payload = await request.json()
    new_password = payload.get("password")
    if not new_password:
        raise HTTPException(status_code=400, detail="password is required")
    # Super admin guard
    auth_cfg = get_auth_config()
    if username == auth_cfg.default_username:
        raise HTTPException(status_code=403, detail="Cannot reset super admin password here")
    store = get_local_user_store()
    try:
        store.reset_password(username, new_password)
        return {"ok": True}
    except ValueError as ve:
        raise HTTPException(status_code=404, detail=str(ve))


@auth_router.delete("/local-users/{username}", response_class=JSONResponse)
async def delete_local_user(username: str, current_user: Optional[User] = Depends(get_current_user)):
    """Delete local user (admin only, cannot delete super admin)."""
    _require_admin(current_user)
    auth_cfg = get_auth_config()
    if username == auth_cfg.default_username:
        raise HTTPException(status_code=403, detail="Cannot delete super admin")
    store = get_local_user_store()
    ok = store.delete_user(username)
    return {"ok": ok}


# Self-service change password for local users
@auth_router.post("/local-users/me/change-password", response_class=JSONResponse)
async def change_own_local_password(request: Request, current_user: Optional[User] = Depends(get_current_user)):
    """Allow authenticated users to change their own password by providing current password.

    Rules:
    - Admin users: verify against secrets config and update secrets config
    - Local users: verify against local user store and update local user store
    - LDAP users (no entry in local store) are not allowed here.
    """
    if current_user is None:
        raise HTTPException(status_code=401, detail="not authenticated")

    payload = await request.json()
    current_password = payload.get("current_password")
    new_password = payload.get("new_password")

    if not current_password or not new_password:
        from .i18n_utils import get_password_message
        raise HTTPException(status_code=400, detail=get_password_message("changePasswordRequired"))

    # Validate new password strength
    from .password_manager import password_manager
    from .i18n_utils import get_password_message
    is_valid, error_msg = password_manager.validate_password_strength(new_password)
    if not is_valid:
        raise HTTPException(status_code=400, detail=f"{get_password_message('changePasswordTooWeak')}: {error_msg}")

    # Use unified user storage for all users
    from .unified_user_store import get_unified_user_store
    unified_store = get_unified_user_store()
    unified_user = unified_store.get_user(current_user.username)
    
    if unified_user and unified_user.is_active:
        # User exists in unified storage
        logger.info(LogModule.AUTH, f"Unified user changing password: {_mask_username(current_user.username)}")
        
        # Verify current password
        if not unified_store.verify_credentials(current_user.username, current_password):
            raise HTTPException(status_code=403, detail=get_password_message("changePasswordCurrentIncorrect"))
        
        # Update password in unified storage
        if unified_store.update_password(current_user.username, new_password):
            logger.info(LogModule.AUTH, f"Unified user password updated successfully: {_mask_username(current_user.username)}")
            return {"ok": True}
        else:
            logger.warning(LogModule.AUTH,"Failed to update password in unified storage")
            raise HTTPException(status_code=500, detail=get_password_message("changePasswordUpdateFailed"))
    
    else:
        # User not found in unified storage
        raise HTTPException(status_code=403, detail=get_password_message("changePasswordNotLocalUser"))
